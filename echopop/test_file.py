from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
import numpy as np
import pandas as pd
from pathlib import Path 
from echopop.survey import Survey

survey = Survey(init_config_path = "C:/Users/Brandyn/Documents/GitHub/echopop/config_files/initialization_config.yml",
                survey_year_config_path = "C:/Users/Brandyn/Documents/GitHub/echopop/config_files/survey_year_2019_config.yml")
survey.load_survey_data()
survey.load_acoustic_data()
survey.transect_analysis()
survey.kriging_analysis(variogram_parameters=dict(model=["exponential", "bessel"], n_lags=30), variable="biomass_density")

# SIM
# LENGTH_BINS = np.linspace(2, 80, 40)
# AGE_BINS = np.linspace(1, 22, 22)
# SEX_OPTS = ["male", "female", "unsexed"]

# def cartesian_coord(*arrays):
#     grid = np.meshgrid(*arrays)
#     coord_list = [entry.ravel() for entry in grid]
#     points = np.vstack(coord_list).T
#     return points

# data = pd.DataFrame(cartesian_coord(LENGTH_BINS, AGE_BINS, SEX_OPTS), 
#                     columns=["length", "age", "sex"])
# data["biomass"] = 0.0
# data.loc[(data.sex == "male") &
#          (
#              ((data.age == "1.0") & (data.length == "24.0"))
#              | ((data.age == "2.0") & (data.length.isin(["26.0", "28.0"])))
#              | ((data.age == "3.0") & (data.length.isin(["28.0", "30.0", "32.0"])))
#              | ((data.age == "3.0") & (data.length.isin(["30.0", "32.0", "34.0", "36.0"])))
#          ),
#          "biomass"
#     ] = [10, 20, 20, 40, 50, 40, 30, 10]
# data.loc[(data.sex == "female") &
#          (
#              ((data.age == "1.0") & (data.length.isin(["22.0", "24.0"]))
#              | ((data.age == "2.0") & (data.length.isin(["24.0", "26.0"])))
#              | ((data.age == "3.0") & (data.length.isin(["26.0", "30.0", "32.0", "34.0"])))
#              | ((data.age == "3.0") & (data.length.isin(["34.0", "36.0", "38.0", "40.0"])))
#              | ((data.age == "4.0") & (data.length.isin(["40.0", "42.0"]))))
#          ),
#          "biomass"
#     ] = [10, 10, 20, 20, 40, 50, 60, 70, 60, 60, 50, 30, 10]
# data.loc[(data.sex == "unsexed") &
#          (
#              ((data.age == "1.0") & (data.length.isin(["22.0", "24.0"]))
#              | ((data.age == "2.0") & (data.length.isin(["26.0"]))))
#          ),
#          "biomass"
#     ] = [10, 20, 10]
# data_all = (
#     data
#     .groupby(["age", "length"])
#     .sum()
#     .reset_index()
#     .assign(sex="all")
# )

# data = pd.concat([data, data_all], ignore_index=True)
# data = data.astype({"age": float, "length": float, "biomass": float, "sex": str})

###################
FILENAME = "C:/Users/Brandyn/Documents/custom_format.xlsx"
sexes = ["all", "male", "female"]
data = survey.analysis["transect"]["biology"]["population"]["tables"]["biomass"]["aged_biomass_df"]
###################
# df1 = survey.analysis["transect"]["biology"]["proportions"]["weight"]["aged_weight_proportions_df"]
# self = survey
data = lambda v: v.data.analysis
title = "{SEX}"
filename = lambda v: Path(v.save_directory) / "transect_aged_biomass_dataframe.xlsx"

def pivot_aged_weight_proportions(
    weight_distribution: pd.DataFrame,
    age1_exclusion: bool,
    stratum_name: str,
) -> Dict[str, pd.DataFrame]:

    # Stack the age distributions
    weight_stk = weight_distribution.sum().reset_index(name="weight")

    # Concatenate a copy representing "all" fish
    weight_stk_all = (
        weight_stk.groupby([stratum_name, "age_bin"], observed=False)["weight"].sum().reset_index()
    )
    # ---- Add to the original stacked dataframe
    weight_stk_full = pd.concat([weight_stk, weight_stk_all.assign(sex="all")], ignore_index=True)

    # Convert age bins into number ages
    weight_stk_full["age"] = weight_stk_full["age_bin"].apply(lambda x: x.mid).astype(float)

    # Pivot the table
    weight_pvt = weight_stk_full.pivot_table(columns=["sex", stratum_name], 
                                             index=["age"], 
                                             values="weight")

    # Zero out the age-1 category, if they are being excluded
    if age1_exclusion:
        weight_pvt.loc[1] = 0.0

    # Convert to proportions
    weight_prop_pvt = (weight_pvt / weight_pvt.sum()).fillna(0.0)

    # Create table dictionary
    tables = {sex: weight_prop_pvt.loc[:, sex].T for sex in ["all", "male", "female"]}

    # Return the output
    return tables

def pivot_dataframe_data(
    transect_df: pd.DataFrame,
    tables_dict: Dict[str, pd.DataFrame],
    sex: str,
    stratum_name: str,
    contrasts: List[str] = ["transect_num", "longitude", "latitude"]
):

    # Filter the correct sex for the aged weight proportions
    aged_weight_proportions_df = tables_dict[sex]

    # Assign the correct biomass column name (based on sex)
    if sex != "all":
        biomass_name = f"biomass_{sex}" if f"biomass_{sex}" in transect_df.columns else "biomass"
    else: 
        biomass_name = "biomass"

    # Filter the transect dataframe
    transect_data = (
        transect_df.filter(contrasts + [stratum_name, biomass_name])
    )

    # Set the index to the strata
    transect_data.set_index([stratum_name], inplace=True)

    # Merge the transect and weight proportions dataframes
    aged_transect_data = pd.merge(transect_data, 
                                  aged_weight_proportions_df, 
                                  left_index=True, 
                                  right_index=True)

    # Propagate the biomass column over the age column proportions
    aged_transect_data.loc[:, 1.0:] = (
        aged_transect_data.loc[:, 1.0:].mul(aged_transect_data[biomass_name], axis=0)
    )

    # Sort the final dataframe and return
    return aged_transect_data.sort_values(contrasts).reset_index()

def write_aged_dataframe_report(
    tables_dict: Dict[str, pd.DataFrame],
    filepath: Path,
):

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Initialize
    for sex in ["all", "female", "male"]:
        
        # Add/overwrite the sheet
        ws = format_file_sheet(sex.capitalize(), wb)
        
        # Set up superheader representing 'sex'
        if sex != "all":
            ws.append([sex.capitalize()])
        else:        
            ws.append([f"{sex.capitalize()} (Male + Female)"])

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()

    # Remove header formatting
    pd.io.formats.excel.ExcelFormatter.header_style = None
    
    # Open up a writer engine
    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        
        # Iterate through all sexes
        for sex in ["all", "female", "male"]:

            # Subset the data dictionary for the particular sheet
            sheet_data = tables_dict[sex]

            # Add actual column names (i.e. age) and datatable rows
            sheet_data.to_excel(writer, sheet_name=sex.capitalize(), 
                                startrow=1, index=None)
            
def repivot_table(age_length_dataframe: pd.DataFrame, 
                  variable: str,
                  length_dataframe: Optional[pd.DataFrame] = None):

    # Restack the data table
    data_stk = (
        age_length_dataframe.stack(future_stack=True).sum(axis=1).reset_index(name=variable)
    )

    # Concatenate unaged data, if needed
    if length_dataframe is not None:
        # ---- Restack
        unaged_data_stk = length_dataframe.sum(axis=1).reset_index(name=variable)
        # ---- Get the right-most age-bin and add 1
        shifted_bin = data_stk["age_bin"].max() + 1
        # ---- Get the mid-point for later renaming
        shifted_bin_mid = shifted_bin.mid
        # ---- Add age column
        unaged_data_stk["age_bin"] = shifted_bin        
        # ---- Concatenate with the aged data
        data_stk = pd.concat([data_stk, unaged_data_stk], ignore_index=True)
    else: 
        shifted_bin_mid = None

    # Convert the length and age columns from intervals into numerics
    # ---- Length
    data_stk["length"] = data_stk["length_bin"].apply(lambda x: x.mid).astype(float)
    # ---- Age
    data_stk["age"] = data_stk["age_bin"].apply(lambda x: x.mid).astype(float)

    # Pivot
    data_pvt = data_stk.pivot_table(
        columns=["age"],
        index=["length"], 
        values=variable, 
        margins_name="Subtotal",
        margins=True, 
        aggfunc="sum"
    )    

    # Drop the extra axes
    data_proc = data_pvt.rename_axis(columns=None, index=None)

    # Rename the unaged placeholder to "Unaged" and return    
    return data_proc.rename(columns = {shifted_bin_mid: "Unaged"})

def format_table_headers(dataframe: pd.DataFrame, 
                         col_name: str = "Age",
                         row_name: str = "Length (cm)") -> List[str]:

    # Get the shape of the table
    DTSHAPE = dataframe.shape

    # Write data with custom headers and gaps
    # ---- Get midway column
    midway_col_idx = int(np.ceil((DTSHAPE[1] - 1) / 2))

    # Create headers
    headers = [row_name] + [""] * (midway_col_idx - 1) + [col_name]
    # ---- Complete the headers
    return headers + [""] * (DTSHAPE[1] - len(headers))

def format_file_sheet(
    sheetname: str,
    workbook: Workbook,
) -> Worksheet: 

    # Check whether the sheetname already exists, and remove if needed
    if sheetname in workbook.sheetnames:
        workbook.remove(workbook[sheetname])

    # Create the new sheet 
    _ = workbook.create_sheet(sheetname)

    # Activate and return the new sheet
    return workbook[sheetname]

def append_datatable_rows(
    worksheet: Worksheet,
    dataframe: pd.DataFrame
) -> None:

    # Add column names (i.e. age)
    worksheet.append([""] + dataframe.columns.values.tolist())

    # Get the dataset row indices
    ROW_INDICES = dataframe.index.values

    # Populate the file with all rows except for the last margins row
    # ---- Iterate along the length bins
    for row in ROW_INDICES[:-1]:
        # ---- Get the row 
        row_data = dataframe.loc[row, :]
        # ---- Concatenate row name with the rest of the data
        row_data = [row_data.name] + list(row_data)
        # ---- Append
        worksheet.append(row_data)

    # Append the final row
    # ---- Only apply if 'Subtotal' margin included
    if "Subtotal" in ROW_INDICES:
        worksheet.append(
            ["Subtotal"] + dataframe.loc["Subtotal"].values[:-1].tolist()    
        )

def append_table_aggregates(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    sex: Literal["all", "female", "male"], 
    tables_dict: Dict[str, pd.DataFrame],
) -> None:

    # Get the aged sums
    if "Unaged" in dataframe.columns:
        sum_age = dataframe.loc["Subtotal"].values[:-2]
        # ---- Unaged sum
        sum_unaged = dataframe.loc["Subtotal", "Unaged"]
    else:
        sum_age = dataframe.loc["Subtotal"].values[:-1]   
        # ---- Unaged sum
        sum_unaged = 0.0

    # Total number across all fish
    total_aged = sum_age.sum() + sum_unaged

    # Calculate the age-1 proportion
    # ---- Proportion
    age1_proportion = dataframe.loc["Subtotal", 1] / sum_age.sum()
    # ---- Compute age-2+ values
    age2_aged = total_aged*(1-age1_proportion)

    # Add next row 
    # ---- Age 1+
    age_1_row = [
        "Total (age1+)",
        total_aged
    ]
    # ---- Add the "Over Age" value
    age_1_row = age_1_row + [
        "",
        "Over Age:",
        sum_age.sum(),
        ""
    ]
    # ---- Add sexed 
    if sex == "all":
        age_1_row = age_1_row + [
            "Male+Female:",
            tables_dict["female"].iloc[:-1, :-1].sum().sum() 
            + tables_dict["male"].iloc[:-1, :-1].sum().sum()
        ]    
    # ---- Append
    worksheet.append(age_1_row)

    # Add next row
    # ---- Age-2+
    age_all_row = [
        "Total (age2+)",
        age2_aged
    ]
    # ---- Add sexed 
    
    if sex == "all":
        age_all_row = (
            age_all_row 
            + [""] * 4
            + [
                "Male+Female:",
                tables_dict["female"].iloc[:-1, 1:-1].sum().sum() 
                + tables_dict["male"].iloc[:-1, 1:-1].sum().sum() 
            ]
        )
    # ---- Append
    worksheet.append(age_all_row)

def append_sheet_label(
    worksheet: Worksheet,
    title: str,
    sex: str,
) -> None:
    
    # Add empty row
    worksheet.append([""])

    # Add sheet label
    sheet_label = title.replace("{SEX}", sex.capitalize())
    # ---- Append
    worksheet.append(
        [""] * 9 + [sheet_label]
    )

def initialize_workbook(
    filepath: Path
) -> Workbook:
    
    # Check if workbook already exists
    if filepath.exists():
        wb = load_workbook(filepath)
    # ---- If not, generate it and remove the default "Sheet" that is created
    else:
        wb = Workbook()
        wb.remove(wb["Sheet"])

    # Return
    return wb

def write_age_length_table_report(
    tables_dict: Dict[str, pd.DataFrame],
    table_title: str,
    filepath: Path,
) -> None:

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Iterate through all sexes
    for sex in ["all", "female", "male"]:

        # Subset the data dictionary for the particular sheet
        sheet_data = tables_dict[sex]

        # Add/overwrite the sheet
        ws = format_file_sheet(sex.capitalize(), wb)

        # Set up the headers for the file sheet
        ws.append(format_table_headers(sheet_data))

        # Add actual column names (i.e. age) and datatable rows
        append_datatable_rows(ws, sheet_data)

        # Append the data table aggregates
        append_table_aggregates(ws, sheet_data, sex, tables_dict)

        # Append the trailing sheet label along the bottom-most row
        append_sheet_label(ws, table_title, sex)

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()

####################################################################################################
from typing import Any, Callable, Dict, Literal, List, Optional, Union
from echopop import Survey
from echopop.biology import filter_species

data_dict = survey

# Filter out non-target species
length_data, specimen_data = filter_species(
    [
        data_dict.input["biology"]["length_df"],
        data_dict.input["biology"]["specimen_df"],
    ],
    data_dict.analysis["settings"]["transect"]["species_id"],
)

stratum_name = data_dict.analysis["settings"]["transect"]["stratum_name"]

# Meld the specimen and length dataframes together for downstream calculations
# aggregate_lengths = specimen_data.meld(
#     length_data, contrasts=["haul_num", "species_id", "length_bin", "sex", "group_sex"]
# )

# filepath = self.save_directory / "aged_length_haul_counts_table.xlsx"

aggregate_lengths = specimen_data.copy()
aggregate_lengths.loc[aggregate_lengths.sex == "male", :].loc[~np.isnan(aggregate_lengths.age)]
# ----
full_df = pd.concat([aggregate_lengths.loc[aggregate_lengths.group_sex == "sexed", :], 
                     aggregate_lengths.loc[aggregate_lengths.group_sex == "sexed", :].assign(sex="all")])
full_df
full_df = full_df.loc[~np.isnan(full_df.age)]
full_df[(full_df.sex == "male") & (~np.isnan(full_df.length))]
full_df["length"] = full_df["length_bin"].apply(lambda x: x.mid).astype(float)
full_pvt = full_df.pivot_table(index=["sex", "length_bin"], columns=["haul_num"], values="length", aggfunc="count", observed=False)
full_pvt.loc["male"].sum().sum()
sex = "male"
table_title = "Aged Length-Haul Counts ({SEX})"
haul_data = full_pvt.loc[sex, :]
haul_stk = haul_data.stack(future_stack=True).reset_index(name="count")
haul_stk.loc[:, "length"] = haul_stk["length_bin"].apply(lambda x: x.mid).astype(float)

haul_agg = haul_stk.pivot_table(index=["length"], 
                                columns=["haul_num"], 
                                values="count", 
                                margins=True, 
                                margins_name="Subtotal", 
                                aggfunc="sum")

# Drop the extra axes
haul_proc = haul_agg.rename_axis(columns=None, index=None)

# Create the workbook, if needed
wb = initialize_workbook(filepath)
# tables_dict = haul_agg
# Subset the data dictionary for the particular sheet
# sheet_data = tables_dict[sex]
sheet_data = haul_agg

# Add/overwrite the sheet
ws = format_file_sheet(sex.capitalize(), wb)

# Set up the headers for the file sheet
ws.append(format_table_headers(sheet_data, col_name=f"Haul Number ({sex.capitalize()})"))

# Add actual column names (i.e. age) and datatable rows
append_datatable_rows(ws, sheet_data)

# Add grand total
total_row = ["Total", sheet_data.iloc[:-1, :-1].sum().sum()]
# ---- Extend, if needed
if sex == "all": 
    total_row + [
        "",
        "Male+Female",
        full_pvt.loc["female"].sum().sum() 
        + full_pvt.loc["male"].sum().sum()
    ]
# ---- Append
ws.append(total_row)

#
ws.append([""])
ws.append([""])

#
append_sheet_label(ws, table_title, sex)

# Save the workbook
wb.save(filepath)
# ---- Close
wb.close()


reports = ["kriged_length_age_biomass", "transect_length_age_abundance", "transect_length_age_biomass"]
test_cls = FEATReports(survey, reports)
self = test_cls

class FEATReports: 
       
    def __init__(self, 
                 survey: Survey,
                 reports: List[
                     Literal[
                        "kriged_aged_biomass_mesh",
                        "kriged_biomass_mesh",
                        "kriged_length_age_abundance",
                        "kriged_length_age_biomass",
                        "transect_aged_biomass",
                        "transect_length_age_abundance",
                        "transect_length_age_biomass",    
                        "transect_population_estimates",                         
                     ]
                 ],
                 save_directory: Optional[Union[str, Path]] = None,
                 ):
    
        # Assign table-type(s)
        self.reports = reports
        
        # Assign the data
        self.data = survey

        # Save directory
        if not save_directory:
            # ---- Get the save directory configured within the `Survey` configuration
            survey_directory = Path(survey.config["data_root_dir"])
            # ---- Create the save directory
            save_directory = survey_directory / "reports"
            # ---- Create the directory, if needed
            save_directory.mkdir(parents=True, exist_ok=True)
            # ---- Assign the save directory
            self.save_directory = save_directory
        else:
            # ---- Validate existence
            if not Path(save_directory).exists():
                raise FileNotFoundError(
                    f"The reports save directory '{Path(save_directory).as_posix()}' does not "
                    f"exist!"
                )
            # ---- Assign the save directory
            self.save_directory = save_directory            
   
    def generate(
        self
    ):

        # self.reports = [
        #     "transect_length_age_abundance",
        #     "transect_length_age_biomass"
        # ]
        
        # Validate the reports
        report_methods = self.get_report_type(
            self.reports
        )
        
        #

        # k = "kriged_length_age_biomass"
        # sub = report_methods[k]
        # data = sub["data"]
        # sub["data"](self)
        # title = sub["title"]
        # filename = sub["filename"]

        # report_methods[k].get("function")(self, ** )        
        report_files = [v.get("function")(self, **v) 
                        for k, v in report_methods.items()]
        
        # Map methods to the appropriate report-type
        # ---- Join over lines
        report_str = "\n   ".join(f"-'{file}'" for file in report_files)
        # ---- Print
        print(f"The following report tables were generated:\n   {report_str}")

    @classmethod
    def get_report_type(
        cls,
        requested_reports = List[str],
    ):
        
        # Find matching methods
        mapped_report_methods = {
            k: cls.__METHOD_REFERENCE__.get(k) for k in requested_reports
            if k in cls.__METHOD_REFERENCE__
        }
        
        # Return
        return mapped_report_methods  

    # REPORT METHODS
    def kriged_aged_biomass_mesh_report(
        self,
        data: Callable[[pd.DataFrame], pd.DataFrame],
        filename: Callable[[Union[str, Path]], Path],
    ):
        
        # Get the data dictionary
        data_dict = data(self)

        # Get 'exclude_age1' argument value
        age1_exclusion = data_dict.analysis["settings"]["transect"]["exclude_age1"]

        # Get the stratum name
        stratum_name = data_dict.analysis["settings"]["transect"]["stratum_name"]

        # Get the weight distributions
        weight_distribution = (
            data_dict.analysis["transect"]["biology"]["distributions"]["weight"]\
                ["aged_length_weight_tbl"]
            .copy()
        )

        # Reformat the aged weight proportions tables
        tables = pivot_aged_weight_proportions(
            weight_distribution, age1_exclusion, stratum_name
        )

        # Get mesh dataframe
        mesh_df = data_dict.results["kriging"]["mesh_results_df"].copy()
        # ---- Filter
        mesh_df_copy = mesh_df.filter(["latitude", "longitude", stratum_name, "biomass"]).copy()

        # Combine the pivot tables with the transect tdata
        mesh_pvt_tables = {sex: pivot_dataframe_data(mesh_df_copy, tables, sex, stratum_name, 
                                                     contrasts=["longitude", "latitude"]) 
                           for sex in ["all", "female", "male"]}

        # Write the *.xlsx sheet
        write_aged_dataframe_report(mesh_pvt_tables, filename(self))

        # Return the filepath name
        return filename(self).as_posix()

    def kriged_biomass_mesh_report(
        self,
        data: Callable[[pd.DataFrame], pd.DataFrame],
        filename: Callable[[Union[str, Path]], Path],
        **kwargs,
    ):

        # Get the stratum name
        stratum_name = self.data.analysis["settings"]["transect"]["stratum_name"]

        # Get apportionment information
        apportion_df = data(self).analysis["transect"]["biology"]["proportions"]["weight"]\
            ["aged_unaged_sex_weight_proportions_df"].copy()
        # ---- Pivot
        apportion_pvt = apportion_df.pivot_table(index=["stratum_num"], columns=["sex"])
        # ---- Compute the proportions across strata
        stratum_prop = (
            apportion_pvt["weight_proportion_overall_aged"] 
            + apportion_pvt["weight_proportion_overall_unaged"]
        )
        
        # Get mesh results
        mesh_df = data(self).results["kriging"]["mesh_results_df"].copy()
        # ---- Set the index
        mesh_df.set_index(stratum_name, inplace=True)

        # Merge the mesh and apportionment dataframes
        merged_mesh = pd.merge(mesh_df, stratum_prop, left_index=True, right_index=True)
        # ---- Propagate the proportions across biomass estimates
        merged_mesh.loc[:, "female":"male"] = (
            merged_mesh.loc[:, "female":"male"].mul(merged_mesh["biomass"], axis=0)
        )
        # ---- Rename the columns
        merged_mesh.rename(columns={"female": "biomass_female", "male": "biomass_male"}, 
                           inplace=True)

        # Back-calculate the kriged standard deviation
        merged_mesh.loc[:, "kriged_sd"] = (
            merged_mesh.loc[:, "kriged_mean"] * merged_mesh.loc[:, "sample_cv"]
        )

        # Update and filter the columns that will be exported
        output_df = merged_mesh.reset_index().filter(
            ["latitude", "longitude", stratum_name, "biomass", "biomass_female", "biomass_male", 
             "sample_cv", "kriged_sd"]
        )
        
        # Save the *.xlsx sheet
        output_df.to_excel(filename(self), sheet_name="Sheet1", index=None)

        # Return the filepath name
        return filename(self).as_posix()

    def kriged_length_age_biomass_report(
        self,
        data: Callable[[pd.DataFrame], pd.DataFrame],
        title: str,
        filename: Callable[[Union[str, Path]], Path],
        **kwargs,
    ):

        # Initialize a pivot table from the dataset
        dataset = data(self).pivot_table(index=["sex", "length_bin"],
                                         columns=["age_bin"],
                                         values=["biomass_apportioned"],
                                         aggfunc="sum",
                                         observed=False)
        
        # Repivot the datasets for each sex
        tables = {sex: repivot_table(dataset.loc[sex, :] * 1e-9, "biomass") 
                  for sex in ["all", "male", "female"]}     

        # Write the *.xlsx sheet
        write_age_length_table_report(tables, title, filename(self))

        # Return the filepath name
        return filename(self).as_posix()

    def transect_aged_biomass_report(
        self,
        data: Callable[[pd.DataFrame], pd.DataFrame],
        filename: Callable[[Union[str, Path]], Path],
        **kwargs,
    ):
        
        # Get the data dictionary
        data_dict = data(self)

        # Get 'exclude_age1' argument value
        age1_exclusion = data_dict["settings"]["transect"]["exclude_age1"]

        # Get the stratum name
        stratum_name = data_dict["settings"]["transect"]["stratum_name"]

        # Get the weight distributions
        weight_distribution = (
            data_dict["transect"]["biology"]["distributions"]["weight"]["aged_length_weight_tbl"]
            .copy()
        )

        # Get transect dataframe
        transect_df = data_dict["transect"]["acoustics"]["adult_transect_df"].copy()

        # Reformat the aged weight proportions tables
        tables = pivot_aged_weight_proportions(
            weight_distribution, age1_exclusion, stratum_name
        )

        # Combine the pivot tables with the transect tdata
        transect_pvt_tables = {sex: pivot_dataframe_data(transect_df, tables, sex, stratum_name) 
                               for sex in ["all", "female", "male"]}

        # Write the *.xlsx sheet
        write_aged_dataframe_report(transect_pvt_tables, filename(self))

        # Return the filepath name
        return filename(self).as_posix()

    def transect_length_age_abundance_report(
        self,
        data: Callable[[pd.DataFrame], pd.DataFrame],
        title: str,
        filename: Callable[[Union[str, Path]], Path],
        **kwargs,
    ):
        
        # Repivot the datasets for each sex
        # ---- First get the complete aged table       
        tables = {sex: 
            repivot_table(age_length_dataframe=data(self)["aged_abundance_df"].loc[sex, :], 
                          length_dataframe=data(self)["unaged_abundance_df"].loc[sex, :],
                          variable="abundance") 
                  for sex in ["all", "male", "female"]}

        # Write the *.xlsx sheet
        write_age_length_table_report(tables, title, filename(self))

        # Return the filepath name
        return filename(self).as_posix()
    
    def transect_length_age_biomass_report(
        self,
        data: Callable[[pd.DataFrame], pd.DataFrame],
        title: str,
        filename: Callable[[Union[str, Path]], Path],
        **kwargs,
    ) -> str:  

        # Repivot the datasets for each sex
        tables = {sex: repivot_table(data(self).loc[sex, :] * 1e-9, "biomass") 
                  for sex in ["all", "male", "female"]}     

        # Write the *.xlsx sheet
        write_age_length_table_report(tables, title, filename(self))

        # Return the filepath name
        return filename(self).as_posix()

    def transect_population_results_report(
        self,
        data: Callable[[pd.DataFrame], pd.DataFrame],
        filename: Callable[[Union[str, Path]], Path],
        **kwargs,
    ):

        # Save the *.xlsx sheet
        data(self).to_excel(
            filename(self), sheet_name="Sheet1", index=None
        )

        # Return the filepath name
        return filename(self).as_posix()

    # CLASS REFERENCE DICTIONARIES
    # Reference dictionary: report methods
    __METHOD_REFERENCE__ = {
        "kriged_aged_biomass_mesh": {
            "function": kriged_aged_biomass_mesh_report,
            "data": lambda v: v.data,
            "title": None,
            "filename": lambda v: 
                Path(v.save_directory) / "kriged_aged_biomass_dataframe.xlsx",
        },
        "kriged_biomass_mesh": {
            "function": kriged_biomass_mesh_report,
            "data": lambda v: v.data.results["kriging"]["mesh_results_df"],
            "title": None,
            "filename": lambda v: Path(v.save_directory) / "kriging_mesh_biomass_dataframe.xlsx",
        },
        "kriged_length_age_biomass": {
            "function": kriged_length_age_biomass_report,
            "data": (
                lambda v:
                    v.data.results["kriging"]["tables"]["overall_apportionment_df"]
            ),
            "title": "Kriged Acoustically Weighted Biomass (mmt) ({SEX})",
            "filename": lambda v:
                Path(v.save_directory) / "kriged_length_age_biomass_table.xlsx"
        },
        "transect_aged_biomass": {
            "function": transect_aged_biomass_report,
            "data": lambda v: v.data.analysis,
            "title": None,
            "filename": lambda v: 
                Path(v.save_directory) / "transect_aged_biomass_dataframe.xlsx",
        },
        "transect_length_age_abundance": {
            "function": transect_length_age_abundance_report,
            "data": (
                lambda v: 
                    v.data.analysis["transect"]["biology"]["population"]["tables"]["abundance"]
            ),
            "title": "Transect-based Acoustically Weighted Abundance ({SEX})",
            "filename": lambda v:
                Path(v.save_directory) / "transect_length_age_abundance_table.xlsx"
        },
        "transect_length_age_biomass": {
            "function": transect_length_age_biomass_report,
            "data": lambda v: 
                v.data.analysis["transect"]["biology"]["population"]["tables"]["biomass"]\
                    ["aged_biomass_df"],
            "title": "Transect-based Acoustically Weighted Biomass (mmt) ({SEX})",
            "filename": lambda v:
                Path(v.save_directory) / "transect_length_age_biomass_table.xlsx"
        },
        "transect_population_results": {
            "function": transect_population_results_report,
            "data": lambda v:
                v.data.analysis["transect"]["acoustics"]["adult_transect_df"],
            "title": None,
            "filename": lambda v:
                Path(v.save_directory) / "transect_population_results_dataframe.xlsx"
        }
    }