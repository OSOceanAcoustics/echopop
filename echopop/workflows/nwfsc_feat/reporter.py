import os
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Union

import numpy as np
import pandas as pd
import pandas.io.formats.excel as pdif
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ... import utils
from . import apportionment

####################################################################################################
# FILE WRITING UTILITY
# --------------------------------------------------------------------------------------------------


def initialize_workbook(filepath: Path) -> Workbook:
    """
    Initialize an Excel workbook at `filepath`. If a workbook already exists at the path it will
    be removed and a fresh workbook object returned.

    Parameters
    ----------
    filepath : pathlib.Path
        Destination path for the workbook.

    Returns
    -------
    openpyxl.workbook.workbook.Workbook
        A new workbook object (the file on disk is not yet written until `.save()` is called).

    Raises
    ------
    TypeError
        If `filepath` is not a Path-like object.
    PermissionError
        If an existing file cannot be removed due to permissions.
    """

    # Check path-type
    if not isinstance(filepath, Path):
        raise TypeError("Filepath must be a `pathlib.Path`.")

    # If file exists, remove it to start fresh
    if filepath.exists():
        try:
            # Use Path.unlink for atomic removal
            filepath.unlink()
        except Exception as e:
            # Fall-back to os.remove if needed, then raise for visibility
            try:
                os.remove(str(filepath))
            except Exception:
                raise PermissionError(f"Could not remove existing file '{filepath}': {e}")

    # Initialize workbook
    wb = Workbook()

    # Remove default sheet created by `openpyxl`
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Return
    return wb


def format_file_sheet(
    sheetname: str,
    workbook: Workbook,
) -> Worksheet:
    """
    Create (or replace) and return an openpyxl Worksheet in `workbook` named `sheetname`.

    Parameters
    ----------
    sheetname : str
        Name for the worksheet to create.
    workbook : openpyxl.workbook.workbook.Workbook
        Workbook object to add/replace the sheet in.

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
        The created worksheet.

    Raises
    ------
    TypeError
        If `workbook` is not an openpyxl Workbook.
    ValueError
        If `sheetname` is empty.
    """

    # Check typing
    if not isinstance(workbook, Workbook):
        raise TypeError("Workbook must be an `openpyxl.Workbook` instance")
    if not sheetname or not isinstance(sheetname, str):
        raise ValueError("Sheetname must be a non-empty string")

    # Check whether the sheetname already exists, and remove if needed
    if sheetname in workbook.sheetnames:
        workbook.remove(workbook[sheetname])

    # Create the new sheet
    _ = workbook.create_sheet(sheetname)

    # Activate and return the new sheet
    return workbook[sheetname]


def format_table_headers(
    dataframe: pd.DataFrame, col_name: str = "Age", row_name: str = "Length (cm)"
) -> List[str]:
    """
    Generate header row for an Excel sheet representing a pivot table with a left row label and
    a top column label.

    Parameters
    ----------
    dataframe : pandas.DataFrame
        DataFrame that will be written to the sheet (used for determining width).
    col_name : str, optional
        Name to print at the far-right top (default "Age").
    row_name : str, optional
        Name to print at the far-left top (default "Length (cm)").

    Returns
    -------
    List[str]
        A single list representing the row to append as header to the worksheet.

    Raises
    ------
    TypeError
        If `dataframe` is not a pandas DataFrame.
    """

    # Check typing
    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError("Dataframe must be a `pandas.DataFrame`.")

    # Get the shape of the table
    DTSHAPE = dataframe.shape

    # Write data with custom headers and gaps
    # ---- Get midway column
    midway_col_idx = int(np.ceil((DTSHAPE[1] - 1) / 2))

    # Create headers
    headers = [row_name] + [""] * (midway_col_idx - 1) + [col_name]
    # ---- Complete the headers
    return headers + [""] * (DTSHAPE[1] - len(headers))


def append_datatable_rows(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    """
    Append rows of a pivoted data table into an openpyxl worksheet.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet to append the data into.
    dataframe : pandas.DataFrame
        Pivoted DataFrame where index represents row labels (e.g. length bins) and columns
        represent ages (plus a 'Subtotal' margin row).

    Raises
    ------
    TypeError
        If `worksheet` is not an openpyxl Worksheet or `dataframe` is not a pandas DataFrame.
    ValueError
        If the DataFrame index is empty.
    """

    # Check typing and data shape
    if not hasattr(worksheet, "append"):
        raise TypeError("Worksheet must be an `openpyxl.Worksheet`.")
    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError("dataframe must be a `pandas.DataFrame`.")
    if dataframe.index.size == 0:
        raise ValueError("Dataframe has no rows to write.")

    # Add column names (i.e. age)
    worksheet.append([0] + dataframe.columns.values.tolist())

    # Get the dataset row indices
    ROW_INDICES = dataframe.index.values

    # Populate the file with all rows except for the last margins row
    # ---- Iterate along the length bins
    for row in ROW_INDICES[:-1]:
        # ---- Get the row
        row_data = dataframe.loc[row, :]
        # ---- Concatenate row name with the rest of the data
        row_data = [row_data.name] + list(row_data)
        # ---- Append
        worksheet.append(row_data)

    # Append the final row
    # ---- Only apply if 'Subtotal' margin included
    if "Subtotal" in ROW_INDICES:
        worksheet.append(["Subtotal"] + dataframe.loc["Subtotal"].values[:-1].tolist())


def append_table_aggregates(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    sex: Literal["all", "female", "male"],
    tables_dict: Dict[str, pd.DataFrame],
) -> None:
    """
    Append summary aggregate rows (Total age1+, Total age2+) and optional combined sex totals to
    the bottom-left of the worksheet.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet to append to.
    dataframe : pandas.DataFrame
        The pivot table DataFrame that was written to the sheet.
    sex : {"all", "female", "male"}
        Sex label for which aggregates are computed.
    tables_dict : dict
        Dictionary containing other sex tables required for combined aggregates
        (expected keys: "male", "female" when `sex` == "all").

    Raises
    ------
    TypeError
        For wrong input types.
    KeyError
        If required tables are missing from `tables_dict`.
    """

    # Check typing, methods, and keys
    if not hasattr(worksheet, "append"):
        raise TypeError("Worksheet must be an `openpyxl.Worksheet`.")
    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError("Dataframe must be a `pandas.DataFrame`.")
    if sex not in ("all", "female", "male"):
        raise ValueError("Sex must be one of 'all', 'female', 'male'.")

    # Get the aged sums
    # ----
    if "Un-aged" in dataframe.columns:
        # ---- Aged sum
        sum_age = dataframe.loc["Subtotal"].values[:-2]
        # ---- Unaged sum
        sum_unaged = dataframe.loc["Subtotal", "Un-aged"]
    else:
        # ---- Aged sum
        sum_age = dataframe.loc["Subtotal"].values[:-1]
        # ---- Unaged sum
        sum_unaged = 0.0

    # Total number across all fish
    total_aged = sum_age.sum()

    # Calculate the age-1 proportion
    # ---- Proportion
    if total_aged == 0:
        # ---- Avoid divide-by-zero error
        age1_proportion = 0.0
    else:
        age1_proportion = dataframe.loc["Subtotal", 1] / total_aged
    # ---- Compute age-2+ values
    age2_aged = (total_aged + sum_unaged) * (1 - age1_proportion)

    # Add next row
    # ---- Age 1+
    age_1_row = ["Total (age1+)", total_aged + sum_unaged]
    # ---- Add the "Over Age" value
    age_1_row = age_1_row + ["", "Over Age:", sum_age.sum() + sum_unaged, ""]
    # ---- Add sexed
    if sex == "all":
        if "female" not in tables_dict or "male" not in tables_dict:
            raise KeyError("Tables must contain 'female' and 'male' keys when sex == 'all'")
        # ---- Calculate total
        sexed_total = (
            tables_dict["female"].iloc[:-1, :-1].sum().sum()
            + tables_dict["male"].iloc[:-1, :-1].sum().sum()
        )
        # ---- Append
        age_1_row = age_1_row + ["Male+Female:", sexed_total]
    # ---- Append
    worksheet.append(age_1_row)

    # Add next row
    # ---- Age-2+
    age_all_row = ["Total (age2+)", age2_aged]
    # ---- Add sexed
    if sex == "all":
        age_all_row = age_all_row + [""] * 4 + ["Male+Female:", sexed_total * (1 - age1_proportion)]
    # ---- Append
    worksheet.append(age_all_row)


def append_sheet_label(
    worksheet: Worksheet,
    title: str,
    sex: str,
) -> None:
    """
    Append a trailing label describing the sheet and sex.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet to append to.
    title : str
        Title text template. May contain "{SEX}" which will be replaced with sex.capitalize().
    sex : str
        Sex string used to format the title.

    Raises
    ------
    TypeError
        If inputs are of incorrect types.
    """

    # Check typing
    if not hasattr(worksheet, "append"):
        raise TypeError("Worksheet must be an `openpyxl.Worksheet`")
    if not isinstance(title, str) or not isinstance(sex, str):
        raise TypeError("Title and sex must be strings.")

    # Add empty row
    worksheet.append([""])

    # Add sheet label
    sheet_label = title.replace("{SEX}", sex.capitalize())
    # ---- Append
    worksheet.append([""] * 9 + [sheet_label])


####################################################################################################
# DATA WRANGLING
# --------------------------------------------------------------------------------------------------


def pivot_aged_dataframe(
    geodata: pd.DataFrame,
    propdata: pd.DataFrame,
    variable: str,
    sex: str,
    contrasts: List[str],
) -> pd.DataFrame:
    """
    Multiply population estimates by stratified proportions and return a pivot-ready DataFrame.

    Parameters
    ----------
    geodata : pandas.DataFrame
        Georeferenced data with population estimates by stratum (index aligns with propdata).
    propdata : pandas.DataFrame
        Proportion table indexed compatibly with `geodata`.
    variable : str
        Column name in `geodata` to propagate by proportions (e.g., 'biomass').
    sex : str
        Sex selection used for naming column selection within `geodata`. If sex != 'all' the code
        attempts to use f"{variable}_{sex}" if present otherwise `variable`.
    contrasts : list
        List of column names retained from `geodata` (e.g., ["latitude", "longitude"]).

    Returns
    -------
    pandas.DataFrame
        The combined DataFrame with propagated per-age (or per-weight) estimates ready for later
        formatting.

    Raises
    ------
    TypeError
        If inputs are wrong types.
    KeyError
        If required columns are missing.
    """

    # Type checking
    if not isinstance(geodata, pd.DataFrame) or not isinstance(propdata, pd.DataFrame):
        raise TypeError("Georeference Dataframe and proportions must both be `pandas.DataFrame`s.")
    if not isinstance(variable, str):
        raise TypeError("Variable must be a string.")
    if not isinstance(contrasts, list):
        raise TypeError("Contrasts must be a list of column names.")

    # Assign the correct population estimate column name (based on sex)
    if sex != "all":
        var_name = f"{variable}_{sex}" if f"{variable}_{sex}" in geodata.columns else variable
    else:
        var_name = variable

    # Check for column
    if var_name not in geodata.columns:
        raise KeyError(f"Variable '{var_name}' not found in geodata columns.")

    # Filter the transect dataframe
    geodata_subset = geodata.filter(contrasts + [var_name])

    # Reindex the proportions
    propdata_ridx = propdata.reindex(geodata.index)

    # Propragate the population estimates over their respectgive stratified proportions
    var_propagated = propdata_ridx.mul(geodata_subset[var_name], axis=0)

    # Merge
    aged_geodata = pd.concat([geodata_subset, var_propagated], axis=1)

    # Sort the final dataframe and return
    return aged_geodata


def repivot_table(
    age_length_dataframe: pd.DataFrame,
    variable: str,
    length_dataframe: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    Re-pivot a stacked age-length table into a matrix with numeric age and length indices.

    This function accepts a previously stacked age-length dataframe created by some internal
    calculations and returns a table with numeric 'length' rows and 'age' columns. If an
    unaged length table is supplied it will be appended as an 'Un-aged' column.

    Parameters
    ----------
    age_length_dataframe : pandas.DataFrame
        Input table to restack and repivot.
    variable : str
        Name for the value column after restacking.
    length_dataframe : pandas.DataFrame, optional
        Optional unaged length totals to append as an 'Un-aged' column.

    Returns
    -------
    pandas.DataFrame
        Pivoted table ready for export (with 'Subtotal' margins included).

    Raises
    ------
    TypeError
        If inputs are of incorrect types.
    """

    # Type checking
    if not isinstance(age_length_dataframe, pd.DataFrame):
        raise TypeError("'age_length_dataframe' must be a `pandas.DataFrame`.")
    if length_dataframe is not None and not isinstance(length_dataframe, pd.DataFrame):
        raise TypeError("'length_dataframe' must be a `pandas.DataFrame` or `None`.")

    # Restack the data table
    data_stk = age_length_dataframe.stack(future_stack=True)
    # ---- Sum across columns if strata are present
    if any(["stratum" in name for name in age_length_dataframe.columns.names]):
        data_stk = data_stk.sum(axis=1)
    # ---- Reset index
    data_stk = data_stk.reset_index(name=variable)

    # Concatenate unaged data, if needed
    if length_dataframe is not None:
        # ---- Restack
        unaged_data_stk = length_dataframe.sum(axis=1).reset_index(name=variable)
        # ---- Get the right-most age-bin and add 1
        shifted_bin = data_stk["age_bin"].max() + 1
        # ---- Get the mid-point for later renaming
        shifted_bin_mid = shifted_bin.mid
        # ---- Add age column
        unaged_data_stk["age_bin"] = shifted_bin
        # ---- Concatenate with the aged data
        data_stk = pd.concat([data_stk, unaged_data_stk], ignore_index=True)
    else:
        shifted_bin_mid = None

    # Convert the length and age columns from intervals into numerics
    # ---- Length
    data_stk["length"] = data_stk["length_bin"].apply(lambda x: x.mid).astype(float)
    # ---- Age
    data_stk["age"] = data_stk["age_bin"].apply(lambda x: x.mid).astype(float)

    # Pivot
    data_pvt = data_stk.pivot_table(
        columns=["age"],
        index=["length"],
        values=variable,
        margins_name="Subtotal",
        margins=True,
        aggfunc="sum",
    )

    # Drop the extra axes
    data_proc = data_pvt.rename_axis(columns=None, index=None)

    # Rename the unaged placeholder to "Unaged" and return
    if shifted_bin_mid is None:
        return data_proc
    else:
        return data_proc.rename(columns={shifted_bin_mid: "Un-aged"})


def pivot_aged_weight_proportions(
    age_weight_data: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    """
    Convert age-weight data into sex-specific proportion tables.

    Parameters
    ----------
    age_weight_data : pandas.DataFrame
        DataFrame indexed by intervals (length or age bins) with columns for sexes
        (must include 'male' and 'female' DataFrames as outermost keys).

    Returns
    -------
    dict
        Mapping of sex -> DataFrame (keys: "all", "male", "female") where each DataFrame
        is transposed and proportions sum to 1 across ages.

    Raises
    ------
    TypeError
        If input is not a DataFrame.
    """

    # Type checking
    if not isinstance(age_weight_data, pd.DataFrame):
        raise TypeError("'age_weight_data' must be a `pandas.DataFrame`.")

    # Get all column level names
    col_levels = list(age_weight_data.columns.names)

    # Move 'sex' to the front if it's not already
    if col_levels[0] != "sex":
        col_levels.remove("sex")
        col_levels = ["sex"] + col_levels
        age_weight_data.columns = age_weight_data.columns.reorder_levels(col_levels)

    # Add 'all' category
    age_weights_all = age_weight_data["male"] + age_weight_data["female"]
    # ---- Convert into a MultiIndex
    age_weights_all.columns = pd.MultiIndex.from_product([["all"], age_weights_all.columns])

    # Concatenate the sex-specific and overall data tables
    age_weight_full = pd.concat([age_weight_data, age_weights_all], axis=1)

    # Convert the interval indices to numerics
    age_weight_full.index = (
        pd.Series(age_weight_full.index).apply(lambda x: x.mid).values.to_numpy().astype(int)
    )

    # Calculate proportions
    age_weight_proportions = (age_weight_full / age_weight_full.sum()).fillna(0.0)

    # Create dictionary with (transposed) sex-specific tables
    age_weight_tables = {
        sex: age_weight_proportions.loc[:, sex].T for sex in ["all", "male", "female"]
    }

    # Return the output
    return age_weight_tables


def pivot_haul_tables(
    dataframe: pd.DataFrame,
    sex: str,
) -> pd.DataFrame:
    """
    Produce a standardized haul counts/length table for a given sex.

    Parameters
    ----------
    dataframe : pandas.DataFrame
        A full pivot table with a MultiIndex columns (sex, length_bin) or single-level 'all'.
    sex : str
        Sex selection 'male', 'female', or 'all'.

    Returns
    -------
    pandas.DataFrame
        Table oriented for writing to Excel with Subtotal row/column.

    Raises
    ------
    KeyError
        If requested sex is not present in columns.
    """

    # Type checking
    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError("Dataframe must be a `pandas.DataFrame`.")
    if sex not in ["male", "female", "all"]:
        raise ValueError("Sex must be one of 'male', 'female', 'all'.")

    # Subset the data into the specific sex
    haul_sex = dataframe.loc[:, sex].copy()

    # Convert the interval indices to numerics
    # ---- Get new columns
    new_columns = pd.Series(haul_sex.columns).apply(lambda x: x.mid).values
    # ---- Convert to `numpy.array`
    if len(new_columns) == 1:
        haul_sex.columns = new_columns
    else:
        haul_sex.columns = new_columns.to_numpy()

    # Transpose
    haul_sex = haul_sex.T

    # Add margins
    # ---- Across length bins
    haul_sex["Subtotal"] = haul_sex.sum(axis=1).astype(int)
    # ---- Across hauls
    haul_sex.loc["Subtotal"] = pd.concat(
        [haul_sex.iloc[:, :-1].sum(axis=0).astype(int), pd.Series([None], dtype="Int64")]
    )

    # Drop column and index names
    haul_sex.rename_axis(columns=None, index=None, inplace=True)

    # Return the updated table
    return haul_sex


def prepare_aged_biomass_dataframes(
    geodata_tables: Dict[str, pd.DataFrame],
):
    """
    Standardize column naming and ordering for aged biomass tables for each sex.

    This mutates the input dictionary in-place to rename columns consistently and order them.

    Parameters
    ----------
    geodata_tables : dict
        Dictionary keyed by sex ('all','male','female') with DataFrames for each.

    Raises
    ------
    KeyError
        If expected keys are missing.
    TypeError
        If invalid types are provided.
    """

    # Type checking
    if not isinstance(geodata_tables, dict):
        raise TypeError("'geodata_tables' must be a `dict`.")

    # Column name mapping for renaming
    RENAME_MAP = {
        "latitude": "Lat",
        "longitude": "Lon",
        "transect_num": "Transect",
    }

    # Column ordering
    COLUMN_ORDER = ["Transect", "Lat", "Lon", "stratum"]

    # Iterate through each table to update the formatting
    for sex in ["all", "male", "female"]:
        if sex not in geodata_tables:
            raise KeyError(f"'geodata_tables' missing required key '{sex}'.")
        # ---- Subset sex
        geodata_sex = geodata_tables[sex]
        if not isinstance(geodata_sex, pd.DataFrame):
            raise TypeError("Each entry of 'geodata_tables' must be a `pandas.DataFrame`.")
        # ---- Rename biomass column
        if sex == "all":
            geodata_sex.rename(columns={"biomass": "wgt_total"}, inplace=True)
        else:
            geodata_sex.rename(columns={f"biomass_{sex}": f"wgt_{sex}"}, inplace=True)
        # ---- Get the stratum index name before resetting
        stratum_name = geodata_sex.index.name
        # ---- Reset the index
        geodata_sex.reset_index(inplace=True)
        # ---- Rename the stratum column
        geodata_sex.rename(columns={stratum_name: "stratum"}, inplace=True)
        # ---- Rename the remaining columns
        geodata_sex.rename(columns=RENAME_MAP, inplace=True)
        # ---- Update column ordering
        sex_columns = (
            COLUMN_ORDER
            + geodata_sex.filter(regex=r"wgt+").columns.tolist()
            + geodata_sex.filter(regex=r"\d+").columns.tolist()
        )
        # ---- Update the dictionary table
        geodata_tables[sex] = geodata_sex.filter(sex_columns)


####################################################################################################
# EXCEL REPORT WRITING
# --------------------------------------------------------------------------------------------------


def write_aged_dataframe_report(
    tables_dict: Dict[str, pd.DataFrame],
    filepath: Path,
    sheetnames: Dict[str, str],
) -> None:
    """
    Write three sheets (all/male/female) of the aged dataframe report to an Excel file.

    Parameters
    ----------
    tables_dict : dict
        Dictionary containing DataFrames with keys 'all', 'male', 'female'.
    filepath : pathlib.Path
        Output xlsx file path.
    sheetnames : dict
        Mapping sex -> sheetname.

    Raises
    ------
    TypeError
        If types are incorrect.
    KeyError
        If expected keys are missing.
    """

    # Type checking
    if not isinstance(filepath, Path):
        raise TypeError("'filepath' must be a `pathlib.Path`.")
    if not isinstance(tables_dict, dict):
        raise TypeError("'tables_dict' must be a `dict`.")
    for k in ["all", "male", "female"]:
        if k not in tables_dict:
            raise KeyError(f"'tables_dict' must contain '{k}'.")
        if not isinstance(tables_dict[k], pd.DataFrame):
            raise TypeError(f"'tables_dict['{k}']' must be a `pandas.DataFrame`.")
    if not isinstance(sheetnames, dict):
        raise TypeError("Sheetnames must be a `dict`.")

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Iterate through each sex to construct the full file
    for sex in ["all", "male", "female"]:
        # ---- Add/Overwrite sheet
        ws = format_file_sheet(sheetnames[sex], wb)
        # ---- Set up the superheader
        if sex != "all":
            ws.append([sex.capitalize()])
        else:
            ws.append([f"{sex.capitalize()} (Male + Female)"])

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()

    # Remove header formatting
    pdif.ExcelFormatter.header_style = None

    # Open up a writer engine
    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:

        # Iterate through all sexes
        for sex in ["all", "male", "female"]:

            # Subset the data dictionary for the particular sheet
            sheet_data = tables_dict[sex]

            # Add actual column names (i.e. age) and datatable rows
            sheet_data.to_excel(writer, sheet_name=sheetnames[sex], startrow=1, index=None)


def prepare_age_length_tables(
    aged_full_table: pd.DataFrame,
    unaged_full_table: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    """
    Add 'Un-aged' column, margin rows and return sex-specific pivot tables for writing.

    Parameters
    ----------
    aged_full_table : pandas.DataFrame
        Concatenated aged tables by sex and 'all'.
    unaged_full_table : pandas.DataFrame
        Unaged counts for each sex (and 'all').

    Returns
    -------
    dict
        Mapping sex -> formatted DataFrame.

    Raises
    ------
    TypeError
        If inputs are not DataFrames.
    """

    # Type checking
    if not isinstance(aged_full_table, pd.DataFrame) or not isinstance(
        unaged_full_table, pd.DataFrame
    ):
        raise TypeError("Both the aged and unaged tables must be `pandas.DataFrame`s.")

    # Initialize dictionary
    sex_tables = {}

    # Iterate through each sex
    for sex in ["male", "female", "all"]:
        # Concatenate the aged and unaged tables
        # ---- Create copy of aged table
        full_table = aged_full_table.copy()[sex]
        # ---- Add "unaged" column
        full_table["Un-aged"] = unaged_full_table[sex]

        # Add margins
        # ---- Across length bins
        full_table["Subtotal"] = full_table.sum(axis=1)
        # ---- Across age bins (and unaged)
        full_table.loc["Subtotal"] = np.concatenate(
            [full_table.iloc[:, :-1].sum(axis=0).to_numpy(), np.array([np.nan])]
        )

        # Add to dictionary
        sex_tables[sex] = full_table

    # Return tables
    return sex_tables


def write_age_length_table_report(
    tables_dict: Dict[str, pd.DataFrame],
    filepath: Path,
    sheetnames: Dict[str, str],
    variable: str,
    type: str,
) -> None:
    """
    Write age-length pivot tables for each sex into an Excel workbook with header formatting and
    aggregates.

    Parameters
    ----------
    tables_dict : dict
        Mapping sex -> DataFrame (as returned by prepare_age_length_tables).
    filepath : pathlib.Path
        Output path.
    sheetnames : dict
        Mapping sex -> sheetname.
    variable : str
        Label to use in sheet footer (e.g., 'abundance' or 'biomass').
    type : str
        Descriptor for the sheet label (e.g., 'Kriged' or 'Un-kriged').

    Raises
    ------
    TypeError
        If types are invalid.
    KeyError
        If table dict missing expected keys.
    """

    # Type checking
    if not isinstance(tables_dict, dict):
        raise TypeError("'tables_dict' must be a `dict`.")
    for k in ["male", "female", "all"]:
        if k not in tables_dict:
            raise KeyError(f"'tables_dict' must include key '{k}'.")

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Iterate through all sexes
    for sex in ["male", "female", "all"]:

        # Subset the data dictionary for the particular sheet
        sheet_data = tables_dict[sex]

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetnames[sex], wb)

        # Set up the headers for the file sheet
        ws.append(format_table_headers(sheet_data))

        # Add actual column names (i.e. age) and datatable rows
        append_datatable_rows(ws, sheet_data)

        # Append the data table aggregates
        append_table_aggregates(ws, sheet_data, sex, tables_dict)

        # Append the trailing sheet label along the bottom-most row
        append_sheet_label(ws, f"{type} Acoustically Weighted {variable.capitalize()} ({sex})", sex)

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()


def write_haul_report(
    tables_dict: Dict[str, pd.DataFrame],
    sheetnames: Dict[str, Any],
    filepath: Path,
) -> None:
    """
    Write haul count tables for male/female/all to an xlsx file.

    Parameters
    ----------
    tables_dict : dict
        Mapping sex -> DataFrame produced by pivot_haul_tables.
    sheetnames : dict
        Mapping sex -> sheetname.
    filepath : pathlib.Path
        Output xlsx file.

    Raises
    ------
    TypeError, KeyError
    """

    # Type checking
    if not isinstance(filepath, Path):
        raise TypeError("'filepath' must be a `pathlib.Path`.")
    if not isinstance(tables_dict, dict):
        raise TypeError("'tables_dict' must be a `dict`.")
    for k in ["male", "female", "all"]:
        if k not in tables_dict:
            raise KeyError(f"'tables_dict' missing '{k}'.")

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Iterate through all sexes
    for sex in ["male", "female", "all"]:

        # Subset the data dictionary for the particular sheet
        sheet_data = tables_dict[sex]

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetnames[sex], wb)

        # Set up the headers for the file sheet
        ws.append(format_table_headers(sheet_data, col_name=f"Haul Number ({sex.capitalize()})"))

        # Add actual column names (i.e. age) and datatable rows
        append_datatable_rows(ws, sheet_data)

        # Add grand total
        total_row = ["Total", sheet_data.iloc[:-1, :-1].sum().sum()]
        # ---- Extend, if needed
        if sex == "all":
            total_row = total_row + [
                "",
                "Male+Female",
                tables_dict["female"].iloc[:-1, :-1].sum().sum()
                + tables_dict["male"].iloc[:-1, :-1].sum().sum(),
            ]
        # ---- Append
        ws.append(total_row)

        # Add two new rows
        ws.append([""])

        # Append the trailing sheet label along the bottom-most row
        append_sheet_label(ws, f"Aged Length-Haul Counts ({sex})", sex)

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()


####################################################################################################
# Reporter class
# --------------------------------------------------------------------------------------------------


class Reporter:
    """
    Reporter - utility class for writing a suite of Excel reports used in the FEAT pipeline.

    This class centralizes common report generation routines (age-length tables, kriging inputs,
    haul counts, transect reports, etc.) and provides consistent validation and file handling.

    Attributes
    ----------
    save_directory : pathlib.Path
        Directory where generated reports will be saved. Created if it does not exist.
    verbose : bool
        If True, print brief status messages when files are written.

    Examples
    --------
    Basic usage
    >>> from pathlib import Path
    >>> reports = Reporter(Path("reports_out"), verbose=True)
    >>> # Write a kriging input sheet
    >>> reports.kriging_input_report(transect_df, "krig_input.xlsx", "KrigeInput")

    Creating an aged-length haul counts report
    >>> # bio_data must include ['sex','length','length_bin','haul_num']
    >>> sheetnames = {"male": "Male", "female": "Female", "all": "All"}
    >>> reports.aged_length_haul_counts_report(bio_data, "haul_counts.xlsx", sheetnames)

    Creating kriged age-length abundance/biomass reports
    >>> # `datatables` expected to contain keys 'aged' and 'unaged' (pandas DataFrames)
    >>> exclude_filter = {"stratum": ["exclude_this_stratum"]}
    >>> reports.kriged_length_age_abundance_report(datatables, exclude_filter, "krig_abund.xlsx",
    sheetnames)
    >>> reports.kriged_length_age_biomass_report(datatables, exclude_filter, "krig_biomass.xlsx",
    sheetnames)

    Preparing and writing transect reports
    >>> reports.transect_length_age_abundance_report(datatables, "transect_abund.xlsx", sheetnames)
    >>> reports.transect_length_age_biomass_report(datatable, "transect_biomass.xlsx", sheetnames)

    Notes
    -----
    - Methods that accept `datatables` expect a dict-like container with at minimum an 'aged'
      DataFrame.
    - The helper function build_age_length_full_tables is used internally to consolidate logic
      used across multiple report methods (unifying indexing, stacking/unstacking and scaling).
    """

    def __init__(self, save_directory: Union[str, Path], verbose: bool = True):
        """
        Initialize a Reporter instance.

        Parameters
        ----------
        save_directory : str or pathlib.Path
            Directory where reports will be saved. If the directory does not exist it will be
            created. A ValueError is raised if the path cannot be used to create a directory.
        verbose : bool, optional
            When True (default) the instance will print brief status messages when files are
            written.

        Raises
        ------
        TypeError
            If `save_directory` is not a str or pathlib.Path or `verbose` is not bool.
        PermissionError
            If the directory cannot be created due to filesystem permissions or other IO issues.

        Examples
        --------
        >>> from pathlib import Path
        >>> reports = Reporter(Path("reports_out"), verbose=True)
        >>> # directory is created if missing and reports.verbose is True
        """

        # Ensure Path-typing
        save_directory = Path(save_directory)

        # Type checking
        if not isinstance(save_directory, Path):
            raise TypeError("'save_directory' must be a str or `pathlib.Path`.")

        # Validate existence -- create path if missing
        try:
            if not save_directory.exists():
                save_directory.mkdir(parents=True, exist_ok=True)
            # ---- Store attribute
            self.save_directory = save_directory
        except Exception as e:
            raise PermissionError(
                f"The save directory '{save_directory.as_posix()}' could not be accessed or "
                f"created due to the following error(s):\n{e}"
            )

        # Store verbosity
        self.verbose = verbose

    def aged_length_haul_counts_report(
        self,
        filename: str,
        sheetnames: Dict[str, str],
        bio_data: pd.DataFrame,
    ) -> None:
        """
        Create and write an aged-length haul counts Excel report.

        The method builds pivot tables of haul-level length counts stratified by sex and writes
        three sheets (male, female, all) to an Excel workbook in `self.save_directory`.

        Parameters
        ----------
        filename : str
            Name of the Excel file to create (relative to self.save_directory).
        sheetnames : dict
            Mapping with keys 'male','female','all' and values for the worksheet names to use.
        bio_data : pandas.DataFrame
            Raw biological specimen data. Required columns:
            - 'sex' (values expected: 'male', 'female', potentially 'unsexed')
            - 'length' (numeric)
            - 'length_bin' (interval-like objects with .mid)
            - 'haul_num' (identifier for haul)

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `bio_data` is not a pandas DataFrame or `filename` is not a str or `sheetnames` not
            a dict.
        KeyError
            If `bio_data` is missing any of the required columns or `sheetnames` lacks required
            keys.
        ValueError
            If `bio_data` has no rows after filtering to male/female (no content to pivot).

        Examples
        --------
        >>> from pathlib import Path
        >>> reports = Reporter(Path("out"))
        >>> sheetnames = {"male":"Male", "female":"Female", "all":"All"}
        >>> reports.aged_length_haul_counts_report(bio_df, "haul_counts.xlsx", sheetnames)
        """

        # Initial checks
        # ---- Typing
        if not isinstance(bio_data, pd.DataFrame):
            raise TypeError("'bio_data' must be a `pandas.DataFrame`.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetnames, dict):
            raise TypeError("'sheetnames' must be a `dict`.")
        # ---- Columns
        required_cols = {"sex", "length", "length_bin", "haul_num"}
        missing = required_cols.difference(bio_data.columns)
        if missing:
            raise KeyError(f"'bio_data' missing required columns: {missing}.")

        # Update the filepath
        filepath = self.save_directory / filename

        # Create copy
        bio_data = bio_data.copy()

        # Filter out unsexed -- i.e. keep only sexed fish
        bio_data_cleaned = bio_data.loc[bio_data["sex"].isin(["male", "female"])]

        # Create a pivot table
        bio_pvt = bio_data_cleaned.pivot_table(
            columns=["sex", "length_bin"],
            index=["haul_num"],
            values="length",
            aggfunc="count",
            observed=False,
        )

        # Add a category for "all"
        bio_data_all = bio_data.loc[bio_data["sex"].isin(["male", "female", "unsexed"])]
        # ---- Pivot
        bio_pvt_all = bio_data_all.pivot_table(
            columns=["length_bin"],
            index=["haul_num"],
            values="length",
            aggfunc="count",
            observed=False,
        )
        # ---- Convert into a MultiIndex
        bio_pvt_all.columns = pd.MultiIndex.from_product([["all"], bio_pvt_all.columns])

        # Concatenate the sex-specific and overall data tables
        bio_pvt_full = pd.concat([bio_pvt, bio_pvt_all], axis=1).fillna(0.0)
        # ---- Regain typing and sorting, if needed
        bio_pvt_full.sort_index(inplace=True)
        bio_pvt_full = bio_pvt_full.astype(int)

        # Create dictionary containing the sex-specific tables
        haul_pvt_tables = {
            sex: pivot_haul_tables(bio_pvt_full, sex) for sex in ["male", "female", "all"]
        }

        # Save the *.xlsx sheet
        write_haul_report(haul_pvt_tables, sheetnames, filepath)

        # Verbosity
        if self.verbose:
            print(f"Aged-length haul counts report saved to '{filepath.as_posix()}'.")

    def kriged_aged_biomass_mesh_report(
        self,
        filename: str,
        sheetnames: Dict[str, str],
        kriged_data: pd.DataFrame,
        weight_data: pd.DataFrame,
        kriged_stratum_link: Dict[str, str],
        exclude_filter: Dict[str, Any] = {},
    ) -> None:
        """
        Produce an Excel workbook containing kriged age-specific biomass sheets.

        High-level steps:
        - Normalize the weight/age distribution (age-weight proportions) per stratum
        - Rename `kriged_data` stratum columns using `kriged_stratum_link`
        - Multiply kriged biomass by age/sex proportions to produce sex-specific aged biomass
        - Format/rename columns and write three sheets (all/female/male)

        Parameters
        ----------
        filename : str
            Output filename within `self.save_directory`.
        sheetnames : dict
            Mapping sex -> worksheet name with keys 'all', 'female', 'male'.
        kriged_data : pandas.DataFrame
            Kriged mesh with one row per mesh cell and at least the stratum identifier and
            biomass-related columns. The stratum must be mapped to `weight_data` using
            `kriged_stratum_link`.
        weight_data : pandas.DataFrame
            Multi-indexed DataFrame representing age-weight distributions. Expected to have a
            MultiIndex for columns that includes 'sex' and a single stratum level (e.g. 'stratum').
        kriged_stratum_link : dict
            Mapping from column names in `kriged_data` to the stratum names used by `weight_data`.
        exclude_filter : dict, optional
            Optional filters passed to `utils.apply_filters` to zero-out or exclude strata before
            computing age proportions.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `kriged_data` or `weight_data` are not pandas DataFrames or other args have wrong
            types.
        ValueError
            If `weight_data` does not contain exactly one stratum-level name in its column
            MultiIndex.
        KeyError
            If expected strata (after renaming) are missing.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> sheetnames = {"all":"All", "female":"Fem", "male":"Male"}
        >>> reports.kriged_aged_biomass_mesh_report("krig_biomass.xlsx", kriged_df, weight_df,
        {'old':'stratum'}, sheetnames)
        """

        # Type checking
        if not isinstance(kriged_data, pd.DataFrame) or not isinstance(weight_data, pd.DataFrame):
            raise TypeError("'kriged_data' and 'weight_data' must be `pandas.DataFrame`s.")
        if not isinstance(sheetnames, dict) or not isinstance(kriged_stratum_link, dict):
            raise TypeError("'sheetnames' and 'kriged_stratum_link' must be `dict`s.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")

        # Create DataFrame copies
        kriged_data = kriged_data.copy()
        weight_data = weight_data.copy()

        # Update the filepath
        filepath = self.save_directory / filename

        # Get the stratum name
        stratum_name = list(set(weight_data.columns.names).difference(["age_bin", "sex"]))

        # Sum across lengths
        age_weight_sums = weight_data.sum(axis=0).unstack(["sex"] + stratum_name)

        # Apply exclusion filter, if supplied
        age_weight_sums_filtered = utils.apply_filters(
            age_weight_sums, exclude_filter=exclude_filter, replace_value=0.0
        )

        # Compute the pivoted weight proportions
        age_weight_proportions = pivot_aged_weight_proportions(age_weight_sums_filtered)

        # Set the stratum name based on the link function
        kriged_data.rename(columns=kriged_stratum_link, inplace=True)

        # Set the index
        kriged_data.set_index(stratum_name, inplace=True)

        # Combine the overall biomass estimates with the age-specific estimates
        kriged_tables = {
            sex: pivot_aged_dataframe(
                kriged_data, age_weight_proportions[sex], "biomass", sex, ["latitude", "longitude"]
            )
            for sex in ["all", "female", "male"]
        }

        # Format the tables
        prepare_aged_biomass_dataframes(kriged_tables)

        # Write/save the report
        write_aged_dataframe_report(kriged_tables, filepath, sheetnames)

        # Verbosity
        if self.verbose:
            print(f"Kriged aged biomass mesh report saved to '{filepath.as_posix()}'.")

    def kriged_mesh_results_report(
        self,
        filename: str,
        sheetname: str,
        kriged_data: pd.DataFrame,
        kriged_stratum: str,
        kriged_variable: str,
        sigma_bs_data: pd.DataFrame,
        sigma_bs_stratum: str,
    ) -> None:
        """
        Write a single-sheet kriged mesh results report with uncertainty metrics.

        Parameters
        ----------
        filename : str
            File name to save (in `self.save_directory`).
        sheetname : str
            Excel worksheet name to use.
        kriged_data : pandas.DataFrame
            Kriged mesh containing at minimum columns for latitude, longitude, kriged_variable,
            and `cell_cv`. The stratum column is indicated by `kriged_stratum`.
        kriged_stratum : str
            Column name or index name in `kriged_data` identifying stratum; used to align
            `sigma_bs_data`.
        kriged_variable : str
            Name of the kriged estimate column to include (e.g., 'biomass').
        sigma_bs_data : pandas.DataFrame
            DataFrame containing sigma_bs by stratum. Must include `sigma_bs` column and a column
            named by `sigma_bs_stratum` (or index).
        sigma_bs_stratum : str
            Column or index name in `sigma_bs_data` that identifies strata.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If inputs are not pandas DataFrames or strings.
        KeyError
            If required columns are not found in `kriged_data` or `sigma_bs_data`.
        ValueError
            If shapes or indexes cannot be aligned.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> reports.kriged_mesh_results_report("krig_mesh.xlsx", "Mesh", krig_df, "stratum",
        "biomass", sigma_df, "stratum")
        """

        # Type checking
        if not isinstance(kriged_data, pd.DataFrame) or not isinstance(sigma_bs_data, pd.DataFrame):
            raise TypeError("'kriged_data' and 'sigma_bs_data' must be `pandas.DataFrame`s.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetname, str):
            raise TypeError("'sheetname' must be a `str`.")

        # Create DataFrame copies
        kriged_data = kriged_data.copy()
        sigma_bs_data = sigma_bs_data.copy()

        # Update the filepath
        filepath = self.save_directory / filename

        # Check for kriged variable
        if kriged_variable not in kriged_data.columns:
            raise KeyError(f"Column '{kriged_variable}' not found in the kriged mesh DataFrame.")

        # Validate kriged mesh stratum name
        if (
            kriged_stratum not in kriged_data.columns
            and kriged_stratum not in kriged_data.index.names
        ):
            raise KeyError(f"Column '{kriged_stratum}' not found in the kriged mesh DataFrame.")
        else:
            # ---- Reset index
            kriged_data.reset_index(inplace=True)
            # ---- Rename to generic
            kriged_data.rename(columns={kriged_stratum: "stratum"}, inplace=True)
            # ---- Set new index
            kriged_data.set_index("stratum", inplace=True)

        # Validate sigma_bs stratum name
        if (
            sigma_bs_stratum not in sigma_bs_data.columns
            and sigma_bs_stratum not in sigma_bs_data.index.names
        ):
            raise KeyError(f"Column '{sigma_bs_stratum}' not found in the sigma_bs DataFrame.")
        else:
            # ---- Reset index
            sigma_bs_data.reset_index(inplace=True)
            # ---- Rename to generic
            sigma_bs_data.rename(columns={sigma_bs_stratum: "stratum"}, inplace=True)
            # ---- Set new index
            sigma_bs_data.set_index("stratum", inplace=True)
            # ---- Reindex to align with kriged mesh DataFrame
            sigma_bs_data = sigma_bs_data.reindex(kriged_data.index)

        # Add column to kriged data
        kriged_data["sig_b"] = 4.0 * np.pi * sigma_bs_data["sigma_bs"]
        # ---- Reset its index
        kriged_data.reset_index(inplace=True)

        # Add `kriged_sd` to the kriged data
        kriged_data["krig_SD"] = kriged_data[kriged_variable] * kriged_data["cell_cv"]

        # Subset the columns that are required for the report
        output_df = kriged_data.filter(
            [
                "longitude",
                "latitude",
                "stratum",
                "nasc",
                "abundance_male",
                "abundance_female",
                "abundance",
                "biomass_male",
                "biomass_female",
                "biomass",
                "sig_b",
                "cell_cv",
                "krig_SD",
            ]
        )

        # Rename the columns to match required output
        output_df.rename(
            columns={
                "latitude": "Lat",
                "longitude": "Lon",
                "nasc": "NASC",
                "abundance_male": "ntk_male",
                "abundance_female": "ntk_female",
                "abundance": "ntk_total",
                "biomass_male": "wgt_male",
                "biomass_female": "wgt_female",
                "biomass": "wgt_total",
                "cell_cv": "krig_CV",
            },
            inplace=True,
        )

        # Save the *.xlsx sheet
        output_df.to_excel(filepath, sheet_name=sheetname, index=None)

        # Verbosity
        if self.verbose:
            print(f"Kriged mesh report saved to '{filepath.as_posix()}'.")

    def kriged_length_age_abundance_report(
        self,
        filename: str,
        sheetnames: Dict[str, str],
        datatables: Dict[str, pd.DataFrame],
        exclude_filter: Dict[str, Any] = {},
    ) -> None:
        """
        Create kriged age-length abundance reports and write a 3-sheet workbook.

        This method:
        - Builds aged/unaged full tables (male/female/all), optionally redistributing using
          the provided `exclude_filter` via apportion.redistribute_population_table
        - Converts interval indices to numeric midpoints for table rows/columns
        - Writes per-sex pivot tables using internal write helpers

        Parameters
        ----------
        filename : str
            Output workbook filename.
        sheetnames : dict
            Mapping sex -> sheet name (keys: 'male','female','all').
        datatables : dict
            Dictionary with keys:
            - 'aged' : pandas.DataFrame (required)
            - 'unaged' : pandas.DataFrame (optional but recommended)
            The function expects DataFrames with interval-like index levels `length_bin` and
            column levels `age_bin` and `sex` (or compatible structure).
        exclude_filter : dict
            Filter dictionary passed to :func:`apportion.reallocate_population_table` to
            zero/exclude strata.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `datatables` is not a dict or other args are wrong types.
        KeyError
            If required keys are missing from `datatables` (must include 'aged').
        ValueError
            If there's a structural mismatch during table building.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> sheetnames = {"male":"Male","female":"Female","all":"All"}
        >>> reports.kriged_length_age_abundance_report({'aged':aged_df, 'unaged':unaged_df},
        {'stratum':['S1']}, "krig_abund.xlsx", sheetnames)
        """

        # Type checking
        if not isinstance(datatables, dict):
            raise TypeError("'datatables' must be a `dict`.")
        if "aged" not in datatables or "unaged" not in datatables:
            raise KeyError("'datatables' must contain 'aged' and 'unaged' keys.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetnames, dict):
            raise TypeError("'sheetnames' must be a `dict`.")

        # Update the filepath
        filepath = self.save_directory / filename

        # Pull the aged dataset
        aged_table = datatables["aged"].copy().sum(axis=1).unstack(["age_bin", "sex"])

        # Redistribute the aged table, if required
        aged_table = apportionment.reallocate_excluded_estimates(
            aged_table, exclude_filter, ["sex"]
        )

        # Reorient the aged table
        # ---- Convert the indices to numerics
        aged_table.index = (
            pd.Series(aged_table.index, name="length_bin")
            .apply(lambda x: x.mid)
            .values.to_numpy()
            .astype(int)
        )
        # ---- Restack sex
        aged_table = aged_table.stack("sex", future_stack=True)
        # ---- Convert the column indices to numerics
        aged_table.columns = (
            pd.Series(aged_table.columns).apply(lambda x: x.mid).values.to_numpy().astype(int)
        )
        # ---- Swap the index levels
        aged_table = aged_table.unstack("sex").swaplevel(axis=1)

        # Add "all" to the aged table
        aged_table_all = aged_table["male"] + aged_table["female"]
        # ---- Convert into a MultiIndex
        aged_table_all.columns = pd.MultiIndex.from_product([["all"], aged_table_all.columns])
        # ---- Concatenate
        aged_full_table = pd.concat([aged_table, aged_table_all], axis=1)

        # Reorient the unaged tabled
        unaged_table = (
            datatables["unaged"].copy().sum(axis=1).unstack("sex").loc[:, ["male", "female"]]
        )
        # ---- Convert the indices to numerics
        unaged_table.index = (
            pd.Series(unaged_table.index, name="length_bin")
            .apply(lambda x: x.mid)
            .values.to_numpy()
            .astype(int)
        )

        # Add "all" to the unaged table
        unaged_table_all = (unaged_table["male"] + unaged_table["female"]).to_frame("all")
        # ---- Concatenate
        unaged_full_table = pd.concat([unaged_table, unaged_table_all], axis=1)

        # Prepare and format the tables
        sex_tables = prepare_age_length_tables(aged_full_table, unaged_full_table)

        # Create the report
        write_age_length_table_report(
            sex_tables, filepath, sheetnames, variable="abundance", type="Kriged"
        )

        # Verbosity
        if self.verbose:
            print(f"Kriged age-length abundance report saved to '{filepath.as_posix()}'.")

    def kriged_length_age_biomass_report(
        self,
        filename: str,
        sheetnames: Dict[str, str],
        datatables: Dict[str, pd.DataFrame],
        exclude_filter: Dict[str, Any] = {},
    ) -> None:
        """
        Create kriged age-length biomass reports (values converted to metric megatonnes).

        This is similar to `kriged_length_age_abundance_report` but scales the tables by 1e-9
        (to convert from grams/tonnes as required) before writing.

        Parameters
        ----------
        filename : str
            Output filename.
        sheetnames : dict
            Mapping sex -> worksheet name.
        datatables : dict
            Dictionary with 'aged' and optional 'unaged' pandas DataFrames as described elsewhere.
        exclude_filter : dict
            Filter dict forwarded to :func:`apportion.reallocate_population_table`.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If inputs have incorrect types.
        KeyError
            If required keys are missing from datatables.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> reports.kriged_length_age_biomass_report({'aged':aged_df, 'unaged':unaged_df}, {},
        "krig_biomass.xlsx", sheetnames)
        """

        # Type checking
        if not isinstance(datatables, dict):
            raise TypeError("'datatables' must be a `dict`.")
        if "aged" not in datatables or "unaged" not in datatables:
            raise KeyError("'datatables' must contain 'aged' and 'unaged' keys.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetnames, dict):
            raise TypeError("'sheetnames' must be a `dict`.")

        # Update the filepath
        filepath = self.save_directory / filename

        # Pull the aged dataset
        aged_table = datatables["aged"].copy().sum(axis=1).unstack(["age_bin", "sex"])

        # Redistribute the aged table, if required
        aged_table = apportionment.reallocate_excluded_estimates(
            aged_table, exclude_filter, ["sex"]
        )

        # Reorient the aged table
        # ---- Convert the indices to numerics
        aged_table.index = (
            pd.Series(aged_table.index, name="length_bin")
            .apply(lambda x: x.mid)
            .values.to_numpy()
            .astype(int)
        )
        # ---- Restack sex
        aged_table = aged_table.stack("sex", future_stack=True)
        # ---- Convert the column indices to numerics
        aged_table.columns = (
            pd.Series(aged_table.columns).apply(lambda x: x.mid).values.to_numpy().astype(int)
        )
        # ---- Swap the index levels
        aged_table = aged_table.unstack("sex").swaplevel(axis=1)

        # Add "all" to the aged table
        aged_table_all = aged_table["male"] + aged_table["female"]
        # ---- Convert into a MultiIndex
        aged_table_all.columns = pd.MultiIndex.from_product([["all"], aged_table_all.columns])
        # ---- Concatenate
        aged_full_table = pd.concat([aged_table, aged_table_all], axis=1) * 1e-9

        # Reorient the unaged tabled
        unaged_table = (
            datatables["unaged"].copy().sum(axis=1).unstack("sex").loc[:, ["male", "female"]]
        )
        # ---- Convert the indices to numerics
        unaged_table.index = (
            pd.Series(unaged_table.index, name="length_bin")
            .apply(lambda x: x.mid)
            .values.to_numpy()
            .astype(int)
        )

        # Add "all" to the unaged table
        unaged_table_all = (unaged_table["male"] + unaged_table["female"]).to_frame("all")
        # ---- Concatenate
        unaged_full_table = pd.concat([unaged_table, unaged_table_all], axis=1) * 1e-9

        # Prepare and format the tables
        sex_tables = prepare_age_length_tables(aged_full_table, unaged_full_table)

        # Create the report
        write_age_length_table_report(
            sex_tables, filepath, sheetnames, variable="biomass (mmt)", type="Kriged"
        )

        # Verbosity
        if self.verbose:
            print(f"Kriged age-length biomass report saved to '{filepath.as_posix()}'.")

    def kriging_input_report(
        self,
        filename: str,
        sheetname: str,
        transect_data: pd.DataFrame,
    ) -> None:
        """
        Save key kriging input variables to an Excel sheet.

        Parameters
        ----------
        filename : str
            Output file name saved under `self.save_directory`.
        sheetname : str
            Worksheet name to use inside the workbook.
        transect_data : pandas.DataFrame
            Transect-level DataFrame that must contain the following columns:
            ['latitude','longitude','biomass_density','nasc','number_density'].

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `transect_data` is not a pandas DataFrame or filename/sheetname not strings.
        KeyError
            If `transect_data` does not contain the required columns.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> reports.kriging_input_report(transect_df, "krig_input.xlsx", "Input")
        """

        # Checks
        # ---- Typing
        if not isinstance(transect_data, pd.DataFrame):
            raise TypeError("'transect_data' must be a `pandas.DataFrame`.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetname, str):
            raise TypeError("'sheetnames' must be a `str`.")
        # ---- Columns
        required = {"latitude", "longitude", "biomass_density", "nasc", "number_density"}
        missing = required.difference(transect_data.columns)
        if missing:
            raise KeyError(f"'transect_data' missing required columns: {missing}.")

        # Update the filepath
        filepath = self.save_directory / filename

        # Filter the columns
        transect_output = transect_data.filter(
            ["latitude", "longitude", "biomass_density", "nasc", "number_density"]
        )

        # Rename the columns
        transect_output.rename(
            columns={
                "latitude": "Lat",
                "longitude": "Lon",
                "biomass_density": "Biomass density",
                "nasc": "NASC",
                "number_density": "Number density",
            },
            inplace=True,
        )

        # Save the sheet
        transect_output.to_excel(filepath, sheet_name=sheetname, index=None)

        # Verbosity
        if self.verbose:
            print(f"Kriging input report saved to '{filepath.as_posix()}'.")

    def total_length_haul_counts_report(
        self,
        filename: str,
        sheetnames: Dict[str, str],
        bio_data: Dict[str, pd.DataFrame],
    ) -> None:
        """
        Create an Excel report of combined specimen and length haul totals.

        The method expects `bio_data` to contain DataFrames for 'specimen' and 'length' and will
        build pivot tables, combine sex-specific and 'all' columns and write three sheets.

        Parameters
        ----------
        filename : str
            File name to create in `self.save_directory`.
        sheetnames : Dict[str, str]
            Mapping sex -> sheet name.
        bio_data : dict
            Dictionary with keys:
            - 'specimen' : pandas.DataFrame containing specimen records (must include 'sex',
            'length_bin', 'haul_num')
            - 'length' : pandas.DataFrame containing length counts (must include 'sex',
            'length_bin', 'haul_num', 'length_count')

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If arguments have incorrect types.
        KeyError
            If required keys/columns are missing.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> sheetnames = {"male":"Male","female":"Female","all":"All"}
        >>> reports.total_length_haul_counts_report({'specimen':spec_df, 'length':len_df},
        "total_counts.xlsx", sheetnames)
        """

        # Type checking
        if not isinstance(bio_data, dict):
            raise TypeError("'bio_data' must be a `dict`.")
        if "specimen" not in bio_data or "length" not in bio_data:
            raise KeyError("'bio_data' must include 'specimen' and 'length' keys.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetnames, dict):
            raise TypeError("'sheetnames' must be a `dict`.")

        # Update the filepath
        filepath = self.save_directory / filename

        # Extract the DataFrames
        # ---- Specimen
        specimen_data = bio_data["specimen"].copy()
        # ---- Length
        length_data = bio_data["length"].copy()

        # Filter out unsexed -- i.e. keep only sexed fish
        # ---- Specimen
        specimen_data_cleaned = specimen_data.loc[specimen_data["sex"].isin(["male", "female"])]
        # ---- Length
        length_data_cleaned = length_data.loc[length_data["sex"].isin(["male", "female"])]

        # Create pivot tables
        # ---- Specimen
        spe_pvt = specimen_data_cleaned.pivot_table(
            columns=["sex", "length_bin"],
            index=["haul_num"],
            values="length",
            aggfunc="count",
            observed=False,
        )
        # ---- Length
        len_pvt = length_data_cleaned.pivot_table(
            columns=["sex", "length_bin"],
            index=["haul_num"],
            values="length_count",
            aggfunc="sum",
            observed=False,
        )

        # Add a category for "all": Specimen
        # ---- Subset
        specimen_data_all = specimen_data.loc[
            specimen_data["sex"].isin(["male", "female", "unsexed"])
        ]
        # ---- Pivot
        specimen_pvt_all = specimen_data_all.pivot_table(
            columns=["length_bin"],
            index=["haul_num"],
            values="length",
            aggfunc="count",
            observed=False,
        )
        # ---- Convert into a MultiIndex
        specimen_pvt_all.columns = pd.MultiIndex.from_product([["all"], specimen_pvt_all.columns])

        # Add a category for "all": Length
        # ---- Subset
        length_data_all = length_data.loc[length_data["sex"].isin(["male", "female", "unsexed"])]
        # ---- Pivot
        length_pvt_all = length_data_all.pivot_table(
            columns=["length_bin"],
            index=["haul_num"],
            values="length_count",
            aggfunc="sum",
            observed=False,
        )
        # ---- Convert into a MultiIndex
        length_pvt_all.columns = pd.MultiIndex.from_product([["all"], length_pvt_all.columns])

        # Concatenate the sex-specific and overall data tables
        # ---- Specimen
        spe_pvt_full = pd.concat([spe_pvt, specimen_pvt_all], axis=1)
        # ---- Length
        len_pvt_full = pd.concat([len_pvt, length_pvt_all], axis=1)

        # Combine the datasets
        full_pvt = len_pvt_full.add(spe_pvt_full, fill_value=0)
        full_pvt = full_pvt.replace([np.inf, -np.inf], 0).fillna(0).astype(int)

        # Create dictionary containing the sex-specific tables
        haul_pvt_tables = {
            sex: pivot_haul_tables(full_pvt, sex) for sex in ["male", "female", "all"]
        }

        # Save the *.xlsx sheet
        write_haul_report(haul_pvt_tables, sheetnames, filepath)

        # Verbosity
        if self.verbose:
            print(f"Total haul length counts report saved to '{filepath.as_posix()}'.")

    def transect_aged_biomass_report(
        self,
        filename: str,
        sheetnames: Dict[str, str],
        transect_data: pd.DataFrame,
        weight_data: pd.DataFrame,
        exclude_filter: Dict[str, Any] = {},
    ) -> None:
        """
        Produce a transect-level aged biomass workbook (three sex-specific sheets).

        Steps:
        - Create age-weight proportions per stratum from `weight_data`
        - Apply optional `exclude_filter` to zero/out strata
        - Multiply transect biomass by age/sex proportions and write per-sex sheets

        Parameters
        ----------
        filename : str
            Output filename.
        sheetnames : dict
            Mapping sex -> worksheet name.
        transect_data : pandas.DataFrame
            Transect-level summary data with a stratum identifier consistent with `weight_data`.
        weight_data : pandas.DataFrame
            Multi-indexed age-weight table whose column names include 'sex' and exactly one stratum
            level.
        exclude_filter : dict, optional
            Passed to `utils.apply_filters` to zero excluded strata (default: {}).

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If inputs have wrong types.
        ValueError
            If `weight_data` does not contain exactly one stratum-level in columns.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> reports.transect_aged_biomass_report(transect_df, weight_df, "transect_biomass.xlsx",
        sheetnames)
        """

        # Type checking
        if not isinstance(transect_data, pd.DataFrame):
            raise TypeError("'transect_data' must be a `pandas.DataFrame`.")
        if not isinstance(weight_data, pd.DataFrame):
            raise TypeError("'weight_data' must be a `pandas.DataFrame`.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetnames, dict):
            raise TypeError("'sheetnames' must be a `dict`.")

        # Create copy
        transect_data = transect_data.copy()
        weight_data = weight_data.copy()

        # Update the filepath
        filepath = self.save_directory / filename

        # Get the stratum name
        stratum_name = list(set(weight_data.columns.names).difference(["age_bin", "sex"]))

        # Sum across lengths
        age_weight_sums = weight_data.sum(axis=0).unstack(["age_bin"]).T

        # Apply exclusion filter, if supplied
        age_weight_sums_filtered = utils.apply_filters(
            age_weight_sums, exclude_filter=exclude_filter, replace_value=0.0
        )

        # Compute the pivoted weight proportions
        age_weight_proportions = pivot_aged_weight_proportions(age_weight_sums_filtered)

        # Set the index
        transect_data.set_index(stratum_name, inplace=True)

        # Combine the overall abundance estimates with the age-specific estimates
        transect_tables = {
            sex: pivot_aged_dataframe(
                transect_data,
                age_weight_proportions[sex],
                "biomass",
                sex,
                ["transect_num", "latitude", "longitude"],
            )
            for sex in ["all", "female", "male"]
        }

        # Format the tables
        prepare_aged_biomass_dataframes(transect_tables)

        # Write/save the report
        write_aged_dataframe_report(transect_tables, filepath, sheetnames)

        # Verbosity
        if self.verbose:
            print(f"Transect aged biomass report saved to '{filepath.as_posix()}'.")

    def transect_length_age_abundance_report(
        self,
        filename: str,
        sheetnames: Dict[
            str,
            str,
        ],
        datatables: Dict[str, pd.DataFrame],
    ) -> None:
        """
        Write an un-kriged transect length-age abundance workbook (3 sheets).

        Parameters
        ----------
        filename : str
            Output filename saved into `self.save_directory`.
        sheetnames : dict
            Mapping sex -> worksheet name.
        datatables : dict
            Dictionary with keys 'aged' (required) and 'unaged' (optional). Expected structure
            matches the other age-length methods (interval indices, age_bin and sex columns).

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If inputs are of incorrect types.
        KeyError
            If required keys are missing.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> reports.transect_length_age_abundance_report({'aged':aged_df,'unaged':unaged_df},
        "transect_abund.xlsx", sheetnames)
        """

        # Type checking
        if not isinstance(datatables, dict):
            raise TypeError("'datatables' must be a `dict`.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetnames, dict):
            raise TypeError("'sheetnames' must be a `dict`.")

        # Update the filepath
        filepath = self.save_directory / filename

        # Pull the aged dataset
        aged_table = datatables["aged"].sum(axis=1).unstack(["age_bin", "sex"])

        # Reorient the aged table
        # ---- Convert the indices to numerics
        aged_table.index = (
            pd.Series(aged_table.index, name="length_bin")
            .apply(lambda x: x.mid)
            .values.to_numpy()
            .astype(int)
        )
        # ---- Restack sex
        aged_table = aged_table.stack("sex", future_stack=True)
        # ---- Convert the column indices to numerics
        aged_table.columns = (
            pd.Series(aged_table.columns).apply(lambda x: x.mid).values.to_numpy().astype(int)
        )
        # ---- Swap the index levels
        aged_table = aged_table.unstack("sex").swaplevel(axis=1)

        # Add "all" to the aged table
        aged_table_all = aged_table["male"] + aged_table["female"]
        # ---- Convert into a MultiIndex
        aged_table_all.columns = pd.MultiIndex.from_product([["all"], aged_table_all.columns])
        # ---- Concatenate
        aged_full_table = pd.concat([aged_table, aged_table_all], axis=1)

        # Reorient the unaged tabled
        unaged_table = datatables["unaged"].sum(axis=1).unstack("sex").loc[:, ["male", "female"]]
        # ---- Convert the indices to numerics
        unaged_table.index = (
            pd.Series(unaged_table.index, name="length_bin")
            .apply(lambda x: x.mid)
            .values.to_numpy()
            .astype(int)
        )

        # Add "all" to the unaged table
        unaged_table_all = (unaged_table["male"] + unaged_table["female"]).to_frame("all")
        # ---- Concatenate
        unaged_full_table = pd.concat([unaged_table, unaged_table_all], axis=1)

        # Prepare and format the tables
        sex_tables = prepare_age_length_tables(aged_full_table, unaged_full_table)

        # Create the report
        write_age_length_table_report(
            sex_tables, filepath, sheetnames, variable="abundance", type="Un-kriged"
        )

        # Verbosity
        if self.verbose:
            print(f"Transect age-length abundance report saved to '{filepath.as_posix()}'.")

    def transect_length_age_biomass_report(
        self,
        filename: str,
        sheetnames: Dict[
            str,
            str,
        ],
        datatable: pd.DataFrame,
    ) -> None:
        """
        Write an un-kriged transect age-length biomass workbook.

        This function expects `datatable` to be a DataFrame with interval-like index for length
        bins and column levels for age_bin and sex. It converts interval midpoints, composes
        male/female/all tables and writes an Excel workbook with three sheets (male/female/all).

        Parameters
        ----------
        filename : str
            Output workbook filename (relative to save_directory).
        sheetnames : dict
            Mapping sex -> worksheet name (keys: 'male','female','all').
        datatable : pandas.DataFrame
            A DataFrame representing aged counts/weights indexed by intervals (length_bin) with
            column levels ['age_bin','sex'] (or compatible). Values are expected to be numeric.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `datatable` is not a pandas DataFrame or filename/sheetnames have wrong types.
        ValueError
            If `datatable` lacks expected structure (index/columns).

        Examples
        --------
        >>> reports = Reporter("out")
        >>> reports.transect_length_age_biomass_report(datatable, "transect_biomass.xlsx",
        sheetnames)
        """

        # Type checking
        if not isinstance(datatable, pd.DataFrame):
            raise TypeError("'datatable' must be a `pandas.DataFrame`.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetnames, dict):
            raise TypeError("'sheetname' must be a `dict`.")

        # Create copy
        datatable = datatable.copy()

        # Update the filepath
        filepath = self.save_directory / filename

        # Pull the aged dataset
        aged_table = datatable.sum(axis=1).unstack(["age_bin", "sex"])

        # Reorient the aged table
        # ---- Convert the indices to numerics
        aged_table.index = (
            pd.Series(aged_table.index, name="length_bin")
            .apply(lambda x: x.mid)
            .values.to_numpy()
            .astype(int)
        )
        # ---- Restack sex
        aged_table = aged_table.stack("sex", future_stack=True)
        # ---- Convert the column indices to numerics
        aged_table.columns = (
            pd.Series(aged_table.columns).apply(lambda x: x.mid).values.to_numpy().astype(int)
        )
        # ---- Swap the index levels
        aged_table = aged_table.unstack("sex").swaplevel(axis=1)

        # Add "all" to the aged table
        aged_table_all = aged_table["male"] + aged_table["female"]
        # ---- Convert into a MultiIndex
        aged_table_all.columns = pd.MultiIndex.from_product([["all"], aged_table_all.columns])
        # ---- Concatenate
        aged_full_table = pd.concat([aged_table, aged_table_all], axis=1)

        # Initialize dictionary
        sex_tables = {}

        # Iterate through each sex
        for sex in ["male", "female", "all"]:
            # Concatenate the aged and unaged tables
            # ---- Create copy of aged table
            full_table = aged_full_table.copy()[sex]
            # ---- Add margins
            # -------- Across length bins
            full_table["Subtotal"] = full_table.sum(axis=1)
            # ---- Across age bins
            full_table.loc["Subtotal"] = np.concatenate(
                [full_table.iloc[:, :-1].sum(axis=0).to_numpy(), np.array([np.nan])]
            )

            # Add to dictionary
            sex_tables[sex] = full_table

        # Create the report
        write_age_length_table_report(
            sex_tables, filepath, sheetnames, variable="biomass (mmt)", type="Un-kriged"
        )

        # Verbosity
        if self.verbose:
            print(f"Transect age-length biomass report saved to '{filepath.as_posix()}'.")

    def transect_population_results_report(
        self,
        filename: str,
        sheetname: str,
        transect_data: pd.DataFrame,
        weight_strata_data: pd.DataFrame,
        sigma_bs_stratum: pd.DataFrame,
        stratum_name: str,
    ) -> None:
        """
        Write transect-level population results with FEAT-friendly column names.

        The method renames, aligns and exports the transect-level results along with
        computed `sig_b` and `wgt_per_fish` derived from `sigma_bs_stratum` and
        `weight_strata_data` respectively.

        Parameters
        ----------
        filename : str
            Output workbook filename.
        sheetname : str
            Worksheet name to use.
        transect_data : pandas.DataFrame
            Transect-level DataFrame. Must contain a stratum identifier that matches `stratum_name`.
        weight_strata_data : pandas.DataFrame
            DataFrame (possibly multi-indexed) providing weight-per-fish or related metrics.
        sigma_bs_stratum : pandas.DataFrame
            DataFrame indexed or columned by stratum that includes a numeric `sigma_bs` column.
        stratum_name : str
            Column or index name to use as the stratum key to align data between inputs.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If arguments are not the expected types.
        KeyError
            If `sigma_bs_stratum` lacks `sigma_bs` or `stratum_name` cannot be found.

        Examples
        --------
        >>> reports = Reporter("out")
        >>> reports.transect_population_results_report(transect_df, weight_df, sigma_df, "stratum",
        "transect_pop.xlsx", "PopResults")
        """

        # Type checking
        if not isinstance(transect_data, pd.DataFrame):
            raise TypeError("'transect_data' must be a `pandas.DataFrame`.")
        if not isinstance(weight_strata_data, pd.DataFrame):
            raise TypeError("'weight_strata_data' must be a `pandas.DataFrame`.")
        if not isinstance(sigma_bs_stratum, pd.DataFrame):
            raise TypeError("'sigma_bs_stratum' must be a `pandas.DataFrame`.")
        if not isinstance(stratum_name, str):
            raise TypeError("'stratum_name' must be a `str`.")
        if not isinstance(filename, str):
            raise TypeError("'filename' must be a `str`.")
        if not isinstance(sheetname, str):
            raise TypeError("'sheetname' must be a `str`.")

        # Column name mapping for renaming
        RENAME_MAP = {
            "transect_num": "Transect",
            "region_id": "Region ID",
            "distance_s": "VL_start",
            "distance_e": "VL_end",
            "latitude": "Lat",
            "longitude": "Lon",
            f"{stratum_name}": "stratum",
            "bottom_depth": "Depth",
            "nasc": "NASC",
            "abundance_male": "ntk_male",
            "abundance_female": "ntk_female",
            "abundance": "ntk_total",
            "biomass_male": "wgt_male",
            "biomass_female": "wgt_female",
            "biomass": "wgt_total",
            "number_density_male": "nntk_male",
            "number_density_female": "nntk_female",
            "number_density": "nntk_total",
            "biomass_density_male": "nwgt_male",
            "biomass_density_female": "nwgt_female",
            "biomass_density": "nwgt_total",
            "layer_mean_depth": "Layer Depth",
            "layer_height": "Layer height",
            "transect_spacing": "Transect spacing",
            "distance_interval": "interval",
            "nasc_proportion": "mix_coef",
            "sig_b": "sig_b",
            "wgt_per_fish": "wgt_per_fish",
        }

        # Update the filepath
        filepath = self.save_directory / filename

        # Create copy
        transect_data = transect_data.copy()
        weight_strata_data = weight_strata_data.copy()
        sigma_bs_stratum = sigma_bs_stratum.copy()

        # Set index
        transect_data.set_index(stratum_name, inplace=True)

        # Reindex the sigma_bs values
        sigma_bs_stratum = sigma_bs_stratum.reindex(transect_data.index)

        # Reindex average stratum weights
        weight_strata_data = weight_strata_data["all"].reindex(transect_data.index)

        # Add 'sig_b' column
        transect_data["sig_b"] = 4.0 * np.pi * sigma_bs_stratum["sigma_bs"]

        # Add average weight column
        transect_data["wgt_per_fish"] = weight_strata_data

        # Reset the index
        transect_data.reset_index(inplace=True)

        # Check columns
        required = set(RENAME_MAP.keys())
        missing = required.difference(transect_data.columns)
        if missing:
            raise KeyError(f"'transect_data' missing required columns: {missing}.")

        # Rename columns
        transect_data.rename(columns=RENAME_MAP, inplace=True)

        # Filter and sort the columns
        transect_output = transect_data.filter(list(RENAME_MAP.values()))

        # Save the *.xlsx sheet
        transect_output.to_excel(filepath, sheet_name=sheetname, index=None)

        # Verbosity
        if self.verbose:
            print(f"Transect population results report saved to '{filepath.as_posix()}'.")
