from pathlib import Path

import yaml
from openpyxl import load_workbook


def load_configuration(init_config_path: Path, survey_year_config_path: Path):
    """
    Loads the biological, NASC, and stratification
    data using parameters obtained from the configuration
    files.

    Parameters
    ----------
    init_config_path : pathlib.Path
        A string specifying the path to the initialization YAML file
    survey_year_config_path : pathlib.Path
        A string specifying the path to the survey year YAML file

    Notes
    -----
    This function parses the configuration files and incorporates them into
    the Survey class object. This initializes the `config` attribute that
    becomes available for future reference and functions.
    """
    # Validate configuration files
    # ---- Retrieve the module directory to begin mapping the configuration file location
    # ---- current_directory = os.path.dirname(os.path.abspath(__file__))

    # ---- Build the full configuration file paths and verify they exist
    config_files = [init_config_path, survey_year_config_path]
    config_existence = [init_config_path.exists(), survey_year_config_path.exists()]

    # ---- Error evaluation and print message (if applicable)
    if not all(config_existence):
        missing_config = [
            files for files, exists in zip(config_files, config_existence) if not exists
        ]
        raise FileNotFoundError(f"The following configuration files do not exist: {missing_config}")

    # Read configuration files
    # ---- If configuration file existence is confirmed, proceed to reading in the actual files
    # ---- !!! TODO: Incorporate a configuration file validator that enforces required variables
    # ---- and formatting
    init_config_params = yaml.safe_load(init_config_path.read_text())
    survey_year_config_params = yaml.safe_load(survey_year_config_path.read_text())

    # Validate that initialization and survey year configuration parameters do not intersect
    config_intersect = set(init_config_params.keys()).intersection(
        set(survey_year_config_params.keys())
    )

    # Error evaluation, if applicable
    if config_intersect:
        raise RuntimeError(
            f"""The initialization and survey year configuration files comprise the following
            intersecting variables: {config_intersect}"""
        )

    # Format dictionary that will parameterize the `config` class attribute
    # ---- Join the initialization and survey year parameters into a single dictionary
    config_to_add = {**init_config_params, **survey_year_config_params}

    # ---- Amend length/age distribution locations within the configuration attribute
    config_to_add["biometrics"] = {
        "bio_hake_len_bin": init_config_params["bio_hake_len_bin"],
        "bio_hake_age_bin": init_config_params["bio_hake_age_bin"],
    }

    del config_to_add["bio_hake_len_bin"], config_to_add["bio_hake_age_bin"]

    # ---- Pass 'full_params' to the class instance
    return config_to_add


def validate_data_columns(
    file_name: Path, sheet_name: str, config_map: list, validation_settings: dict
):
    """
    Opens a virtual instance of each .xlsx file to validate the presence
    of require data column/variable names

    Parameters
    ----------
    file_name: Path
        File path of data
    sheet_name: str
        Name of Excel sheet containing data
    config_map: list
        A list parsed from the file name that indicates how data attributes
        within `self` are organized
    validation_settings: dict
        The subset CONFIG_MAP settings that contain the target column names
    """

    # Open connection with the workbook and specific sheet
    # This is useful for not calling the workbook into memory and allows for parsing
    # only the necessary rows/column names
    try:
        workbook = load_workbook(file_name, read_only=True)

        # If multiple sheets, iterate through
        sheet_name = [sheet_name] if isinstance(sheet_name, str) else sheet_name

        for sheets in sheet_name:
            sheet = workbook[sheets]

            # Validate that the expected columns are contained within the parsed
            # column names of the workbook
            if "vario_krig_para" in config_map:
                data_columns = [list(row) for row in zip(*sheet.iter_rows(values_only=True))][0]
            else:
                data_columns = {col.value for col in sheet[1]}

            # Error evaluation and print message (if applicable)
            if not set(validation_settings.keys()).issubset(set(data_columns)):
                missing_columns = set(validation_settings.keys()) - set(data_columns)
                raise ValueError(f"Missing columns in the Excel file: {missing_columns}")

        # Close connection to the work book
        workbook.close()

    except Exception as e:
        print(f"Error reading file '{str(file_name)}': {e}")
