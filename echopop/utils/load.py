import copy
from pathlib import Path
from typing import List, Optional, Union

import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook

from ..core import BIODATA_HAUL_MAP, CONFIG_MAP, DATA_STRUCTURE, LAYER_NAME_MAP
from .data_structure_utils import map_imported_datasets
from .validate import CONFIG_DATA_MODEL, CONFIG_INIT_MODEL


def load_configuration(init_config_path: Path, survey_year_config_path: Path):
    """
    Loads the biological, NASC, and stratification
    data using parameters obtained from the configuration
    files.

    Parameters
    ----------
    init_config_path : pathlib.Path
        A string specifying the path to the initialization YAML file
    survey_year_config_path : pathlib.Path
        A string specifying the path to the survey year YAML file

    Notes
    -----
    This function parses the configuration files and incorporates them into
    the Survey class object. This initializes the `config` attribute that
    becomes available for future reference and functions.
    """
    # Validate configuration files
    # Retrieve the module directory to begin mapping the configuration file location
    # current_directory = os.path.dirname(os.path.abspath(__file__))

    # Build the full configuration file paths and verify they exist
    config_files = [init_config_path, survey_year_config_path]
    config_existence = [init_config_path.exists(), survey_year_config_path.exists()]

    # Error evaluation and print message (if applicable)
    if not all(config_existence):
        missing_config = [
            files for files, exists in zip(config_files, config_existence) if not exists
        ]
        raise FileNotFoundError(f"The following configuration files do not exist: {missing_config}")

    # Read configuration files
    # ---- Initialization
    init_config_params = yaml.safe_load(init_config_path.read_text())
    # -------- Validate
    valid_init_config_params = CONFIG_INIT_MODEL(
        init_config_path.as_posix(), **init_config_params
    ).model_dump(exclude_none=True)
    # ---- Survey year data
    survey_year_config_params = yaml.safe_load(survey_year_config_path.read_text())
    # -------- Validate
    valid_survey_year_config_params = CONFIG_DATA_MODEL(
        survey_year_config_path.as_posix(), **survey_year_config_params
    ).model_dump(exclude_none=True)

    # Validate that initialization and survey year configuration parameters do not intersect
    config_intersect = set(valid_init_config_params.keys()).intersection(
        set(valid_survey_year_config_params.keys())
    )

    # Error evaluation, if applicable
    if config_intersect:
        raise RuntimeError(
            f"The initialization and survey year configuration files comprise the following"
            f"intersecting variables: {', '.join(config_intersect)}"
        )

    # Format dictionary that will parameterize the `config` class attribute
    # Join the initialization and survey year parameters into a single dictionary
    config_to_add = {**valid_init_config_params, **valid_survey_year_config_params}

    # Amend length/age distribution locations within the configuration attribute
    config_to_add["biometrics"] = {
        "bio_hake_len_bin": valid_init_config_params["bio_hake_len_bin"],
        "bio_hake_age_bin": valid_init_config_params["bio_hake_age_bin"],
    }

    del config_to_add["bio_hake_len_bin"], config_to_add["bio_hake_age_bin"]

    # Pass 'full_params' to the class instance
    return config_to_add


def load_dataset(
    input_dict: dict, configuration_dict: dict, dataset_type: Optional[Union[str, List[str]]] = None
):
    """
    Loads the biological, NASC, and stratification
    data using parameters obtained from the configuration
    files. This will generate data attributes associated with the tags
    defined in both the configuration yml files and the reference CONFIG_MAP
    and LAYER_NAME_MAP dictionaries.

    Parameters
    ----------
    input_dict: dict
        A dictionary containing the loaded survey data.
    configuration_dict: dict
        Dictionary that contains all of the `Survey`-object configurations found within
        the `config` attribute.
    dataset_type: Union[str, List[str]], optional
        A string (or list of strings) corresponding to the named datasets defined in both
        'CONFIG_MAP' and the dataset definitions located in the file configuration .yaml.
    """

    # Re-initialize the input keys, if needed
    # ---- Get name of the proposed input dictionary keys
    input_keys = [
        (
            LAYER_NAME_MAP[key]["superlayer"][0]
            if LAYER_NAME_MAP[key]["superlayer"]
            else LAYER_NAME_MAP[key]["name"]
        )
        for key in CONFIG_MAP.keys()
        if key in list(dataset_type)
    ]
    # ---- Map the complete datasets
    imported_data = map_imported_datasets(input_dict)
    # ---- Re-initialize if data already loaded to avoid duplication issues
    if set(input_keys).issubset(imported_data):
        # ---- Reset the relevant keys
        input_dict.update({key: copy.deepcopy(DATA_STRUCTURE["input"][key]) for key in input_keys})

    # Check whether data files defined from the configuration file exists
    # ---- Generate flat JSON table comprising all configuration parameter names
    flat_configuration_table = pd.json_normalize(configuration_dict).filter(regex="filename")

    # Convert `dataset_type` to list if needed and defined
    if dataset_type is not None and not isinstance(dataset_type, list):
        dataset_type = [dataset_type]

    # Get the subset table if specific `dataset_type` is defined
    if dataset_type:
        # ---- Get the outermost dictionary keys
        outer_keys = flat_configuration_table.columns.str.split(".").str[0]
        # ---- Get the associated column names
        matching_columns = flat_configuration_table.columns[outer_keys.isin(dataset_type)]
        # ---- Filter the columns
        flat_configuration_table = flat_configuration_table.filter(matching_columns)
    # ---- Default to `CONFIG_MAP` keys otherwise
    else:
        dataset_type = list(CONFIG_MAP.keys())
    # ---- Parse the flattened configuration table to identify data file names and paths
    parsed_filenames = flat_configuration_table.values.flatten()
    # ---- Evaluate whether either file is missing
    data_existence = [
        (Path(configuration_dict["data_root_dir"]) / file).exists() for file in parsed_filenames
    ]

    # Assign the existence status to each configuration file for error evaluation
    # ---- Error evaluation and print message (if applicable)
    if not all(data_existence):
        missing_data = parsed_filenames[~np.array(data_existence)]
        raise FileNotFoundError(f"The following data files do not exist: {missing_data}")

    # Get the applicable `CONFIG_MAP` keys for the defined datasets
    expected_datasets = set(CONFIG_MAP.keys()).intersection(dataset_type)

    # Data validation and import
    # ---- Iterate through known datasets and datalayers
    for dataset in list(expected_datasets):

        for datalayer in [*configuration_dict[dataset].keys()]:

            # Define validation settings from CONFIG_MAP
            validation_settings = CONFIG_MAP[dataset][datalayer]

            # Define configuration settings w/ file + sheet names
            config_settings = configuration_dict[dataset][datalayer]

            # Create reference index of the dictionary path
            config_map = [dataset, datalayer]

            # Define the data layer name
            # ---- Based on the lattermost portion of the file path string
            # Create list for parsing the hard-coded API dictionary
            if dataset == "biological":
                for region_id in [*configuration_dict[dataset][datalayer].keys()]:

                    # Get file and sheet name
                    file_name = (
                        Path(configuration_dict["data_root_dir"])
                        / config_settings[region_id]["filename"]
                    )
                    sheet_name = config_settings[region_id]["sheetname"]

                    # Update `config_map`
                    if len(config_map) == 2:
                        config_map = config_map + [region_id]
                    else:
                        config_map[2] = region_id

                    # Validate column names of this iterated file
                    validate_data_columns(file_name, sheet_name, config_map, validation_settings)

                    # Validate datatypes within dataset and make appropriate changes to dtypes
                    # ---- This first enforces the correct dtype for each imported column
                    # ---- This then assigns the imported data to the correct class attribute
                    read_validated_data(
                        input_dict,
                        configuration_dict,
                        file_name,
                        sheet_name,
                        config_map,
                        validation_settings,
                    )
            else:
                file_name = Path(configuration_dict["data_root_dir"]) / config_settings["filename"]
                sheet_name = config_settings["sheetname"]

                # If multiple sheets, iterate through
                # ---- If multiple sheets, then this needs to be converted accordingly
                sheet_name = [sheet_name] if isinstance(sheet_name, str) else sheet_name

                for sheets in sheet_name:
                    # Update if INPFC
                    if sheets.lower() == "inpfc":
                        # Update validation settings from CONFIG_MAP
                        validation_settings = CONFIG_MAP[dataset]["inpfc_strata"]

                        # Update configuration key map
                        config_map = [dataset, "inpfc_strata"]

                    elif datalayer == "geo_strata":
                        # Update validation settings from CONFIG_MAP
                        validation_settings = CONFIG_MAP[dataset][datalayer]

                        # Update configuration key map
                        config_map = [dataset, datalayer]

                    # Validate datatypes within dataset and make appropriate changes to dtypes
                    # (if necessary)
                    # ---- This first enforces the correct dtype for each imported column
                    # ---- This then assigns the imported data to the correct class attribute
                    validate_data_columns(file_name, sheets, config_map, validation_settings)

                    # Read in data and add to `Survey` object
                    read_validated_data(
                        input_dict,
                        configuration_dict,
                        file_name,
                        sheets,
                        config_map,
                        validation_settings,
                    )

    # Update the data format of various inputs within `Survey`
    prepare_input_data(input_dict, configuration_dict)


def read_validated_data(
    input_dict: dict,
    configuration_dict: dict,
    file_name: Path,
    sheet_name: str,
    config_map: list,
    validation_settings: dict,
):
    """
    Reads in data and validates the data type of each column/variable

    Parameters
    ----------
    input_dict: dict
        Dictionary represent the `input` attribute for the `Survey`-class object
    file_name: Path
        The file name without the prepended file path
    sheet_name: str
        The Excel sheet name containing the target data
    config_map: list
        A list parsed from the file name that indicates how data attributes
        within `self` are organized
    validation_settings: dict
        The subset CONFIG_MAP settings that contain the target column names
    """

    # Based on the configuration settings, read the Excel files into memory. A format
    # exception is made for 'kriging.vario_krig_para' since it requires additional
    # data wrangling (i.e. transposing) to resemble the same dataframe format applied
    # to all other data attributes.
    if "vario_krig_para" in config_map:
        # Read Excel file into memory and then transpose
        df_initial = pd.read_excel(file_name, header=None).T

        # Take the values from the first row and redfine them as the column headers
        df_initial.columns = df_initial.iloc[0]
        df_initial = df_initial.drop(0)

        # Slice only the columns that are relevant to the echopop module functionality
        valid_columns = list(set(validation_settings.keys()).intersection(set(df_initial.columns)))
        df_filtered = df_initial[valid_columns]

        # Ensure the order of columns in df_filtered matches df_initial
        df_filtered = df_filtered[df_initial.columns]

        # Apply data types from validation_settings to the filtered DataFrame
        df = df_filtered.apply(
            lambda col: col.astype(
                validation_settings.get(col.name, type(df_filtered.iloc[0][col.name]))
            )
        )

    else:
        # Read Excel file into memory -- this only reads in the required columns
        df = pd.read_excel(file_name, sheet_name=sheet_name, usecols=validation_settings.keys())

        # Apply data types from validation_settings to the filtered DataFrame
        df = df.apply(lambda col: col.astype(validation_settings.get(col.name, type(col[0]))))

    # Assign the data to their correct data attributes/keys
    if LAYER_NAME_MAP[config_map[0]]["superlayer"] == []:
        sub_attribute = LAYER_NAME_MAP[config_map[0]]["name"]
    else:
        sub_attribute = LAYER_NAME_MAP[config_map[0]]["superlayer"][0]

    # Step 2: Determine whether the dataframe already exists
    if sub_attribute in ["biology", "statistics", "spatial"]:
        if sub_attribute == "biology":
            # Add US / CAN as a region index
            df["region"] = config_map[2]

            # Apply CAN haul number offset
            if config_map[2] == "CAN":
                df["haul_num"] += configuration_dict["CAN_haul_offset"]

        # A single dataframe per entry is expected, so no other fancy operations are needed
        if sheet_name.lower() == "inpfc":
            df_list = [input_dict[sub_attribute]["inpfc_strata_df"], df]
            input_dict[sub_attribute]["inpfc_strata_df"] = pd.concat(df_list)
        else:
            if config_map[0] == "kriging":
                df_list = [input_dict[sub_attribute]["kriging"][config_map[1] + "_df"], df]
                input_dict[sub_attribute]["kriging"][config_map[1] + "_df"] = pd.concat(df_list)
            else:
                df_list = [input_dict[sub_attribute][config_map[1] + "_df"], df]
                input_dict[sub_attribute][config_map[1] + "_df"] = pd.concat(df_list)
    # TODO: This can be refactored out
    elif sub_attribute == "acoustics":

        # Toggle through including and excluding age-1
        if config_map[1] == "no_age1":
            df = df.rename(
                columns={
                    "NASC": "NASC_no_age1",
                    "haul_num": "haul_no_age1",
                    "stratum_num": "stratum_no_age1",
                }
            )
        else:
            df = df.rename(
                columns={
                    "NASC": "NASC_all_ages",
                    "haul_num": "haul_all_ages",
                    "stratum_num": "stratum_all_ages",
                }
            )

        if input_dict["acoustics"]["nasc_df"].shape == (0, 0):
            input_dict["acoustics"]["nasc_df"] = df
        else:
            column_to_add = df.columns.difference(
                input_dict["acoustics"]["nasc_df"].columns
            ).tolist()
            input_dict["acoustics"]["nasc_df"][column_to_add] = df[column_to_add]
    else:
        raise ValueError(
            """Unexpected data attribute structure. Check API settings located in"""
            """the configuration YAML and core.py"""
        )


def validate_data_columns(
    file_name: Path, sheet_name: str, config_map: list, validation_settings: dict
):
    """
    Opens a virtual instance of each .xlsx file to validate the presence
    of require data column/variable names

    Parameters
    ----------
    file_name: Path
        File path of data
    sheet_name: str
        Name of Excel sheet containing data
    config_map: list
        A list parsed from the file name that indicates how data attributes
        within `self` are organized
    validation_settings: dict
        The subset CONFIG_MAP settings that contain the target column names
    """

    # Open connection with the workbook and specific sheet
    # This is useful for not calling the workbook into memory and allows for parsing
    # only the necessary rows/column names
    try:
        workbook = load_workbook(file_name, read_only=True)

        # If multiple sheets, iterate through
        sheet_name = [sheet_name] if isinstance(sheet_name, str) else sheet_name

        for sheets in sheet_name:
            sheet = workbook[sheets]

            # Validate that the expected columns are contained within the parsed
            # column names of the workbook
            if "vario_krig_para" in config_map:
                data_columns = [list(row) for row in zip(*sheet.iter_rows(values_only=True))][0]
            else:
                data_columns = {col.value for col in sheet[1]}

            # Error evaluation and print message (if applicable)
            if not set(validation_settings.keys()).issubset(set(data_columns)):
                missing_columns = set(validation_settings.keys()) - set(data_columns)
                raise ValueError(f"Missing columns in the Excel file: {missing_columns}")

        # Close connection to the work book
        workbook.close()

    except Exception as e:
        print(f"Error reading file '{str(file_name)}': {e}")


def write_haul_to_transect_key(configuration_dict: dict, verbose: bool):
    """
    Function for writing the haul-transect mapping key .xlsx file.

    Parameters
    ----------
    configuration_dict: dict
        Dictionary containing file information, directories, and other values stored within the
        configuration YAML files.
    verbose: bool
        Console messages that will print various messages, updates, etc. when set to True.
    """

    # Get the haul-to-transect mapping settings
    haul_to_transect_settings = configuration_dict["haul_to_transect_mapping"]

    # Get root directory
    root_directory = configuration_dict["data_root_dir"]

    # Validate that `gear_data` exists in configuration settings
    if "gear_data" not in configuration_dict:
        raise KeyError(
            "Directory information for 'gear_data' missing from file configuration settings."
        )

    # Generate haul-to-transect map key
    # ---- Get filename template
    name_template = haul_to_transect_settings["save_file_template"]
    # ---- Define datatypes
    dtypes = {key: values["type"] for key, values in BIODATA_HAUL_MAP.items()}
    # ---- Define column name conversion
    col_names = {key: values["name"] for key, values in BIODATA_HAUL_MAP.items()}
    # ---- Get gear data dictionary
    gear_data = configuration_dict["gear_data"]
    # ---- Get survey year
    survey_year = configuration_dict["survey_year"]
    # ---- Swap out the {YEAR} component
    name_template = name_template.replace("{YEAR}", str(survey_year))
    # ---- Iterate through regions
    for region in gear_data.keys():
        # ---- Update {COUNTRY_CODE} component
        save_file = name_template.replace("{COUNTRY}", region + ".xlsx")
        # ---- Get directory settings
        dir_settings = haul_to_transect_settings["file_settings"][region]
        # ---- Get directory
        save_dir = dir_settings["directory"]
        # ---- Create filepath
        filepath = Path(root_directory) / gear_data[region]["filename"]
        # ---- Get sheetname
        sheet = gear_data[region]["sheetname"]
        # ---- Read file
        data = pd.read_excel(filepath, sheet_name=sheet, usecols=BIODATA_HAUL_MAP.keys()).dropna(
            subset=BIODATA_HAUL_MAP.keys()
        )
        # ---- Update datatypes
        data = data.astype(dtypes)
        # ---- Change column names
        data = data.rename(columns=col_names).reset_index(drop=True)

        # ---- Write the file
        data.to_excel(
            excel_writer=str(Path(root_directory + save_dir) / save_file),
            sheet_name=dir_settings["sheetname"],
            index=False,
        )
        # ---- Print out message
        if verbose:
            print(
                f"Haul-to-transect mapping file for '{region}' saved at "
                f"'{Path(root_directory + save_dir) / save_file}'."
            )


def prepare_input_data(input_dict: dict, configuration_dict: dict):
    """
    Rearranges and organizes data formats of the initial file inputs

    Parameters
    ----------
    input_dict: dict
        Dictionary corresponding to the `input` attribute belonging to `Survey`-class
    configuration_dict: dict
        Dictionary corresponding to the `config` attribute belonging to `Survey`-class
    """

    # Map datasets that are present
    imported_data = map_imported_datasets(input_dict)

    if "biology" in imported_data:
        # Generate length and age vectors
        # ---- Length vector
        length_bins = np.linspace(
            configuration_dict["biometrics"]["bio_hake_len_bin"][0],
            configuration_dict["biometrics"]["bio_hake_len_bin"][1],
            configuration_dict["biometrics"]["bio_hake_len_bin"][2],
            dtype=np.float64,
        )
        # ---- Age vector
        age_bins = np.linspace(
            configuration_dict["biometrics"]["bio_hake_age_bin"][0],
            configuration_dict["biometrics"]["bio_hake_age_bin"][1],
            configuration_dict["biometrics"]["bio_hake_age_bin"][2],
            dtype=np.float64,
        )

        # Discretize these values into discrete intervals
        # ---- Calculate binwidths
        # -------- Length
        length_binwidth = np.mean(np.diff(length_bins / 2.0))
        # -------- Age
        age_binwidth = np.mean(np.diff(age_bins / 2.0))
        # ---- Center the bins within the binwidths
        # -------- Length
        length_centered_bins = np.concatenate(
            ([length_bins[0] - length_binwidth], length_bins + length_binwidth)
        )
        # -------- Age
        age_centered_bins = np.concatenate(([age_bins[0] - age_binwidth], age_bins + age_binwidth))

        # Merge the vector and centered bins into dataframes that will be added into the `input`
        # attribute
        # ---- Generate DataFrame for length
        length_bins_df = pd.DataFrame({"length_bins": length_bins})
        # -------- Discretize the bins as categorical intervals
        length_bins_df["length_intervals"] = pd.cut(
            length_bins_df["length_bins"], length_centered_bins
        )
        # ---- Generate DataFrame for age
        age_bins_df = pd.DataFrame({"age_bins": age_bins})
        # -------- Discretize the bins as categorical intervals
        age_bins_df["age_intervals"] = pd.cut(age_bins_df["age_bins"], age_centered_bins)
        # ---- Update `input` attribute
        # -------- Length
        input_dict["biology"]["distributions"]["length_bins_df"] = length_bins_df
        # -------- Age
        input_dict["biology"]["distributions"]["age_bins_df"] = age_bins_df
        # -------- Delete the duplicate configuration keys
        # del configuration_dict["biometrics"]

        # Relabel sex to literal words among biological data
        # ---- Specimen
        input_dict["biology"]["specimen_df"]["sex"] = np.where(
            input_dict["biology"]["specimen_df"]["sex"] == int(1),
            "male",
            np.where(input_dict["biology"]["specimen_df"]["sex"] == int(2), "female", "unsexed"),
        )
        # -------- Sex group
        input_dict["biology"]["specimen_df"]["group_sex"] = np.where(
            input_dict["biology"]["specimen_df"]["sex"] != "unsexed", "sexed", "unsexed"
        )
        # ---- Length
        input_dict["biology"]["length_df"]["sex"] = np.where(
            input_dict["biology"]["length_df"]["sex"] == int(1),
            "male",
            np.where(input_dict["biology"]["length_df"]["sex"] == int(2), "female", "unsexed"),
        )
        # -------- Sex group
        input_dict["biology"]["length_df"]["group_sex"] = np.where(
            input_dict["biology"]["length_df"]["sex"] != "unsexed", "sexed", "unsexed"
        )

        # Discretize the age and length bins of appropriate biological data
        # ---- Specimen
        input_dict["biology"]["specimen_df"] = input_dict["biology"]["specimen_df"].bin_variable(
            [length_centered_bins, age_centered_bins], ["length", "age"]
        )
        # ---- Length
        input_dict["biology"]["length_df"] = input_dict["biology"]["length_df"].bin_variable(
            length_centered_bins, "length"
        )

    # SPATIAL
    if "spatial" in imported_data:
        # Update column names
        # ---- `geo_strata`
        input_dict["spatial"]["geo_strata_df"].columns = input_dict["spatial"][
            "geo_strata_df"
        ].columns.str.replace(" ", "_")
        # ---- `inpfc_strata`
        input_dict["spatial"]["inpfc_strata_df"].columns = input_dict["spatial"][
            "inpfc_strata_df"
        ].columns.str.replace(" ", "_")
        # ---- `inpfc_strata`: rename stratum column name to avoid conflicts
        input_dict["spatial"]["inpfc_strata_df"].rename(
            columns={"stratum_num": "stratum_inpfc"}, inplace=True
        )

        # Bin data
        # ---- Create latitude intervals to bin the strata
        latitude_bins = np.concatenate(
            [[-90], input_dict["spatial"]["inpfc_strata_df"]["northlimit_latitude"], [90]]
        )
        # ---- Add categorical intervals
        input_dict["spatial"]["inpfc_strata_df"]["latitude_interval"] = pd.cut(
            input_dict["spatial"]["inpfc_strata_df"]["northlimit_latitude"] * 0.99, latitude_bins
        )

    # ACOUSTICS + SPATIAL
    if set(["acoustics", "spatial"]).issubset(imported_data):
        # Bin NASC transects into appropriate INPFC strata
        input_dict["acoustics"]["nasc_df"]["stratum_inpfc"] = (
            pd.cut(
                input_dict["acoustics"]["nasc_df"]["latitude"],
                latitude_bins,
                right=True,
                labels=range(len(latitude_bins) - 1),
            )
        ).astype(int) + 1

    # BIOLOGY + SPATIAL
    if set(["biology", "spatial"]).issubset(imported_data):

        # Merge haul numbers and spatial information across biological variables
        # ---- Consolidate information linking haul-transect-stratum indices
        input_dict["biology"]["haul_to_transect_df"] = input_dict["biology"][
            "haul_to_transect_df"
        ].merge(input_dict["spatial"]["strata_df"], on="haul_num", how="outer")
        # ---- Create interval key for haul numbers to assign INPFC stratum
        haul_bins = np.sort(
            np.unique(
                np.concatenate(
                    [
                        input_dict["spatial"]["inpfc_strata_df"]["haul_start"] - int(1),
                        input_dict["spatial"]["inpfc_strata_df"]["haul_end"],
                    ]
                )
            )
        )
        # ---- Quantize the INPFC dataframe hauls based on strata
        input_dict["spatial"]["inpfc_strata_df"]["haul_bin"] = pd.cut(
            (
                input_dict["spatial"]["inpfc_strata_df"]["haul_start"]
                + input_dict["spatial"]["inpfc_strata_df"]["haul_end"]
            )
            / 2,
            haul_bins,
        )
        # ---- Rename `stratum_num` column
        input_dict["spatial"]["inpfc_strata_df"].rename(
            columns={"stratum_num": "stratum_inpfc"}, inplace=True
        )
        # ---- Define haul bins with `haul_to_transect_df`
        input_dict["biology"]["haul_to_transect_df"]["haul_bin"] = pd.cut(
            input_dict["biology"]["haul_to_transect_df"]["haul_num"], haul_bins
        )
        # ---- Define INPFC stratum for `haul_to_transect_df`
        input_dict["biology"]["haul_to_transect_df"] = input_dict["biology"][
            "haul_to_transect_df"
        ].merge(input_dict["spatial"]["inpfc_strata_df"][["stratum_inpfc", "haul_bin"]], how="left")
        # ---- Distribute this information to other biological variables
        # -------- Specimen
        input_dict["biology"]["specimen_df"] = input_dict["biology"]["specimen_df"].merge(
            input_dict["biology"]["haul_to_transect_df"], how="left"
        )
        # -------- Length
        input_dict["biology"]["length_df"] = input_dict["biology"]["length_df"].merge(
            input_dict["biology"]["haul_to_transect_df"], how="left"
        )
        # -------- Catch
        input_dict["biology"]["catch_df"] = input_dict["biology"]["catch_df"].merge(
            input_dict["biology"]["haul_to_transect_df"], how="left"
        )

    # STATISTICS
    if set(["statistics"]).issubset(imported_data):
        # Reorganize kriging/variogram parameters
        # ---- Kriging
        # -------- Generate dictionary comprising kriging model configuration
        kriging_params = (
            input_dict["statistics"]["kriging"]["vario_krig_para_df"]
            .filter(regex="krig[.]")
            .rename(columns=lambda x: x.replace("krig.", ""))
            .rename(columns={"ratio": "anisotropy", "srad": "search_radius"})
            .to_dict(orient="records")[0]
        )
        # -------- Concatenate configuration settings for kriging
        kriging_params.update(configuration_dict["kriging_parameters"])
        # ---- Variogram
        # -------- Generate dictionary comprising variogram model configuration
        variogram_params = (
            input_dict["statistics"]["kriging"]["vario_krig_para_df"]
            .filter(regex="vario[.]")
            .rename(columns=lambda x: x.replace("vario.", ""))
            .rename(
                columns={
                    "lscl": "correlation_range",
                    "powr": "decay_power",
                    "hole": "hole_effect_range",
                    "res": "lag_resolution",
                    "nugt": "nugget",
                }
            )
            .to_dict(orient="records")[0]
        )
        # ---- Update the input attribute with the reorganized parameters
        input_dict["statistics"]["variogram"].update({"model_config": variogram_params})
        input_dict["statistics"]["kriging"].update({"model_config": kriging_params})
        # -------- Delete the duplicate dataframe
        del input_dict["statistics"]["kriging"]["vario_krig_para_df"]


def validate_config_structure(yaml_data, config_spec):
    """
    Validate configuration YAML dictionary structure and entry types
    """

    # Helper function for validating dictionary key names/structure
    def validate_dict(data, spec, branch=""):
        for key, value in spec.items():
            current_branch = f"{branch}.{key}" if branch else key
            if key == "ANY":
                # If the spec key is "ANY", check all keys in the data against the "ANY" spec
                if not isinstance(data, dict):
                    raise ValueError(
                        f"Expected a dictionary at {current_branch}, got {type(data).__name__}"
                    )
                for sub_key in data:
                    validate_value(data[sub_key], spec[key], f"{current_branch}.{sub_key}")
            else:
                if key not in data:
                    raise KeyError(f"Missing key: {current_branch}")
                validate_value(data[key], value, current_branch)

    # Helper function for parsing values entered at different points throughout the dictionary
    def validate_value(data, spec, branch):
        if isinstance(spec, dict):
            validate_dict(data, spec, branch)
        elif isinstance(spec, type):
            if not is_valid_type(data, spec):
                raise TypeError(
                    f"Expected type {spec.__name__} at {branch}, got {type(data).__name__}"
                )
        elif isinstance(spec, list):
            if not isinstance(data, list):
                raise TypeError(f"Expected a list at {branch}, got {type(data).__name__}")
            if len(spec) != 1:
                raise ValueError(
                    f"Spec list at {branch} should contain exactly one type element, got {spec}"
                )
            for index, item in enumerate(data):
                validate_value(item, spec[0], f"{branch}[{index}]")
        else:
            raise ValueError(f"Unknown spec type at {branch}: {spec}")

    # Helper function for assessing the data type of each nested value
    def is_valid_type(data, spec_type):
        if spec_type is float:
            return isinstance(data, (int, float))
        return isinstance(data, spec_type)

    try:
        validate_dict(yaml_data, config_spec)
    except (KeyError, TypeError, ValueError) as e:
        print(f"Validation error: {e}")
