from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple, Union

import numpy as np
import pandas as pd
import pandas.io.formats.excel as pdif
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..biology import filter_species
from ..survey import Survey

####################################################################################################
# FILE WRITING UTILITY
# --------------------------------------------------------------------------------------------------


def initialize_workbook(filepath: Path) -> Workbook:
    """
    Initialize Excel workbook
    """

    # Check if workbook already exists
    if filepath.exists():
        wb = load_workbook(filepath)
    # ---- If not, generate it and remove the default "Sheet" that is created
    else:
        wb = Workbook()
        wb.remove(wb["Sheet"])

    # Return
    return wb


def format_file_sheet(
    sheetname: str,
    workbook: Workbook,
) -> Worksheet:
    """
    Initialize and format Excel worksheet(s)
    """

    # Check whether the sheetname already exists, and remove if needed
    if sheetname in workbook.sheetnames:
        workbook.remove(workbook[sheetname])

    # Create the new sheet
    _ = workbook.create_sheet(sheetname)

    # Activate and return the new sheet
    return workbook[sheetname]


def format_table_headers(
    dataframe: pd.DataFrame, col_name: str = "Age", row_name: str = "Length (cm)"
) -> List[str]:
    """
    Format worksheet headers independent of the input data
    """

    # Get the shape of the table
    DTSHAPE = dataframe.shape

    # Write data with custom headers and gaps
    # ---- Get midway column
    midway_col_idx = int(np.ceil((DTSHAPE[1] - 1) / 2))

    # Create headers
    headers = [row_name] + [""] * (midway_col_idx - 1) + [col_name]
    # ---- Complete the headers
    return headers + [""] * (DTSHAPE[1] - len(headers))


def append_datatable_rows(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    """
    Write data rows into Excel sheet
    """

    # Add column names (i.e. age)
    worksheet.append([0] + dataframe.columns.values.tolist())

    # Get the dataset row indices
    ROW_INDICES = dataframe.index.values

    # Populate the file with all rows except for the last margins row
    # ---- Iterate along the length bins
    for row in ROW_INDICES[:-1]:
        # ---- Get the row
        row_data = dataframe.loc[row, :]
        # ---- Concatenate row name with the rest of the data
        row_data = [row_data.name] + list(row_data)
        # ---- Append
        worksheet.append(row_data)

    # Append the final row
    # ---- Only apply if 'Subtotal' margin included
    if "Subtotal" in ROW_INDICES:
        worksheet.append(["Subtotal"] + dataframe.loc["Subtotal"].values[:-1].tolist())


def append_table_aggregates(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    sex: Literal["all", "female", "male"],
    tables_dict: Dict[str, pd.DataFrame],
) -> None:
    """
    Add table aggregate calculations to bottom-left margin of worksheet
    """

    # Get the aged sums
    if "Unaged" in dataframe.columns:
        sum_age = dataframe.loc["Subtotal"].values[:-2]
        # ---- Unaged sum
        sum_unaged = dataframe.loc["Subtotal", "Unaged"]
    else:
        sum_age = dataframe.loc["Subtotal"].values[:-1]
        # ---- Unaged sum
        sum_unaged = 0.0

    # Total number across all fish
    total_aged = sum_age.sum() + sum_unaged

    # Calculate the age-1 proportion
    # ---- Proportion
    age1_proportion = dataframe.loc["Subtotal", 1] / sum_age.sum()
    # ---- Compute age-2+ values
    age2_aged = total_aged * (1 - age1_proportion)

    # Add next row
    # ---- Age 1+
    age_1_row = ["Total (age1+)", total_aged]
    # ---- Add the "Over Age" value
    age_1_row = age_1_row + ["", "Over Age:", sum_age.sum(), ""]
    # ---- Add sexed
    if sex == "all":
        age_1_row = age_1_row + [
            "Male+Female:",
            tables_dict["female"].iloc[:-1, :-1].sum().sum()
            + tables_dict["male"].iloc[:-1, :-1].sum().sum(),
        ]
    # ---- Append
    worksheet.append(age_1_row)

    # Add next row
    # ---- Age-2+
    age_all_row = ["Total (age2+)", age2_aged]
    # ---- Add sexed

    if sex == "all":
        age_all_row = (
            age_all_row
            + [""] * 4
            + [
                "Male+Female:",
                tables_dict["female"].iloc[:-1, 1:-1].sum().sum()
                + tables_dict["male"].iloc[:-1, 1:-1].sum().sum(),
            ]
        )
    # ---- Append
    worksheet.append(age_all_row)


def append_sheet_label(
    worksheet: Worksheet,
    title: str,
    sex: str,
) -> None:
    """
    Append sheet title to tail of Excel worksheet
    """

    # Add empty row
    worksheet.append([""])

    # Add sheet label
    sheet_label = title.replace("{SEX}", sex.capitalize())
    # ---- Append
    worksheet.append([""] * 9 + [sheet_label])


####################################################################################################
# DATA WRANGLING
# --------------------------------------------------------------------------------------------------


def pivot_dataframe_data(
    transect_df: pd.DataFrame,
    tables_dict: Dict[str, pd.DataFrame],
    sex: str,
    stratum_name: str,
    contrasts: List[str] = ["transect_num", "longitude", "latitude"],
):
    """
    Pivot DataFrame object for row and column indexing
    """

    # Filter the correct sex for the aged weight proportions
    aged_weight_proportions_df = tables_dict[sex]

    # Assign the correct biomass column name (based on sex)
    if sex != "all":
        biomass_name = f"biomass_{sex}" if f"biomass_{sex}" in transect_df.columns else "biomass"
    else:
        biomass_name = "biomass"

    # Filter the transect dataframe
    transect_data = transect_df.filter(contrasts + [stratum_name, biomass_name])

    # Set the index to the strata
    transect_data.set_index([stratum_name], inplace=True)

    # Merge the transect and weight proportions dataframes
    aged_transect_data = pd.merge(
        transect_data, aged_weight_proportions_df, left_index=True, right_index=True
    )

    # Propagate the biomass column over the age column proportions
    aged_transect_data.loc[:, 1.0:] = aged_transect_data.loc[:, 1.0:].mul(
        aged_transect_data[biomass_name], axis=0
    )

    # Sort the final dataframe and return
    return aged_transect_data.sort_values(contrasts).reset_index()


def repivot_table(
    age_length_dataframe: pd.DataFrame,
    variable: str,
    length_dataframe: Optional[pd.DataFrame] = None,
):
    """
    Repivot existing pivot table(s)
    """

    # Restack the data table
    data_stk = age_length_dataframe.stack(future_stack=True)
    # ---- Sum across columns if strata are present
    if any(["stratum" in name for name in age_length_dataframe.columns.names]):
        data_stk = data_stk.sum(axis=1)
    # ---- Reset index
    data_stk = data_stk.reset_index(name=variable)

    # Concatenate unaged data, if needed
    if length_dataframe is not None:
        # ---- Restack
        unaged_data_stk = length_dataframe.sum(axis=1).reset_index(name=variable)
        # ---- Get the right-most age-bin and add 1
        shifted_bin = data_stk["age_bin"].max() + 1
        # ---- Get the mid-point for later renaming
        shifted_bin_mid = shifted_bin.mid
        # ---- Add age column
        unaged_data_stk["age_bin"] = shifted_bin
        # ---- Concatenate with the aged data
        data_stk = pd.concat([data_stk, unaged_data_stk], ignore_index=True)
    else:
        shifted_bin_mid = None

    # Convert the length and age columns from intervals into numerics
    # ---- Length
    data_stk["length"] = data_stk["length_bin"].apply(lambda x: x.mid).astype(float)
    # ---- Age
    data_stk["age"] = data_stk["age_bin"].apply(lambda x: x.mid).astype(float)

    # Pivot
    data_pvt = data_stk.pivot_table(
        columns=["age"],
        index=["length"],
        values=variable,
        margins_name="Subtotal",
        margins=True,
        aggfunc="sum",
    )

    # Drop the extra axes
    data_proc = data_pvt.rename_axis(columns=None, index=None)

    # Rename the unaged placeholder to "Unaged" and return
    return data_proc.rename(columns={shifted_bin_mid: "Un-aged"})


def pivot_aged_weight_proportions(
    weight_distribution: pd.DataFrame,
    age1_exclusion: bool,
    stratum_name: str,
) -> Dict[str, pd.DataFrame]:
    """
    Pivot distributions of aged weight proportions
    """

    # Stack the age distributions
    weight_stk = weight_distribution.sum().reset_index(name="weight")

    # Concatenate a copy representing "all" fish
    weight_stk_all = (
        weight_stk.groupby([stratum_name, "age_bin"], observed=False)["weight"].sum().reset_index()
    )
    # ---- Add to the original stacked dataframe
    weight_stk_full = pd.concat([weight_stk, weight_stk_all.assign(sex="all")], ignore_index=True)

    # Convert age bins into number ages
    weight_stk_full["age"] = weight_stk_full["age_bin"].apply(lambda x: x.mid).astype(float)

    # Pivot the table
    weight_pvt = weight_stk_full.pivot_table(
        columns=["sex", stratum_name], index=["age"], values="weight"
    )

    # Zero out the age-1 category, if they are being excluded
    if age1_exclusion:
        weight_pvt.loc[1] = 0.0

    # Convert to proportions
    weight_prop_pvt = (weight_pvt / weight_pvt.sum()).fillna(0.0)

    # Create table dictionary
    tables = {sex: weight_prop_pvt.loc[:, sex].T for sex in ["all", "male", "female"]}

    # Return the output
    return tables


def pivot_haul_tables(
    dataframe: pd.DataFrame,
    sex: str,
) -> pd.DataFrame:
    """
    Pivot haul count datatable(s)
    """

    # Subset the data into the specific sex
    haul_sex = dataframe.loc[sex, :]

    # Stack the table
    haul_stk = haul_sex.stack(future_stack=True).reset_index(name="count")

    # Convert the length bins into number lengths
    haul_stk.loc[:, "length"] = haul_stk["length_bin"].apply(lambda x: x.mid).astype(float)

    # Repivot the table with margin subtotals
    haul_agg = haul_stk.pivot_table(
        index=["length"],
        columns=["haul_num"],
        values="count",
        margins=True,
        margins_name="Subtotal",
        aggfunc="sum",
    )

    # Return the new pivot table with column and index names removed
    return haul_agg.rename_axis(columns=None, index=None)


####################################################################################################
# EXCEL REPORT WRITING
# --------------------------------------------------------------------------------------------------


def write_aged_dataframe_report(
    tables_dict: Dict[str, pd.DataFrame],
    filepath: Path,
):
    """
    Write georeferenced population estimates apportioned across age-classes
    """

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Get sheetnames
    sheetnames = {
        "all": "Sheet1",
        "male": "Sheet2",
        "female": "Sheet3",
    }

    # Renaming schema
    column_renaming = {
        "latitude": "Lat",
        "longitude": "Lon",
        "stratum_inpfc": "stratum",
        "stratum_num": "stratum",
        "biomass": "wgt_{SEX}",
    }

    # Initialize
    for sex in ["all", "male", "female"]:

        # Swap out "{SEX}"
        column_renaming_copy = column_renaming.copy()
        # ---- Apply to "biomass"
        if sex == "all":
            column_renaming_copy["biomass"] = "wgt_total"
        else:
            column_renaming_copy["biomass"] = column_renaming_copy["biomass"].replace("{SEX}", sex)

        # Format columns
        # ---- Rename
        tables_dict[sex] = tables_dict[sex].rename(columns=column_renaming_copy)
        # ---- Reorder columns
        column_order = ["Lat", "Lon", "stratum", column_renaming_copy["biomass"]]
        # ---- Filter based on this order
        column_order = column_order + tables_dict[sex].filter(regex=r"\d+").columns.tolist()
        # ---- Apply
        tables_dict[sex] = tables_dict[sex].filter(column_order)

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetnames[sex], wb)

        # Set up superheader representing 'sex'
        if sex != "all":
            ws.append([sex.capitalize()])
        else:
            ws.append([f"{sex.capitalize()} (Male + Female)"])

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()

    # Remove header formatting
    pdif.ExcelFormatter.header_style = None

    # Open up a writer engine
    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:

        # Iterate through all sexes
        for sex in ["all", "male", "female"]:

            # Subset the data dictionary for the particular sheet
            sheet_data = tables_dict[sex]

            # Add actual column names (i.e. age) and datatable rows
            sheet_data.to_excel(writer, sheet_name=sheetnames[sex], startrow=1, index=None)


def write_age_length_table_report(
    tables_dict: Dict[str, pd.DataFrame],
    table_title: str,
    filepath: Path,
) -> None:
    """
    Write pivot table comprising population estimates apportioned across age and length bins
    """

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Get sheetnames
    sheetnames = {
        "male": "Sheet1",
        "female": "Sheet2",
        "all": "Sheet3",
    }

    # Iterate through all sexes
    for sex in ["male", "female", "all"]:

        # Subset the data dictionary for the particular sheet
        sheet_data = tables_dict[sex]

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetnames[sex], wb)

        # Set up the headers for the file sheet
        ws.append(format_table_headers(sheet_data))

        # Add actual column names (i.e. age) and datatable rows
        append_datatable_rows(ws, sheet_data)

        # Append the data table aggregates
        append_table_aggregates(ws, sheet_data, sex, tables_dict)

        # Append the trailing sheet label along the bottom-most row
        append_sheet_label(ws, table_title, sex)

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()


def write_haul_report(
    tables_dict: Dict[str, pd.DataFrame],
    table_title: str,
    filepath: Path,
):
    """
    Write haul count table(s)
    """

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Get sheet name
    sheetname = {
        "male": "Sheet1",
        "female": "Sheet2",
        "all": "Sheet3",
    }

    # Iterate through all sexes
    for sex in ["male", "female", "all"]:

        # Subset the data dictionary for the particular sheet
        sheet_data = tables_dict[sex]

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetname[sex], wb)

        # Set up the headers for the file sheet
        ws.append(format_table_headers(sheet_data, col_name=f"Haul Number ({sex.capitalize()})"))

        # Add actual column names (i.e. age) and datatable rows
        append_datatable_rows(ws, sheet_data)

        # Add grand total
        total_row = ["Total", sheet_data.iloc[:-1, :-1].sum().sum()]
        # ---- Extend, if needed
        if sex == "all":
            total_row = total_row + [
                "",
                "Male+Female",
                tables_dict["female"].iloc[:-1, :-1].sum().sum()
                + tables_dict["male"].iloc[:-1, :-1].sum().sum(),
            ]
        # ---- Append
        ws.append(total_row)

        # Add two new rows
        ws.append([""])

        # Append the trailing sheet label along the bottom-most row
        append_sheet_label(ws, table_title, sex)

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()


####################################################################################################
# FEATReports class
# --------------------------------------------------------------------------------------------------


class FEATReports:
    """
    Echopop FEAT report generation function

    This class includes methods processing and writing reports used by the FEAT team

    Parameters
    ----------
    survey: echopop.survey.Survey
        An `echopop.survey.Survey` class object containing all of the relevant results required for
        producing various `*.xlsx`-formatted reports

    reports: Optional[List[Literal["aged_length_haul_counts", "kriged_aged_biomass_mesh", \
        "kriged_biomass_mesh", "kriged_length_age_biomass", "kriging_input", \
            "total_length_haul_counts", "transect_aged_biomass", "transect_length_age_abundance", \
                "transect_length_age_biomass", "transect_population_results"]]]]]
        A list of compatible reports that can be generated. When no list is supplied, then all
        reports are generated. Available options include:

            - *"aged_length_haul_counts"* \n
            Table comprising distributions of counts from aged fish for each haul

            - *"kriged_aged_biomass_mesh"* \n
            Dataframe comprising georeferenced kriged mesh results with biomass distributed across
            all age bins

            - *"kriging_input"* \n
            Dataframe comprising georeferenced abundance, biomass, and NASC values that can be used
            for the kriging analysis. Note that only the 'biomass' column is currently used within
            `Echopop`

            - *"kriged_length_age_abundance"* \n
            Table comprising kriged abundance estimates distributed over age and length bins

            - *"kriged_length_age_biomass"* \n
            Table comprising kriged biomass estimates distributed over age and length bins

            - *"kriged_mesh_results"* \n
            Dataframe comprising georeferenced kriged mesh results with biomass estimates

            - *"transect_aged_biomass"* \n
            Dataframe comprising georeferenced along-transect biomass estimates with biomass
            distributed across age bins

            - *"transect_length_age_abundance"* \n
            Table comprising along-transect abundance estimates distributed over age and length bins

            - *"transect_length_age_biomass"* \n
            Table comprising along-transect biomass estimates distributed over age and length bins

            - *"transect_population_estimates"* \n
            Dataframe comprising georeferenced along-transect population estimates with that
            includes abundance, biomass, biomass density, number density, and values specific to
            each sex

    save_directory: Optional[Union[str, pathlib.Path]]
        A string or Path object specifying the directory where generated reports will be saved.
        When no directory path is supplied for this argument, the `FEATReports` class will produce
        reports in the folder called `reports` contained within `config["data_root_dir"]` stored
        in the input `echopop.survey.Survey` object

    Attributes
    ----------
    data: echopop.survey.Survey
        The `echopop.survey.Survey`-class object containing all of the data used to generate the
        report `*.xlsx` files

    reports: List[str]
        A list of the report-types that will be produced

    save_directory: pathlib.Path
        A Path object comprising the save directory path where report files will be written

    Methods
    -------
    *generate()*
        Report factory method that generates the `*.xlsx` files used by the FEAT team

    Notes
    -----
    The entire workflow including results from `Survey.transect_analysis()` and
    `Survey.kriging_analysis()` are required to produce all of the report files generated by the
    `FEATreports` object

    Examples
    --------
    >>> from echopop.survey import Survey
    >>> survey = Survey(init_config_path, survey_year_config_path)
    >>> survey.load_survey_data()
    >>> survey.load_acoustic_data()
    >>> survey.transect_analysis(verbose=False)
    >>> survey.kriging_analysis(variogram_parameters=dict(model=["exponential", "bessel"],
    ... n_lags=30), variable="biomass_density", verbose=False)
    >>> from echopop.extensions.FEATreports import FEATreports
    >>> reports = ["aged_length_haul_counts", "kriged_mesh_results"]
    >>> FEATReports(survey, reports, save_directory="C:/Users/Data").generate()
    The following report tables were generated:
       -'C:/Users/Data/aged_length_haul_counts_table.xlsx'
       -'C:/Users/Data/kriged_biomass_mesh_dataframe.xlsx'
       -'C:/Users/Data/kriged_biomass_mesh_reduced_dataframe.xlsx'
    >>> FEATReports(survey, save_directory="C:/Users/Data").generate()
    The following report tables were generated:
       -'C:/Users/Data/aged_length_haul_counts_table.xlsx'
       -'C:/Users/Data/kriged_aged_biomass_mesh_dataframe.xlsx'
       -'C:/Users/Data/kriged_aged_biomass_reduced_mesh_dataframe.xlsx'
       -'C:/Users/Data/kriged_biomass_mesh_dataframe.xlsx'
       -'C:/Users/Data/kriged_biomass_mesh_reduced_dataframe.xlsx'
       -'C:/Users/Data/kriged_length_age_biomass_table.xlsx'
       -'C:/Users/Data/kriging_input_dataframe.xlsx'
       -'C:/Users/Data/total_length_haul_counts_table.xlsx'
       -'C:/Users/Data/transect_aged_biomass_dataframe.xlsx'
       -'C:/Users/Data/transect_aged_biomass_reduced_dataframe.xlsx'
       -'C:/Users/Data/transect_length_age_abundance_table.xlsx'
       -'C:/Users/Data/transect_length_age_biomass_table.xlsx'
       -'C:/Users/Data/transect_population_results_dataframe.xlsx'

    """

    def __init__(
        self,
        survey: Survey,
        reports: Optional[
            List[
                Literal[
                    "aged_length_haul_counts",
                    "kriged_aged_biomass_mesh",
                    "kriging_input",
                    "kriged_length_age_abundance",
                    "kriged_length_age_biomass",
                    "kriged_mesh_results",
                    "total_length_haul_counts",
                    "transect_aged_biomass",
                    "transect_length_age_abundance",
                    "transect_length_age_biomass",
                    "transect_population_results",
                ]
            ]
        ] = None,
        save_directory: Optional[Union[str, Path]] = None,
    ):

        # Assign table-type(s)
        if reports is None:
            reports = [
                "aged_length_haul_counts",
                "kriged_aged_biomass_mesh",
                "kriged_mesh_results",
                "kriged_length_age_abundance",
                "kriged_length_age_biomass",
                "kriging_input",
                "total_length_haul_counts",
                "transect_aged_biomass",
                "transect_length_age_abundance",
                "transect_length_age_biomass",
                "transect_population_results",
            ]
        # ---- Assign to attribute
        self.reports = reports

        # Assign the data
        self.data = survey

        # Save directory
        if not save_directory:
            # ---- Get the save directory configured within the `Survey` configuration
            try:
                save_directory = Path(survey.config["report_path"])
            except KeyError as e:
                # ---- Drop traceback
                e.__traceback__ = None
                raise KeyError(
                    "Report save directory not defined. Either a value for argument "
                    "'save_directory' must be fined, or a valid directory must be defined via "
                    "'report_path' in the file configuration `*.yml` file."
                )
            # ---- Create the directory, if needed
            save_directory.mkdir(parents=True, exist_ok=True)
            # ---- Assign the save directory
            self.save_directory = save_directory
        else:
            # ---- Validate existence
            if not Path(save_directory).exists():
                raise FileNotFoundError(
                    f"The reports save directory '{Path(save_directory).as_posix()}' does not "
                    f"exist!"
                )
            # ---- Assign the save directory
            self.save_directory = Path(save_directory)

    def generate(self):
        """
        Report factory method
        """

        # Validate the reports
        unknown_requests, report_methods = self.get_report_type(self.reports)

        # Print unexpected reports
        if len(unknown_requests) > 0:
            # ---- Format join
            unknown_str = "\n   ".join(f"-'{request}'" for request in unknown_requests)
            # ---- Print
            print(
                f"The following requested reports do not match available report-types (use "
                f"`.report_options()` for a complete list/print-out of available "
                f"reports):\n   {unknown_str}"
            )

        # Run all input report generation functions
        report_files = [v.get("function")(self, **v) for _, v in report_methods.items()]

        # Map methods to the appropriate report-type
        # ---- Initialize an empty list
        report_list = sum(
            ([item] if isinstance(item, str) else list(item) for item in report_files), []
        )
        # ---- Join over lines
        report_str = "\n   ".join(f"-'{file}'" for file in report_list)
        # ---- Print
        print(f"The following report tables were generated:\n   {report_str}")

    @classmethod
    def get_report_type(
        cls,
        requested_reports=List[str],
    ) -> Tuple[List[str], Dict[str, Any]]:

        # Identify superfluous requested reports
        unknown_reports = [
            report for report in requested_reports if report not in cls.__METHOD_REFERENCE__
        ]

        # Find matching methods
        mapped_report_methods = {
            k: cls.__METHOD_REFERENCE__.get(k)
            for k in requested_reports
            if k in cls.__METHOD_REFERENCE__
        }

        # Return
        return unknown_reports, mapped_report_methods

    # REPORT METHODS
    def aged_length_haul_counts_report(
        self,
        title: str,
        filename: str,
        **kwargs,
    ):

        # Get the dataset
        (specimen_data,) = filter_species(
            self.data.input["biology"]["specimen_df"],
            self.data.analysis["settings"]["transect"]["species_id"],
        )

        # Consolidate the dataset to also include "all"
        full_data = pd.concat(
            [
                specimen_data.loc[specimen_data.group_sex == "sexed", :],
                specimen_data.loc[specimen_data.group_sex == "sexed", :].assign(sex="all"),
            ]
        )

        # Remove missing data
        full_data = full_data.loc[~np.isnan(full_data.age)]

        # Initial pivot table
        full_pvt = full_data.pivot_table(
            index=["sex", "length_bin"],
            columns=["haul_num"],
            values="length",
            aggfunc="count",
            observed=False,
        )

        # Create data dictionary containing the sex-specific tables
        haul_pvt_tables = {
            sex: pivot_haul_tables(full_pvt, sex) for sex in ["all", "female", "male"]
        }

        # Write the *.xlsx sheet
        # ---- Get filepath
        filepath = self.save_directory / filename
        # ---- Write file
        write_haul_report(haul_pvt_tables, title, filepath)

        # Return the filepath name
        return filepath.as_posix()

    def kriged_aged_biomass_mesh_report(
        self,
        filename: str,
        **kwargs,
    ):

        # Get 'exclude_age1' argument value
        age1_exclusion = self.data.analysis["settings"]["transect"]["exclude_age1"]

        # Get the stratum name
        stratum_name = self.data.analysis["settings"]["transect"]["stratum_name"]

        # Get the weight distributions
        weight_distribution = self.data.analysis["transect"]["biology"]["distributions"]["weight"][
            "aged_length_weight_tbl"
        ].copy()

        # Reformat the aged weight proportions tables
        tables = pivot_aged_weight_proportions(weight_distribution, age1_exclusion, stratum_name)

        # Get mesh dataframe
        mesh_df = self.data.results["kriging"]["mesh_results_df"].copy()
        # ---- Filter
        mesh_df_copy = mesh_df.filter(["latitude", "longitude", stratum_name, "biomass"]).copy()

        # Combine the pivot tables with the transect tdata
        mesh_pvt_tables = {
            sex: pivot_dataframe_data(
                mesh_df_copy, tables, sex, stratum_name, contrasts=["longitude", "latitude"]
            )
            for sex in ["all", "female", "male"]
        }

        # Remove the empty cells
        mesh_pvt_reduced_tables = {
            sex: mesh_pvt_tables[sex].loc[lambda x: x.biomass > 0.0]
            for sex in ["all", "female", "male"]
        }

        # Get filepaths
        filepath = tuple([self.save_directory / f for f in filename])

        # Write the *.xlsx sheet (full dataset)
        write_aged_dataframe_report(mesh_pvt_tables, filepath[0])

        # Write the *.xlsx sheet (reduced dataset)
        write_aged_dataframe_report(mesh_pvt_reduced_tables, filepath[1])

        # Return the filepath name
        return filepath[0].as_posix(), filepath[1].as_posix()

    def kriged_mesh_results_report(
        self,
        filename: str,
        **kwargs,
    ):

        # Get the stratum name
        stratum_name = self.data.analysis["settings"]["transect"]["stratum_name"]

        # Get mesh results
        mesh_df = self.data.results["kriging"]["mesh_results_df"].copy()

        # Back-calculate the kriged standard deviation
        mesh_df.loc[:, "kriged_sd"] = mesh_df.loc[:, "kriged_mean"] * mesh_df.loc[:, "sample_cv"]
        # ---- Set index
        mesh_df.set_index([stratum_name], inplace=True)

        # Get the mean strata `sigma_bs`
        strata_mean_sigma_bs = (
            self.data.analysis["transect"]["acoustics"]["sigma_bs"]["strata_mean_df"]
            .copy()
            .set_index([stratum_name])
            .reindex(mesh_df.index)
        )
        # ---- Add to the dataframe
        mesh_df["sig_b"] = 4.0 * np.pi * strata_mean_sigma_bs["sigma_bs_mean"]
        # ---- Reset the index
        mesh_df.reset_index(inplace=True)

        # Filter the column names
        # ---- Select columns
        output_df = mesh_df.filter(
            [
                "latitude",
                "longitude",
                stratum_name,
                "nasc",
                "abundance_male",
                "abundance_female",
                "abundance",
                "biomass_male",
                "biomass_female",
                "biomass",
                "sig_b",
                "sample_cv",
                "kriged_sd",
            ]
        )
        # ---- Rename
        output_df.rename(
            columns={
                "latitude": "Lat",
                "longitude": "Lon",
                f"{stratum_name}": "stratum",
                "nasc": "NASC",
                "abundance_male": "ntk_male",
                "abundance_female": "ntk_female",
                "abundance": "ntk_total",
                "biomass_male": "wgt_male",
                "biomass_female": "wgt_female",
                "biomass": "wgt_total",
                "sample_cv": "krig_CV",
                "kriged_sd": "krig_SD",
            },
            inplace=True,
        )

        # Get filepaths
        filepath = tuple([self.save_directory / f for f in filename])

        # Save the *.xlsx sheet (full dataset)
        output_df.to_excel(filepath[0], sheet_name="Sheet1", index=None)

        # Reduce the dataframe and then save
        output_df.loc[output_df["NASC"] > 0.0].to_excel(
            filepath[1], sheet_name="Sheet1", index=None
        )

        # Return the filepath name
        return filepath[0].as_posix(), filepath[1].as_posix()

    def kriged_length_age_abundance_report(
        self,
        title: str,
        filename: str,
        **kwargs,
    ):

        # Get the dataset
        dataset = self.data.results["kriging"]["tables"]
        # ---- Aged data
        aged_data = dataset["aged_tbl"]["abundance_apportioned"].copy(0)
        # ---- Unaged data
        unaged_data = dataset["unaged_tbl"]["abundance_apportioned_unaged"].copy()

        # Process the aged data
        # ---- Stack
        aged_stk = aged_data.stack(future_stack=True).reset_index(name="abundance")
        # ---- Expand to include 'all'
        full_aged_stk = pd.concat(
            [
                aged_stk,
                aged_stk.groupby(["length_bin", "age_bin"], observed=False)["abundance"]
                .sum()
                .reset_index()
                .assign(sex="all"),
            ],
            ignore_index=True,
        )
        # ---- Re-pivot
        full_aged_pvt = full_aged_stk.pivot_table(
            index=["sex", "length_bin"], columns=["age_bin"], values="abundance", observed=False
        )

        # Process the unaged data
        # ---- Stack
        unaged_stk = unaged_data.stack(future_stack=True).reset_index(name="abundance")
        # ---- Expand to include 'all'
        full_unaged_stk = pd.concat(
            [
                unaged_stk,
                unaged_stk.groupby(["length_bin"], observed=False)["abundance"]
                .sum()
                .reset_index()
                .assign(sex="all"),
            ],
            ignore_index=True,
        )
        # ---- Re-pivot
        full_unaged_pvt = full_unaged_stk.pivot_table(
            index=["sex", "length_bin"], values="abundance", observed=False
        )

        # Repivot the datasets for each sex
        tables = {
            sex: repivot_table(
                age_length_dataframe=full_aged_pvt.loc[sex, :],
                length_dataframe=full_unaged_pvt.loc[sex, :],
                variable="abundance",
            )
            for sex in ["all", "male", "female"]
        }

        # Get filepath
        filepath = self.save_directory / filename

        # Write the *.xlsx sheet
        write_age_length_table_report(tables, title, filepath)

        # Return the filepath name
        return filepath.as_posix()

    def kriged_length_age_biomass_report(
        self,
        title: str,
        filename: str,
        **kwargs,
    ):

        # Get the dataset
        dataset = self.data.results["kriging"]["tables"]["overall_apportionment_df"]
        # ---- Expand to include 'all'
        dataset = pd.concat(
            [
                dataset,
                dataset.groupby(["age_bin", "length_bin"], observed=False)["biomass_apportioned"]
                .sum()
                .reset_index()
                .assign(sex="all"),
            ]
        )

        # Initialize a pivot table from the dataset
        dataset_pvt = dataset.pivot_table(
            index=["sex", "length_bin"],
            columns=["age_bin"],
            values="biomass_apportioned",
            aggfunc="sum",
            observed=False,
        )

        # Repivot the datasets for each sex
        tables = {
            sex: repivot_table(dataset_pvt.loc[sex, :] * 1e-9, "biomass")
            for sex in ["all", "male", "female"]
        }

        # Get filepath
        filepath = self.save_directory / filename

        # Write the *.xlsx sheet
        write_age_length_table_report(tables, title, filepath)

        # Return the filepath name
        return filepath.as_posix()

    def kriging_input_report(
        self,
        filename: str,
        **kwargs,
    ):

        # Get the dataset
        dataset = self.data.analysis["transect"]["acoustics"]["adult_transect_df"].sort_values(
            ["transect_num", "longitude", "latitude"]
        )

        # Filter the dataset columns
        dataset_filt = dataset.filter(
            ["latitude", "longitude", "biomass_density", "nasc", "number_density"]
        )

        # Get the filepath
        filepath = self.save_directory / filename

        # Rename the columns
        dataset_filt = dataset_filt.rename(
            columns={
                "latitude": "Lat",
                "longitude": "Lon",
                "biomass_density": "Biomass density",
                "nasc": "NASC",
                "number_density": "Number density",
            }
        )

        # and save the *.xlsx sheet
        dataset_filt.to_excel(filepath, sheet_name="Sheet1", index=None)

        # Return the filepath name
        return filepath.as_posix()

    def total_length_haul_counts_report(
        self,
        title: str,
        filename: str,
        **kwargs,
    ):

        # Get the dataset
        length_data, specimen_data = filter_species(
            [
                self.data.input["biology"]["length_df"],
                self.data.input["biology"]["specimen_df"],
            ],
            self.data.analysis["settings"]["transect"]["species_id"],
        )

        # Consolidate the dataset to also include "all"
        # ---- Length data
        full_length_data = pd.concat(
            [length_data.loc[length_data.group_sex == "sexed", :], length_data.assign(sex="all")]
        )
        # ---- Specimen data
        full_specimen_data = pd.concat(
            [
                specimen_data.loc[specimen_data.group_sex == "sexed", :],
                specimen_data.assign(sex="all"),
            ]
        )

        # Remove missing data
        # ---- Length
        full_length_data = full_length_data.loc[~np.isnan(full_length_data.length)]
        # ---- Specimen
        full_specimen_data = full_specimen_data.loc[
            (~np.isnan(full_specimen_data.length)) & (~np.isnan(full_specimen_data.age))
        ]

        # Initial pivot table
        # ---- Length
        full_length_pvt = full_length_data.pivot_table(
            index=["sex", "length_bin"],
            columns=["haul_num"],
            values="length_count",
            aggfunc="sum",
            observed=False,
        )
        # ---- Specimen
        full_specimen_pvt = full_specimen_data.pivot_table(
            index=["sex", "length_bin"],
            columns=["haul_num"],
            values="length",
            aggfunc="count",
            observed=False,
        )

        # Combine the datasets
        full_pvt = full_length_pvt.add(full_specimen_pvt, fill_value=0).astype(int)

        # Create data dictionary containing the sex-specific tables
        haul_pvt_tables = {
            sex: pivot_haul_tables(full_pvt, sex) for sex in ["all", "female", "male"]
        }

        # Get the filepath
        filepath = self.save_directory / filename

        # Write the *.xlsx sheet
        write_haul_report(haul_pvt_tables, title, filepath)

        # Return the filepath name
        return filepath.as_posix()

    def transect_aged_biomass_report(
        self,
        filename: str,
        **kwargs,
    ):

        # Get 'exclude_age1' argument value
        age1_exclusion = self.data.analysis["settings"]["transect"]["exclude_age1"]

        # Get the stratum name
        stratum_name = self.data.analysis["settings"]["transect"]["stratum_name"]

        # Get the weight distributions
        weight_distribution = self.data.analysis["transect"]["biology"]["distributions"]["weight"][
            "aged_length_weight_tbl"
        ].copy()

        # Get transect dataframe
        transect_df = self.data.analysis["transect"]["acoustics"]["adult_transect_df"].copy()

        # Reformat the aged weight proportions tables
        tables = pivot_aged_weight_proportions(weight_distribution, age1_exclusion, stratum_name)

        # Combine the pivot tables with the transect data
        transect_pvt_tables = {
            sex: pivot_dataframe_data(transect_df, tables, sex, stratum_name)
            for sex in ["all", "female", "male"]
        }

        # Get the filepath
        filepath = tuple([self.save_directory / f for f in filename])

        # Write the *.xlsx sheet (full dataset)
        write_aged_dataframe_report(transect_pvt_tables, filepath[0])

        # Reduce the datasets
        transect_pvt_reduced_tables = {
            sex: pivot_dataframe_data(
                transect_df.loc[transect_df.biomass > 0.0], tables, sex, stratum_name
            )
            for sex in ["all", "female", "male"]
        }

        # Write the *.xlsx sheet (reduced dataset)
        write_aged_dataframe_report(transect_pvt_reduced_tables, filepath[1])

        # Return the filepath name
        return filepath[0].as_posix(), filepath[1].as_posix()

    def transect_length_age_abundance_report(
        self,
        title: str,
        filename: str,
        **kwargs,
    ):

        # Get the underlying dataset
        dataset = self.data.analysis["transect"]["biology"]["population"]["tables"][
            "abundance"
        ].copy()

        # Repivot the datasets for each sex
        # ---- First get the complete aged table
        tables = {
            sex: repivot_table(
                age_length_dataframe=dataset["aged_abundance_df"].loc[sex, :],
                length_dataframe=dataset["unaged_abundance_df"].loc[sex, :],
                variable="abundance",
            )
            for sex in ["all", "male", "female"]
        }

        # Get the filepath
        filepath = self.save_directory / filename

        # Write the *.xlsx sheet
        write_age_length_table_report(tables, title, filepath)

        # Return the filepath name
        return filepath.as_posix()

    def transect_length_age_biomass_report(
        self,
        title: str,
        filename: str,
        **kwargs,
    ) -> str:

        # Get the underlying dataset
        dataset = self.data.analysis["transect"]["biology"]["population"]["tables"]["biomass"][
            "aged_biomass_df"
        ].copy()

        # Repivot the datasets for each sex
        tables = {
            sex: repivot_table(dataset.loc[sex, :] * 1e-9, "biomass")
            for sex in ["all", "male", "female"]
        }

        # Get the filepath
        filepath = self.save_directory / filename

        # Write the *.xlsx sheet
        write_age_length_table_report(tables, title, filepath)

        # Return the filepath name
        return filepath.as_posix()

    def transect_population_results_report(
        self,
        filename: str,
        **kwargs,
    ):

        # Get filepaths
        filepath = tuple([self.save_directory / f for f in filename])

        # Get the stratum column name
        stratum_column = self.data.analysis["settings"]["transect"]["stratum_name"]

        # Get the processed nasc data
        nasc_data = self.data.analysis["transect"]["acoustics"]["adult_transect_df"].copy()
        # ---- Set index
        nasc_data.set_index(["transect_num", "longitude", "latitude", stratum_column], inplace=True)

        # Get the input data
        input_data = self.data.input["acoustics"]["nasc_df"].copy()
        # ---- Set index
        input_data.set_index(["transect_num", "longitude", "latitude"], inplace=True)

        # Get the average strata weights
        weight_strata_df = (
            self.data.analysis["transect"]["biology"]["weight"]["weight_stratum_df"].copy()
        ).set_index([stratum_column])

        # Get the average strata sigma_bs
        sigma_bs_strata_df = (
            self.data.analysis["transect"]["acoustics"]["sigma_bs"]["strata_mean_df"].copy()
        ).set_index([stratum_column])

        # Combine the datasets
        # ---- Copy
        output_data = nasc_data.copy()
        # ---- Merge the initial input data
        output_data.loc[
            :,
            [
                "bottom_depth",
                "layer_height",
                "layer_mean_depth",
                "vessel_log_start",
                "region_id",
                "vessel_log_end",
            ],
        ] = input_data.loc[
            :,
            [
                "bottom_depth",
                "layer_height",
                "layer_mean_depth",
                "region_id",
                "vessel_log_start",
                "vessel_log_end",
            ],
        ]
        # ---- Reset index
        output_data.reset_index(inplace=True)
        # ---- Set index to strata
        output_data.set_index([stratum_column], inplace=True)
        # ---- Merge with the average weight values
        output_data.loc[:, "average_weight"] = weight_strata_df.loc[
            weight_strata_df["sex"] == "all", "average_weight"
        ]
        # ---- Merge with the average sigma_bs values
        output_data.loc[:, "sig_b"] = 4.0 * np.pi * sigma_bs_strata_df["sigma_bs_mean"]
        # ---- Reset index
        output_data.reset_index(inplace=True)

        # Compute the distance
        output_data["interval"] = output_data["vessel_log_end"] - output_data["vessel_log_start"]

        # Rename and filter
        # ---- Filter/reorder columns
        report_df = output_data.filter(
            [
                "transect_num",
                "region_id",
                "vessel_log_start",
                "vessel_log_end",
                "latitude",
                "longitude",
                stratum_column,
                "bottom_depth",
                "nasc",
                "abundance_male",
                "abundance_female",
                "abundance",
                "biomass_male",
                "biomass_female",
                "biomass",
                "number_density_male",
                "number_density_female",
                "number_density",
                "biomass_density_male",
                "biomass_density_female",
                "biomass_density",
                "layer_depth_mean",
                "layer_height",
                "transect_spacing",
                "interval",
                "fraction_hake",
                "sig_b",
                "average_weight",
            ]
        )
        # ---- Sort the values
        report_df.sort_values(
            [
                "transect_num",
                "longitude",
                "latitude",
            ],
            inplace=True,
        )
        # ---- Rename columns
        report_df.rename(
            columns={
                "transect_num": "Transect",
                "region_id": "Region ID",
                "vessel_log_start": "VL_start",
                "vessel_log_end": "VL_end",
                "latitude": "Lat",
                "longitude": "Lon",
                f"{stratum_column}": "stratum",
                "bottom_depth": "Depth",
                "nasc": "NASC",
                "abundance_male": "ntk_male",
                "abundance_female": "ntk_female",
                "abundance": "ntk_total",
                "biomass_male": "wgt_male",
                "biomass_female": "wgt_female",
                "biomass": "wgt_total",
                "number_density_male": "nntk_male",
                "number_density_female": "nntk_female",
                "number_density": "nntk_total",
                "biomass_density_male": "nwgt_male",
                "biomass_density_female": "nwgt_female",
                "biomass_density": "nwgt_total",
                "layer_depth_mean": "Layer Depth",
                "layer_height": "Layer height",
                "transect_spacing": "Transect spacing",
                "fraction_hake": "mix_coef",
                "average_weight": "wgt_per_fish",
            },
            inplace=True,
        )

        # Save the *.xlsx sheet
        report_df.to_excel(filepath[0], sheet_name="Sheet1", index=None)

        # Reduce the datasets
        report_df.loc[report_df["NASC"] > 0.0].to_excel(
            filepath[1], sheet_name="Sheet1", index=None
        )

        # Return the filepath name
        return filepath[0].as_posix(), filepath[1].as_posix()

    @staticmethod
    def report_options():
        """
        Helper method for listing viable report-types
        """

        # Get reference
        options = list(FEATReports.__METHOD_REFERENCE__.keys())

        # Join
        options_str = "\n   ".join(f"-'{report}'" for report in options)

        # Print
        print(f"The following report-types are available for export:\n   {options_str}")

        # Return the list
        return list(options)

    # CLASS REFERENCE DICTIONARIES
    # Reference dictionary: report methods
    __METHOD_REFERENCE__ = {
        "aged_length_haul_counts": {
            "function": aged_length_haul_counts_report,
            "title": "Aged Length-Haul Counts ({SEX})",
            "filename": "aged_len_haul_counts_table.xlsx",
        },
        "kriged_aged_biomass_mesh": {
            "function": kriged_aged_biomass_mesh_report,
            "title": None,
            "filename": (
                "EchoPro_kriged_aged_output_0.xlsx",
                "EchoPro_kriged_aged_output_1.xlsx",
            ),
        },
        "kriging_input": {
            "function": kriging_input_report,
            "title": None,
            "filename": "kriging_input.xlsx",
        },
        "kriged_length_age_abundance": {
            "function": kriged_length_age_abundance_report,
            "title": "Kriged Acoustically Weighted Abundance ({SEX})",
            "filename": "kriged_len_age_abundance_table.xlsx",
        },
        "kriged_length_age_biomass": {
            "function": kriged_length_age_biomass_report,
            "title": "Kriged Acoustically Weighted Biomass (mmt) ({SEX})",
            "filename": "kriged_len_age_biomass_table.xlsx",
        },
        "kriged_mesh_results": {
            "function": kriged_mesh_results_report,
            "title": None,
            "filename": (
                "EchoPro_kriged_output_0.xlsx",
                "EchoPro_kriged_output_1.xlsx",
            ),
        },
        "total_length_haul_counts": {
            "function": total_length_haul_counts_report,
            "title": "Un-Aged Length-Haul Counts ({SEX})",
            "filename": "total_len_haul_counts_table.xlsx",
        },
        "transect_aged_biomass": {
            "function": transect_aged_biomass_report,
            "title": None,
            "filename": (
                "EchoPro_un-kriged_aged_output_0.xlsx",
                "EchoPro_un-kriged_aged_output_1.xlsx",
            ),
        },
        "transect_length_age_abundance": {
            "function": transect_length_age_abundance_report,
            "title": "Transect-based Acoustically Weighted Abundance ({SEX})",
            "filename": "un-kriged_len_age_abundance_table.xlsx",
        },
        "transect_length_age_biomass": {
            "function": transect_length_age_biomass_report,
            "title": "Transect-based Acoustically Weighted Biomass (mmt) ({SEX})",
            "filename": "un-kriged_len_age_biomass_table.xlsx",
        },
        "transect_population_results": {
            "function": transect_population_results_report,
            "title": None,
            "filename": (
                "EchoPro_un-kriged_output_0.xlsx",
                "EchoPro_un-kriged_output_1.xlsx",
            ),
        },
    }
