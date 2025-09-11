import os
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple, Union

import numpy as np
import pandas as pd
import pandas.io.formats.excel as pdif
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

####################################################################################################
# FILE WRITING UTILITY
# --------------------------------------------------------------------------------------------------


def initialize_workbook(filepath: Path) -> Workbook:
    """
    Initialize Excel workbook
    """

    # Check if workbook already exists
    if filepath.exists():
        # ---- Delete
        # wb = load_workbook(filepath)
        os.remove(filepath)
    # ---- If not, generate it and remove the default "Sheet" that is created
    # else:
    wb = Workbook()
    wb.remove(wb["Sheet"])

    # Return
    return wb



def format_file_sheet(
    sheetname: str,
    workbook: Workbook,
) -> Worksheet:
    """
    Initialize and format Excel worksheet(s)
    """

    # Check whether the sheetname already exists, and remove if needed
    if sheetname in workbook.sheetnames:
        workbook.remove(workbook[sheetname])

    # Create the new sheet
    _ = workbook.create_sheet(sheetname)

    # Activate and return the new sheet
    return workbook[sheetname]


def format_table_headers(
    dataframe: pd.DataFrame, col_name: str = "Age", row_name: str = "Length (cm)"
) -> List[str]:
    """
    Format worksheet headers independent of the input data
    """

    # Get the shape of the table
    DTSHAPE = dataframe.shape

    # Write data with custom headers and gaps
    # ---- Get midway column
    midway_col_idx = int(np.ceil((DTSHAPE[1] - 1) / 2))

    # Create headers
    headers = [row_name] + [""] * (midway_col_idx - 1) + [col_name]
    # ---- Complete the headers
    return headers + [""] * (DTSHAPE[1] - len(headers))


def append_datatable_rows(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    """
    Write data rows into Excel sheet
    """

    # Add column names (i.e. age)
    worksheet.append([0] + dataframe.columns.values.tolist())

    # Get the dataset row indices
    ROW_INDICES = dataframe.index.values

    # Populate the file with all rows except for the last margins row
    # ---- Iterate along the length bins
    for row in ROW_INDICES[:-1]:
        # ---- Get the row
        row_data = dataframe.loc[row, :]
        # ---- Concatenate row name with the rest of the data
        row_data = [row_data.name] + list(row_data)
        # ---- Append
        worksheet.append(row_data)

    # Append the final row
    # ---- Only apply if 'Subtotal' margin included
    if "Subtotal" in ROW_INDICES:
        worksheet.append(["Subtotal"] + dataframe.loc["Subtotal"].values[:-1].tolist())


def append_table_aggregates(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    sex: Literal["all", "female", "male"],
    tables_dict: Dict[str, pd.DataFrame],
) -> None:
    """
    Add table aggregate calculations to bottom-left margin of worksheet
    """

    # Get the aged sums
    # ----
    if "Un-aged" in dataframe.columns:
        # ---- Aged sum
        sum_age = dataframe.loc["Subtotal"].values[:-2]
        # ---- Unaged sum
        sum_unaged = dataframe.loc["Subtotal", "Un-aged"]
    else:
        # ---- Aged sum
        sum_age = dataframe.loc["Subtotal"].values[:-1]
        # ---- Unaged sum
        sum_unaged = 0.0

    # Total number across all fish
    total_aged = sum_age.sum()

    # Calculate the age-1 proportion
    # ---- Proportion
    age1_proportion = dataframe.loc["Subtotal", 1] / total_aged
    # ---- Compute age-2+ values
    age2_aged = (total_aged + sum_unaged) * (1 - age1_proportion)

    # Add next row
    # ---- Age 1+
    age_1_row = ["Total (age1+)", total_aged + sum_unaged]
    # ---- Add the "Over Age" value
    age_1_row = age_1_row + ["", "Over Age:", sum_age.sum() + sum_unaged, ""]
    # ---- Add sexed
    if sex == "all":
        # ---- Calculate total
        sexed_total = (
            tables_dict["female"].iloc[:-1, :-1].sum().sum()
            + tables_dict["male"].iloc[:-1, :-1].sum().sum()
        )
        # ---- Append
        age_1_row = age_1_row + ["Male+Female:", sexed_total]
    # ---- Append
    worksheet.append(age_1_row)

    # Add next row
    # ---- Age-2+
    age_all_row = ["Total (age2+)", age2_aged]
    # ---- Add sexed
    if sex == "all":
        age_all_row = age_all_row + [""] * 4 + ["Male+Female:", sexed_total * (1 - age1_proportion)]
    # ---- Append
    worksheet.append(age_all_row)


def append_sheet_label(
    worksheet: Worksheet,
    title: str,
    sex: str,
) -> None:
    """
    Append sheet title to tail of Excel worksheet
    """

    # Add empty row
    worksheet.append([""])

    # Add sheet label
    sheet_label = title.replace("{SEX}", sex.capitalize())
    # ---- Append
    worksheet.append([""] * 9 + [sheet_label])


####################################################################################################
# DATA WRANGLING
# --------------------------------------------------------------------------------------------------


def pivot_dataframe_data(
    transect_df: pd.DataFrame,
    tables_dict: Dict[str, pd.DataFrame],
    sex: str,
    stratum_name: str,
    contrasts: List[str] = ["transect_num", "longitude", "latitude"],
):
    """
    Pivot DataFrame object for row and column indexing
    """

    # Filter the correct sex for the aged weight proportions
    aged_weight_proportions_df = tables_dict[sex]

    # Assign the correct biomass column name (based on sex)
    if sex != "all":
        biomass_name = f"biomass_{sex}" if f"biomass_{sex}" in transect_df.columns else "biomass"
    else:
        biomass_name = "biomass"

    # Filter the transect dataframe
    transect_data = transect_df.filter(contrasts + [stratum_name, biomass_name])

    # Set the index to the strata
    transect_data.set_index([stratum_name], inplace=True)

    # Merge the transect and weight proportions dataframes
    aged_transect_data = pd.merge(
        transect_data, aged_weight_proportions_df, left_index=True, right_index=True
    )

    # Propagate the biomass column over the age column proportions
    aged_transect_data.loc[:, 1.0:] = aged_transect_data.loc[:, 1.0:].mul(
        aged_transect_data[biomass_name], axis=0
    )

    # Sort the final dataframe and return
    return aged_transect_data.sort_values(contrasts).reset_index()


def repivot_table(
    age_length_dataframe: pd.DataFrame,
    variable: str,
    length_dataframe: Optional[pd.DataFrame] = None,
):
    """
    Repivot existing pivot table(s)
    """

    # Restack the data table
    data_stk = age_length_dataframe.stack(future_stack=True)
    # ---- Sum across columns if strata are present
    if any(["stratum" in name for name in age_length_dataframe.columns.names]):
        data_stk = data_stk.sum(axis=1)
    # ---- Reset index
    data_stk = data_stk.reset_index(name=variable)

    # Concatenate unaged data, if needed
    if length_dataframe is not None:
        # ---- Restack
        unaged_data_stk = length_dataframe.sum(axis=1).reset_index(name=variable)
        # ---- Get the right-most age-bin and add 1
        shifted_bin = data_stk["age_bin"].max() + 1
        # ---- Get the mid-point for later renaming
        shifted_bin_mid = shifted_bin.mid
        # ---- Add age column
        unaged_data_stk["age_bin"] = shifted_bin
        # ---- Concatenate with the aged data
        data_stk = pd.concat([data_stk, unaged_data_stk], ignore_index=True)
    else:
        shifted_bin_mid = None

    # Convert the length and age columns from intervals into numerics
    # ---- Length
    data_stk["length"] = data_stk["length_bin"].apply(lambda x: x.mid).astype(float)
    # ---- Age
    data_stk["age"] = data_stk["age_bin"].apply(lambda x: x.mid).astype(float)

    # Pivot
    data_pvt = data_stk.pivot_table(
        columns=["age"],
        index=["length"],
        values=variable,
        margins_name="Subtotal",
        margins=True,
        aggfunc="sum",
    )

    # Drop the extra axes
    data_proc = data_pvt.rename_axis(columns=None, index=None)

    # Rename the unaged placeholder to "Unaged" and return
    return data_proc.rename(columns={shifted_bin_mid: "Un-aged"})


def pivot_aged_weight_proportions(
    weight_distribution: pd.DataFrame,
    age1_exclusion: bool,
    stratum_name: str,
) -> Dict[str, pd.DataFrame]:
    """
    Pivot distributions of aged weight proportions
    """

    # Stack the age distributions
    weight_stk = weight_distribution.sum().reset_index(name="weight")

    # Concatenate a copy representing "all" fish
    weight_stk_all = (
        weight_stk.groupby([stratum_name, "age_bin"], observed=False)["weight"].sum().reset_index()
    )
    # ---- Add to the original stacked dataframe
    weight_stk_full = pd.concat([weight_stk, weight_stk_all.assign(sex="all")], ignore_index=True)

    # Convert age bins into number ages
    weight_stk_full["age"] = weight_stk_full["age_bin"].apply(lambda x: x.mid).astype(float)

    # Pivot the table
    weight_pvt = weight_stk_full.pivot_table(
        columns=["sex", stratum_name], index=["age"], values="weight"
    )

    # Zero out the age-1 category, if they are being excluded
    if age1_exclusion:
        weight_pvt.loc[1] = 0.0

    # Convert to proportions
    weight_prop_pvt = (weight_pvt / weight_pvt.sum()).fillna(0.0)

    # Create table dictionary
    tables = {sex: weight_prop_pvt.loc[:, sex].T for sex in ["all", "male", "female"]}

    # Return the output
    return tables


def pivot_haul_tables(
    dataframe: pd.DataFrame,
    sex: str,
) -> pd.DataFrame:
    """
    Pivot haul count datatable(s)
    """

    # Subset the data into the specific sex
    haul_sex = dataframe.loc[sex, :]

    # Stack the table
    haul_stk = haul_sex.stack(future_stack=True).reset_index(name="count")

    # Convert the length bins into number lengths
    haul_stk.loc[:, "length"] = haul_stk["length_bin"].apply(lambda x: x.mid).astype(float)

    # Repivot the table with margin subtotals
    haul_agg = haul_stk.pivot_table(
        index=["length"],
        columns=["haul_num"],
        values="count",
        margins=True,
        margins_name="Subtotal",
        aggfunc="sum",
    )

    # Return the new pivot table with column and index names removed
    return haul_agg.rename_axis(columns=None, index=None)


####################################################################################################
# EXCEL REPORT WRITING
# --------------------------------------------------------------------------------------------------


def write_aged_dataframe_report(
    tables_dict: Dict[str, pd.DataFrame],
    filepath: Path,
):
    """
    Write georeferenced population estimates apportioned across age-classes
    """

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Get sheetnames
    sheetnames = {
        "all": "Sheet1",
        "male": "Sheet2",
        "female": "Sheet3",
    }

    # Renaming schema
    column_renaming = {
        "latitude": "Lat",
        "longitude": "Lon",
        "stratum_inpfc": "stratum",
        "stratum_num": "stratum",
        "biomass": "wgt_{SEX}",
    }

    # Initialize
    for sex in ["all", "male", "female"]:

        # Swap out "{SEX}"
        column_renaming_copy = column_renaming.copy()
        # ---- Apply to "biomass"
        if sex == "all":
            column_renaming_copy["biomass"] = "wgt_total"
        else:
            column_renaming_copy["biomass"] = column_renaming_copy["biomass"].replace("{SEX}", sex)

        # Format columns
        # ---- Rename
        tables_dict[sex] = tables_dict[sex].rename(columns=column_renaming_copy)
        # ---- Reorder columns
        column_order = ["Lat", "Lon", "stratum", column_renaming_copy["biomass"]]
        # ---- Filter based on this order
        column_order = column_order + tables_dict[sex].filter(regex=r"\d+").columns.tolist()
        # ---- Apply
        tables_dict[sex] = tables_dict[sex].filter(column_order)

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetnames[sex], wb)

        # Set up superheader representing 'sex'
        if sex != "all":
            ws.append([sex.capitalize()])
        else:
            ws.append([f"{sex.capitalize()} (Male + Female)"])

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()

    # Remove header formatting
    pdif.ExcelFormatter.header_style = None

    # Open up a writer engine
    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:

        # Iterate through all sexes
        for sex in ["all", "male", "female"]:

            # Subset the data dictionary for the particular sheet
            sheet_data = tables_dict[sex]

            # Add actual column names (i.e. age) and datatable rows
            sheet_data.to_excel(writer, sheet_name=sheetnames[sex], startrow=1, index=None)


def write_age_length_table_report(
    tables_dict: Dict[str, pd.DataFrame],
    table_title: str,
    filepath: Path,
) -> None:
    """
    Write pivot table comprising population estimates apportioned across age and length bins
    """

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Get sheetnames
    sheetnames = {
        "male": "Sheet1",
        "female": "Sheet2",
        "all": "Sheet3",
    }

    # Iterate through all sexes
    for sex in ["male", "female", "all"]:

        # Subset the data dictionary for the particular sheet
        sheet_data = tables_dict[sex]

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetnames[sex], wb)

        # Set up the headers for the file sheet
        ws.append(format_table_headers(sheet_data))

        # Add actual column names (i.e. age) and datatable rows
        append_datatable_rows(ws, sheet_data)

        # Append the data table aggregates
        append_table_aggregates(ws, sheet_data, sex, tables_dict)

        # Append the trailing sheet label along the bottom-most row
        append_sheet_label(ws, table_title, sex)

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()


def write_haul_report(
    tables_dict: Dict[str, pd.DataFrame],
    table_title: str,
    filepath: Path,
):
    """
    Write haul count table(s)
    """

    # Create the workbook, if needed
    wb = initialize_workbook(filepath)

    # Get sheet name
    sheetname = {
        "male": "Sheet1",
        "female": "Sheet2",
        "all": "Sheet3",
    }

    # Iterate through all sexes
    for sex in ["male", "female", "all"]:

        # Subset the data dictionary for the particular sheet
        sheet_data = tables_dict[sex]

        # Add/overwrite the sheet
        # ws = format_file_sheet(sex.capitalize(), wb)
        ws = format_file_sheet(sheetname[sex], wb)

        # Set up the headers for the file sheet
        ws.append(format_table_headers(sheet_data, col_name=f"Haul Number ({sex.capitalize()})"))

        # Add actual column names (i.e. age) and datatable rows
        append_datatable_rows(ws, sheet_data)

        # Add grand total
        total_row = ["Total", sheet_data.iloc[:-1, :-1].sum().sum()]
        # ---- Extend, if needed
        if sex == "all":
            total_row = total_row + [
                "",
                "Male+Female",
                tables_dict["female"].iloc[:-1, :-1].sum().sum()
                + tables_dict["male"].iloc[:-1, :-1].sum().sum(),
            ]
        # ---- Append
        ws.append(total_row)

        # Add two new rows
        ws.append([""])

        # Append the trailing sheet label along the bottom-most row
        append_sheet_label(ws, table_title, sex)

    # Save the workbook
    wb.save(filepath)
    # ---- Close
    wb.close()


####################################################################################################
# FEATReports class
# --------------------------------------------------------------------------------------------------



class FEATReports:

    def __init__(self, save_directory):