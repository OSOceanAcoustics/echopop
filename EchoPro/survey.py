from pathlib import Path
from typing import List, Union
from pathlib import Path
import pandas as pd
import numpy as np
import yaml
import os
from openpyxl import load_workbook
from .computation.acoustics import ts_length_regression, to_linear
from .core import CONFIG_MAP, LAYER_NAME_MAP
### !!! TODO : This is a temporary import call -- this will need to be changed to 
# the correct relative structure (i.e. '.core' instead of 'EchoPro.core' at a future testing step)
import pprint
from .utils.data_structure_utils import pull_nested_dict, push_nested_dict
### !!! TODO : This is a temporary import call -- this will need to be changed to 
# the correct relative structure (i.e. '.utils.data_structure_utils' instead of 
# 'EchoPro.utils.data_structure_utils' at a future testing step)

class Survey:
    """
    EchoPro base class that imports and prepares parameters for
    a survey. Additionally, it includes functions for accessing
    the modules associated with the transect and Kriging variable
    calculations, CV analysis, semi-variogram algorithm, and Kriging.

    Parameters
    ----------
    init_config_path : str or pathlib.Path
        A string specifying the path to the initialization YAML file
    survey_year_config_path : str or pathlib.Path
        A string specifying the path to the survey year YAML file

    Attributes
    ----------
    meta : dict 
        Metadata variable that provides summary information concerning the
        data contained within the class object (e.g. 'self.summary').
    config : dict 
        Configuration settings and parameters that can be referenced for
        various downstream and internal functions.
    data : dict
        Various dictionaries are incorporated into the Survey class object that 
        are directly referenced for various downstream and internal functions. This
        includes attributes such as 'biology', 'acoustics', and 'spatial' that represent
        various nested biological, acoustic, and spatial/stratification datasets imported
        based on the input files defined via the configuration settings.
    
    """
    def __init__(
        self,
        init_config_path: Union[str, Path] ,
        survey_year_config_path: Union[str, Path] ,
    ):
        ### Loading the configuration settings and definitions that are used to 
        # initialize the Survey class object
        # ATTRIBUTE ADDITIONS: `config`
        self.config = self.load_configuration(init_config_path, survey_year_config_path)

        ### Loading the datasets defined in the configuration files
        # EXAMPLE ATTRIBUTE ADDITIONS: `biology`, `spatial`, `acoustics`
        self.load_survey_data()

        # Define length and age distributions
        self.biometric_distributions()

        ### A utility function that helps to map the datasets currently present
        # within Survey object via the `___.summary` property. This also initializes
        # the `meta` attribute
        # ATTRIBUTE ADDITIONS: `meta`
        ##
        # TODO: Fix bug associated with `load_survey_data()` that results in the 
        # 'dict_tree' key within the `___.config` attribute overwriting previous
        # entries when multiple datasets of the same layer name are present, such as
        # 'US' versus 'CAN' for biological data. At present, it defaults to listing 
        # only 'CAN' -- this is not reflected in the actual data attribute structure where
        # both 'US' and 'CAN' are present, denoted by the 'index' column in relevant dataframes.
        self.populate_tree()

        
    @staticmethod
    def load_configuration( init_config_path: Path , 
                            survey_year_config_path: Path ):
        """
        Loads the biological, NASC, and stratification
        data using parameters obtained from the configuration
        files.

        Parameters
        ----------
        init_config_path : pathlib.Path
            A string specifying the path to the initialization YAML file
        survey_year_config_path : pathlib.Path
            A string specifying the path to the survey year YAML file

        Notes
        -----
        This function parses the configuration files and incorporates them into
        the Survey class object. This initializes the `config` attribute that 
        becomes available for future reference and functions.
        """
        ### Validate configuration files
        # Retreive the module directory to begin mapping the configuration file location
        current_directory = os.path.dirname(os.path.abspath(__file__))

        # Build the full configuration file paths
        config_files = [init_config_path, survey_year_config_path]
        config_file_paths = [os.path.join(current_directory, '..', 'config_files', file_name) for file_name in config_files]

        # Check files to ensure they exist
        config_existence = [os.path.exists(file_path) for file_path in config_file_paths]

        # Assign the existence status to each configuration file for error evaluation
        config_status = dict(zip(config_files, config_existence))

        # Error evaluation and print message (if applicable)
        if not all(config_status.values()):
            missing_config = [config_file for config_file, exists in config_status.items() if not exists]
            raise FileNotFoundError(f"The following configuration files do not exist: {missing_config}")

        ### Read configuration files
        # If configuration file existence is confirmed, proceed to reading in the actual files
        ## !!! TODO: Incorporate a configuration file validator that enforces required variables and formatting
        init_config_params = yaml.safe_load(Path(config_file_paths[0]).read_text())
        survey_year_config_params = yaml.safe_load(Path(config_file_paths[1]).read_text())        

        # Validate that initialization and survey year configuration parameters do not intersect
        param_intersect = set(init_config_params.keys()).intersection(set(survey_year_config_params.keys()))

        # Error evaluation, if applicable
        if param_intersect:
            raise RuntimeError(
                f"The initialization and survey year configuration files comprise the following intersecting variables: {param_intersect}"
            )

        ### Format dictionary that will parameterize the `config` class attribute
        # Join the initialization and survey year parameters into a single dictionary
        full_params = {**init_config_params, **survey_year_config_params}

        # ??? TODO: Since this is a static method, this carriage return may not be the most appropriate way to pass 
        # 'full_params' to the class instance
        return full_params
    
    def search_nested_dict( self ,
                            config: dict ,
                            filename: str ,
                            current_path: list = [] ):
        """
        Recursive utility function that searches the configuration attribute
        (a nested dictionary) for a specific 

        Parameters
        ----------
        self
            Survey class object.
        config : dict
            Survey class configuration ('self.config') dictionary.
        filename : str
            Target filename that is being searched within the configuration dictionary.
        current_path : list
            List that records previously evaluated levels/layers of the configuration
        dictionary and steps forward to the next nested level within the dictionary.

        Returns
        -----
        data_layer_names : list
            List comprising the names of nested dictionary layers.
        reference_path : str
            A string representing the configuration tag that is referenced in the 
        CONFIG_MAP API. 
        sheet_name : str
            References the correct sheet name containing the data that is defined
        in the configuration files.
        """
        ### Initialize 'current_path' if not already defined/initialized
        if current_path is None:
                current_path = []

        ### Recursive function 
        # Step forward through the configuration dictionary until the filename input
        # is located.
        for key, value in config.items():
            # Not located ? Keep searching
            if isinstance(value, dict):
                result = self.search_nested_dict(value, filename, current_path + [key])
                if result is not None:
                    return result
            # Located ? Provide the outputs
            elif key == 'filename' and value in filename:
                return current_path,'/'.join(current_path[0:2]),config.get('sheetname')
            
        ### Never present !? 
        ## !!! TODO : In theory, the 'load_configuration' function should validate that 
        # the information required for the output is present. However, if that leaks through
        # any cracks, then this becomes an unchecked recursive function. Consequently, this
        # should be leashed so that it can only dig for so long until it yields some sort of 
        # Value or Runtime Error. 
        return None     
    
    def load_survey_data( self ):
        """
        Loads the biological, NASC, and stratification
        data using parameters obtained from the configuration
        files. This will generate data attributes associated with the tags
        defined in both the configuration yml files and the reference CONFIG_MAP
        and LAYER_NAME_MAP dictionaries.
        """

        ### Check whether data files defined from the configuration file exists
        # Generate flat JSON table comprising all configuration parameter names
        flat_configuration_table = pd.json_normalize(self.config)

        # Parse the flattened configuration table to identify data file names and paths
        parsed_filenames = flat_configuration_table.filter(regex="filename").values.flatten()

        # Append the root directory to the parsed data file names
        data_file_paths = [os.path.join(self.config['data_root_dir'], filename) for filename in parsed_filenames]

        # Evaluate whether either file is missing
        data_existence = [os.path.exists(file_path) for file_path in data_file_paths]

        # Assign the existence status to each configuration file for error evaluation
        data_status = dict(zip(data_file_paths, data_existence))

        # Error evaluation and print message (if applicable)
        if not all(data_status.values()):
            missing_data = [data_file for data_file, exists in data_status.items() if not exists]
            raise FileNotFoundError(f"The following data files do not exist: {missing_data}")
        
        ### Data validation and import
        # Iterate through each file defined/detected in 'data_file_paths'
        for file in data_file_paths:

            # Parse data layers, associated configuration key tag, and the Excel sheet name associated
            # with the data
            data_layers, reference_path, sheet_name = self.search_nested_dict(self.config, file)

            # Reference LAYER_NAME_MAP to update 'self.config' -- this is necessary for both mapping
            # the data attribute structures and reorganizing the dictionary organization within each
            # class attribute later on
            # Pull LAYER_NAME_MAP reference information
            layer_config = LAYER_NAME_MAP.get(data_layers[0])

            # Determine how deep 'data_layers' is represented within the dictionary
            # This is required for mapping dataframes with an appropriately named
            # data layer name appended with '_df' (e.g. 'length' -> 'length_df')
            data_name_bool = set(data_layers).intersection(layer_config['data'])
            data_layer_index = data_layers.index(''.join(data_name_bool))

            #pull_nested_dict(CONFIG_MAP, data_layers)
            # Retrieve/access CONFIG_MAP values defining the expected column names and data types
            validation_settings = pull_nested_dict(CONFIG_MAP, data_layers[:data_layer_index+1])

            ## ??? TODO: Remove Try-Exception ? Or may need more informative error message construction
            # Validate column names
            try:
                # Open connection with the workbook and specific sheet 
                # This is useful for not calling the workbook into memory and allows for parsing
                # only the necessary rows/column names
                workbook = load_workbook(file, read_only=True)
                sheet = workbook[sheet_name]

                # Validate that the expected columns are contained within the parsed 
                # column names of the workbook   
                if reference_path == "kriging/vario_krig_para":
                    data_columns = [list(row) for row in zip(*sheet.iter_rows(values_only=True))][0]
                else:   
                    data_columns = {col.value for col in sheet[1]}

                # Close connection to the work book
                workbook.close()

                # Error evaluation and print message (if applicable)
                if not set(validation_settings.keys()).issubset(set(data_columns)):
                    missing_columns = set(validation_settings.keys()) - set(data_columns)
                    raise ValueError(f"Missing columns in the Excel file: {missing_columns}")
                
            except Exception as e:
                print(f"Error reading file '{file}': {e}")
                continue

            # Based on the configuration settings, read the Excel files into memory. A format
            # exception is made for 'kriging.vario_krig_para' since it requires additional
            # data wrangling (i.e. transposing) to resemble the same dataframe format applied
            # to all other data attributes.
            if reference_path == "kriging/vario_krig_para":
                    
                # Read Excel file into memory and then transpose
                df_initial = pd.read_excel(file, header=None).T

                # Take the values from the first row and redfine them as the column headers
                df_initial.columns = df_initial.iloc[0]
                df_initial = df_initial.drop(0)

                # Slice only the columns that are relevant to the EchoPro module functionality
                valid_columns = list(set(validation_settings.keys()).intersection(set(df_initial.columns)))
                df_filtered = df_initial[valid_columns]

                # Ensure the order of columns in df_filtered matches df_initial
                df_filtered = df_filtered[df_initial.columns]

                # Apply data types from validation_settings to the filtered DataFrame
                df = df_filtered.apply(lambda col: col.astype(validation_settings.get(col.name, type(df_filtered.iloc[0][col.name]))))

            else:

                # Read Excel file into memory -- this only reads in the required columns
                df = pd.read_excel(file, sheet_name=sheet_name, usecols=validation_settings.keys())

                # Apply data types from validation_settings to the filtered DataFrame
                df = df.apply(lambda col: col.astype(validation_settings.get(col.name, type(col[0]))))

                # This is a goofy way of accounting for the 'CAN' and 'US' format -- if these are present,
                # they are added as an 'index' column that can be referenced later on.
                ## ??? TODO : This janky boolean chunk can be replaced with something more efficient
                # such as adding a key to the original configuration files that explicitly state datasets
                # from multiple sources/regions/etc. are present, or automatically merge layers with shared 
                # data tree structures in a later function/step
                if len(data_layers) > 2:
                    df['index'] = data_layers[2]

            ### CONFIGURATION : Map the imported data and save within the 'config' attribute
            # Initialize temporary attribute object
            ## TODO : This can probably be moved to the "IMPORT" section of this function below
            internal_object = self

            # Create a shallow copy of the data layers
            ## ??? TODO : This seems unnecessary, but some issues seemed to arise when this line is excluded.
            # This is probably due to some issue in the local environment where tests were evaluated. 
            dict_tree_names = data_layers.copy()

            # From LAYER_NAME_MAP, rename the original top layer (e.g. 'stratification' -> 'spatial')
            dict_tree_names[0] = layer_config['name']

            # From LAYER_NAME_MAP, add the 'superlayer', if defined, that adjusts the intended dictionary
            # nested tree structure (e.g. 'kriging/vars' -> 'statistics/kriging/vars')
            updated_layers = layer_config['superlayer'] + dict_tree_names

            # Determine how deep 'data_layers' is represented within the dictionary
            # This is required for mapping dataframes with an appropriately named
            # data layer name appended with '_df' (e.g. 'length' -> 'length_df')
            updated_name_bool = set(updated_layers).intersection(layer_config['data'])
            updated_layer_index = updated_layers.index(''.join(updated_name_bool))

            # Append '_df' to the correct data layer name. This will change the name located within the 
            # configuration to provide context of the datatype (in this case, as dataframe).
            updated_layers[updated_layer_index] = ''.join(updated_name_bool)+'_df'

            # Push the new nested tree data path to the config under the key: 'dict_tree'
            push_nested_dict(self.config, data_layers + ['dict_tree'], updated_layers)
            ### IMPORT : Assign the imported data to their intended data attributes that have now been 
            # explicitly defined from the configuration files, validated via 'CONFIG_MAP', reorganized and 
            # saved within 'self.config' via 'LAYER_NAME_MAP'

            # Index the data tree layers: (surface_layer, data_layer] -- so this should exclude the 
            # surface layer and exclude any detected layers beyond those appended with '_df'
            nested_layers = updated_layers[1:updated_layer_index+1]

            # Evaluate whether the surface layer is present -- if it is, then this means the data 
            # attribute already exists.
            if not hasattr(internal_object, updated_layers[0]):
                # Initialize empty dictionary
                d = {}

                # Push the formated nested dictionary to the temporary dictionary -- this will now
                # contain the data dataframe at the appropriate nested layer/level
                push_nested_dict(d, nested_layers, df) 

                # Initialize the attribute and set the value using the temporary dictionary
                setattr(internal_object, updated_layers[0], d)

            # If it is present, then the next step is to enter the data attribute itself to begin
            # amending the dictionary entries
            else:
                # Enter attribute
                internal_object = getattr(internal_object, updated_layers[0])

                # Does the labeled dataframe (e.g. 'length_df') already exist ? If not, then the
                # full data branch is initialized in case more values will be appended later on
                if pull_nested_dict(internal_object, nested_layers) is None:                    
                    # Push the formated nested dictionary to the temporary dictionary -- this will now
                    # contain the data dataframe at the appropriate nested layer/level
                    push_nested_dict(internal_object, nested_layers, df)

                # If it is present, then the 'index' column (e.g. 'US' vs 'CAN') will be required. Otherwise,
                # the pre-existing and proposed/to-be-added dataframes are simply concatenated.    
                else:
                    # Pull the dataframe located within the defined nested dictionary layers/levels
                    pre_df = pull_nested_dict(internal_object, nested_layers)

                    # Push the concatenated dataframe to the appropriate nested layer/level 
                    push_nested_dict(internal_object, nested_layers, pd.concat([pre_df , df])) 
        
        ### Merge haul numbers and regional indices across biological variables
        # Also add strata values/indices here alongside transect numbers 
        ### !!! TODO: rename 'index' column name to 'country' or 'region'
        self.biology['specimen_df'] = self.biology['haul_to_transect_df'].merge(self.biology['specimen_df'], on=['haul_num', 'index'])
        self.biology['length_df'] = self.biology['haul_to_transect_df'].merge(self.biology['length_df'], on=['haul_num', 'index'])
        self.biology['catch_df'] = self.biology['haul_to_transect_df'].merge(self.biology['catch_df'], on=['haul_num', 'index'])

    def biometric_distributions( self ):
        """
        Expand bin parameters into actual bins for length and age distributions
        """

        # Pull the relevant age and length bins and output a dictionary
        biometrics = {
            'length_bins': np.linspace( self.config['bio_hake_len_bin'][0] ,
                                        self.config['bio_hake_len_bin'][1] ,
                                        self.config['bio_hake_len_bin'][2] ,
                                        dtype = np.int64 ) ,
            'age_bins': np.linspace( self.config['bio_hake_age_bin'][0] ,
                                     self.config['bio_hake_age_bin'][1] , 
                                     self.config['bio_hake_age_bin'][2] ) ,
        }

        # Update the configuration so these variables are mappable
        self.config['biometrics'] = {
            'parameters': {
                'bio_hake_len_bin': self.config['bio_hake_len_bin'] ,
                'bio_hake_age_bin': self.config['bio_hake_age_bin']
            } ,
            'dict_tree': ['biology', 'distributions', ['age_bins', 'length_bins']]
        }
        del self.config['bio_hake_len_bin'], self.config['bio_hake_age_bin']

        # Push to biology attribute 
        self.biology['distributions'] = biometrics
    
    @staticmethod
    def add_to_tree( current_layer: dict , 
                    data_layers: np.ndarray , 
                    value: list ):
        """
        Map out the data path/structure of a specific branch within a nested
        data tree/dictionary.
        """

        # Iterate through the next branch that represent the 
        # nested data tree structure of each data attribute
        for i, layer in enumerate(data_layers):
            if i < len(data_layers) - 1:
                if layer not in current_layer:
                    current_layer[layer] = {}
                current_layer = current_layer[layer]
            else:
                if i == len(data_layers) - 1:
                    current_layer.setdefault(layer, []).append(value)
                else:
                    current_layer = current_layer.setdefault(layer, [])

    def populate_tree( self ):
        """
        Construct and populate the data structure tree and append it to the metadata attribute.
        """

        # Initialize the 'self.meta' attribute and the dictionary 'tree'.
        self.meta = {'tree': {}}

        # 'Normalize' the dictionary into a dataframe
        flat_configuration_table = pd.json_normalize(self.config)

        # Parse only the configuration values labeled 'dict_tree' -- then flatten
        tree_map = flat_configuration_table.filter(regex="dict_tree").values.flatten()

        # Iterate through all of the possible values in 'tree_map' to iteratively construct
        # a tree-like map of the various layers/levels/nodes contained within each of the 
        # class data attributes. This enables everything to be viewed in the console via the
        # 'summary' property function defined below.
        for data_layers in tree_map:
            self.add_to_tree(self.meta['tree'], data_layers[:-1], data_layers[-1])

    # Note : This function isn't necessary, but it is certainly helpful for debugging code along the way
    # So it may be worth removing in the future, or may be a helpful QOL utility for users. ¯\_(ツ)_/¯ 
    @property
    def summary( self ):
        """
        This provides a 'summary' property that can be used to quickly reference how the
        data attributes (and their respective nested trees) are organized within the Survey
        class object.
        """
        return pprint.pprint(self.meta.get('tree',{}))
