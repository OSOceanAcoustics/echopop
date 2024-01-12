from pathlib import Path
from openpyxl import load_workbook

def validate_data_columns( file_name: Path ,
                           sheet_name: str ,
                           config_map: list ,
                           validation_settings: dict ):
    """
    Opens a virtual instance of each .xlsx file to validate the presence 
    of require data column/variable names

    Parameters
    ----------
    file_name: Path
        File path of data
    sheet_name: str
        Name of Excel sheet containing data
    config_map: list
        A list parsed from the file name that indicates how data attributes
        within `self` are organized
    validation_settings: dict
        The subset CONFIG_MAP settings that contain the target column names
    """
    
    # Open connection with the workbook and specific sheet 
    # This is useful for not calling the workbook into memory and allows for parsing
    # only the necessary rows/column names 
    try:
        workbook = load_workbook(file_name, read_only=True)
        sheet = workbook[sheet_name]
        # Validate that the expected columns are contained within the parsed 
        # column names of the workbook   
        if 'vario_krig_para' in config_map:
            data_columns = [list(row) for row in zip(*sheet.iter_rows(values_only=True))][0]
        else:   
            data_columns = {col.value for col in sheet[1]}

        # Close connection to the work book
        workbook.close()

        # Error evaluation and print message (if applicable)
        if not set(validation_settings.keys()).issubset(set(data_columns)):
            missing_columns = set(validation_settings.keys()) - set(data_columns)
            raise ValueError(f"Missing columns in the Excel file: {missing_columns}")

    except Exception as e:
        print(f"Error reading file '{str(file_name)}': {e}")